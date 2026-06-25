"""
BRSR PDF & Excel Report Generator.
SEBI-mandated Business Responsibility and Sustainability Reporting (BRSR) —
Essential indicators only, Indian Financial Year (Apr–Mar).

Called from main.py /api/brsr/reports/generate.
PDF reuses the layout machinery from pdf_report.py.
"""

import io
import os
import json
import numpy as np

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate,
    Table, TableStyle, Paragraph, Spacer, PageBreak,
)

# Reuse layout constants + helpers from GRI PDF module
from pdf_report import (
    PAGE_W, PAGE_H, MARGIN_LR, MARGIN_TOP, MARGIN_BOT, AVAIL_W,
    WHITE, BORDER, HDR_TEXT, TOTAL_BG, NOTE_COL,
    _build_doc, _on_page_factory,
    _sec_hdr, _sub, _gap, _notes, _BASE_TS, _std_table, ST,
    _n, _filter_plant, _rate,
)

# ── BRSR AMBER PALETTE ─────────────────────────────────────────────────────────
AMBER  = colors.HexColor('#f59e0b')
AMBER2 = colors.HexColor('#d97706')

# ── CONFIG ────────────────────────────────────────────────────────────────────
_CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'data', 'brsr_config.json')

def _load_config():
    try:
        with open(_CONFIG_PATH, encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

# ── SECTION A — GENERAL DISCLOSURES ──────────────────────────────────────────

def _section_a(cfg, fy):
    org = cfg.get('organization', {})
    name    = org.get('name',             '[Not configured]')
    cin     = org.get('cin',              '[Not configured]')
    address = org.get('registered_address', '[Not configured]')
    contact = org.get('contact_person',   '[Not configured]')
    email   = org.get('contact_email',    '[Not configured]')
    website = org.get('website',          '[Not configured]')

    rows = [
        ['Field', 'Value'],
        ['Legal Name',              name],
        ['CIN',                     cin],
        ['Registered Address',      address],
        ['Reporting Period',        fy or 'Full Dataset'],
        ['SEBI Category',           cfg.get('sebi_category', 'Listed')],
        ['Website',                 website],
        ['BRSR Contact',            contact],
        ['Contact Email',           email],
    ]
    col_w = [AVAIL_W * 0.35, AVAIL_W * 0.65]
    t = Table([[Paragraph(c, ST['th_l'] if i == 0 else ST['td_l'])
                for i, c in enumerate(r)] for r in rows], colWidths=col_w)
    ts = list(_BASE_TS) + [
        ('BACKGROUND', (0, 0), (-1, 0), AMBER),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#fef3c7')),
        ('FONTNAME',   (0, 1), (0, -1), 'Helvetica-Bold'),
    ]
    t.setStyle(TableStyle(ts))
    return [_sec_hdr('SECTION A — General Disclosures'), _gap(), t, _gap()]


# ── SECTION B — MANAGEMENT & PROCESS DISCLOSURES ─────────────────────────────

def _section_b(cfg):
    principles = cfg.get('principles', {})
    rows = [['Principle', 'Title', 'Policy Status', 'Committee']]
    p_labels = {
        'p1': 'P1 — Ethics & Transparency',
        'p2': 'P2 — Sustainable Products',
        'p3': 'P3 — Employee Well-Being',
        'p4': 'P4 — Stakeholder Responsiveness',
        'p5': 'P5 — Human Rights',
        'p6': 'P6 — Environment',
        'p7': 'P7 — Policy Advocacy',
        'p8': 'P8 — Inclusive Growth / CSR',
        'p9': 'P9 — Consumer Responsibility',
    }
    for key in ['p1', 'p2', 'p3', 'p4', 'p5', 'p6', 'p7', 'p8', 'p9']:
        p = principles.get(key, {})
        status  = p.get('policy_status', 'not-adopted')
        status_label = {'adopted': 'Adopted ✓', 'in-progress': 'In Progress', 'not-adopted': 'Not Adopted'}.get(status, status)
        committee = p.get('committee') or '—'
        rows.append([
            p_labels.get(key, key.upper()),
            p.get('policy_title', ''),
            status_label,
            committee,
        ])
    col_w = [AVAIL_W * 0.18, AVAIL_W * 0.35, AVAIL_W * 0.17, AVAIL_W * 0.30]
    def cell(text, is_hdr=False, align=TA_LEFT):
        sty = ST['th_l'] if is_hdr else ST['td_l']
        return Paragraph(str(text), sty)
    tbl_rows = [[cell(c, is_hdr=(i == 0)) for c in r] for i, r in enumerate(rows)]
    t = Table(tbl_rows, colWidths=col_w)
    ts = list(_BASE_TS) + [
        ('BACKGROUND', (0, 0), (-1, 0), AMBER),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]
    t.setStyle(TableStyle(ts))
    return [_sec_hdr('SECTION B — Management & Process Disclosures'), _gap(), t, _gap()]


# ── SECTION C — PRINCIPLE-WISE DISCLOSURES ────────────────────────────────────

def _yearly_sums(df, fy_years, col):
    return {y: float(df[df['ReportingYear'] == y][col].sum()) for y in fy_years}

def _build_p6_energy(df_e, fy_years):
    rows = [
        {'label': 'Total Energy Consumed',           'unit': 'GJ',  'values': _yearly_sums(df_e, fy_years, 'TotalEnergyConsumedGJ')},
        {'label': 'Electricity from Renewables',     'unit': 'GJ',  'values': _yearly_sums(df_e, fy_years, 'ElectricityRenewableGJ')},
        {'label': 'Electricity from Non-Renewables', 'unit': 'GJ',  'values': _yearly_sums(df_e, fy_years, 'ElectricityNonRenewableGJ')},
    ]
    # Add renewable % row
    ren_vals = {}
    for y in fy_years:
        ren = float(df_e[df_e['ReportingYear'] == y]['ElectricityRenewableGJ'].sum())
        non = float(df_e[df_e['ReportingYear'] == y]['ElectricityNonRenewableGJ'].sum())
        ren_vals[y] = round(ren / (ren + non) * 100, 1) if (ren + non) > 0 else 0.0
    rows.append({'label': 'Renewable Share', 'unit': '%', 'values': ren_vals, 'dp': 1})
    return [
        _sub('Principle 6 — Energy (B1/B2): GJ / % renewable'),
        _std_table(rows, fy_years, row_hdr='Energy Metric', w_label=220, default_dp=0),
        _gap(4),
    ]


def _build_p6_water(df_w, fy_years):
    rows = [
        {'label': 'Total Water Withdrawn',                  'unit': 'KL', 'values': _yearly_sums(df_w, fy_years, 'TotalWaterWithdrawn')},
        {'label': 'Water Consumed',                         'unit': 'KL', 'values': _yearly_sums(df_w, fy_years, 'WaterConsumed')},
        {'label': 'Water Withdrawn — Stress Areas',         'unit': 'KL', 'values': _yearly_sums(df_w, fy_years, 'TotalWaterWithdrawnStressArea')},
    ]
    return [
        _sub('Principle 6 — Water (C1/C2): KL'),
        _std_table(rows, fy_years, row_hdr='Water Metric', w_label=230, default_dp=0),
        _gap(4),
    ]


def _build_p6_ghg(df_g, fy_years):
    rows = [
        {'label': 'Scope 1 GHG Emissions',             'unit': 'tCO₂e', 'values': _yearly_sums(df_g, fy_years, 'Scope1TotaltCO2e'), 'dp': 1},
        {'label': 'Scope 2 GHG Emissions (location)',  'unit': 'tCO₂e', 'values': _yearly_sums(df_g, fy_years, 'Scope2LocationBasedtCO2e'), 'dp': 1},
    ]
    return [
        _sub('Principle 6 — GHG Emissions (D1/D2): tCO₂e'),
        _std_table(rows, fy_years, row_hdr='Emissions Category', w_label=230, default_dp=1),
        _gap(4),
    ]


def _build_p6_waste(df_ws, fy_years):
    # GRI306 uses long format: HazardousFlag column + ValueNumber column
    if 'HazardousFlag' in df_ws.columns and 'ValueNumber' in df_ws.columns:
        haz_df    = df_ws[df_ws['HazardousFlag'] == 'Hazardous']
        nonhaz_df = df_ws[df_ws['HazardousFlag'] == 'Non-hazardous']
        haz_vals    = {y: float(haz_df[haz_df['ReportingYear'] == y]['ValueNumber'].sum())    for y in fy_years}
        nonhaz_vals = {y: float(nonhaz_df[nonhaz_df['ReportingYear'] == y]['ValueNumber'].sum()) for y in fy_years}
        rows = [
            {'label': 'Hazardous Waste Generated',     'unit': 't', 'values': haz_vals,    'dp': 2},
            {'label': 'Non-Hazardous Waste Generated', 'unit': 't', 'values': nonhaz_vals, 'dp': 2},
        ]
    else:
        # Fallback for wide-format datasets
        rows = []
        for label, col in [('Hazardous Waste Generated', 'TotalHazardousWasteGenerated'),
                            ('Non-Hazardous Waste Generated', 'TotalNonHazardousWasteGenerated')]:
            if col in df_ws.columns:
                rows.append({'label': label, 'unit': 't', 'values': _yearly_sums(df_ws, fy_years, col), 'dp': 2})
    if not rows:
        return [_sub('Principle 6 — Waste (E1/E2): Data columns not found in dataset.'), _gap(4)]
    return [
        _sub('Principle 6 — Waste (E1/E2): metric tonnes'),
        _std_table(rows, fy_years, row_hdr='Waste Category', w_label=230, default_dp=2),
        _gap(4),
    ]


def _build_p3_safety(df_s, fy_years):
    # GRI403 column aliases — the dataset uses HoursWorked / RecordableInjuries / FatalitiesInjury
    hours_col = 'HoursWorked' if 'HoursWorked' in df_s.columns else 'TotalHoursWorked'
    rec_col   = 'RecordableInjuries' if 'RecordableInjuries' in df_s.columns else 'TotalRecordableIncidents'
    fat_col   = 'FatalitiesInjury' if 'FatalitiesInjury' in df_s.columns else 'Fatalities'

    def rate_by_year(df, y, num_col, den_col):
        sub = df[df['FiscalReportingPeriod'] == y]
        num = float(sub[num_col].sum()) if num_col in sub.columns else 0.0
        den = float(sub[den_col].sum()) if den_col in sub.columns else 0.0
        return _rate(num, den)

    ltifr_vals = {y: rate_by_year(df_s, y, 'LostTimeInjuries', hours_col) for y in fy_years}
    trir_vals  = {y: rate_by_year(df_s, y, rec_col, hours_col) for y in fy_years}
    fat_vals   = {y: int(df_s[df_s['FiscalReportingPeriod'] == y][fat_col].sum()) if fat_col in df_s.columns else 0 for y in fy_years}

    rows = [
        {'label': 'LTIFR (per 200,000 hrs)',  'unit': 'rate', 'values': ltifr_vals, 'dp': 2},
        {'label': 'TRIR (per 200,000 hrs)',   'unit': 'rate', 'values': trir_vals,  'dp': 2},
        {'label': 'Fatalities',               'unit': 'cases', 'values': fat_vals,  'dp': 0},
    ]
    return [
        _sub('Principle 3 — Occupational Health & Safety (GRI 403-9): Rates per 200,000 hours'),
        _std_table(rows, fy_years, row_hdr='Safety Metric', w_label=230, default_dp=2),
        _gap(4),
    ]


def _placeholder_section(principle_key, title, description):
    msg = f'{principle_key} — {title}: Data collection pending. {description}'
    t = Table([[Paragraph(msg, ST['nt'])]], colWidths=[AVAIL_W])
    t.setStyle(TableStyle(list(_BASE_TS) + [
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fef3c7')),
    ]))
    return [t, _gap(4)]


def _build_p3_workforce(df_wf, all_fys):
    """P3 Workforce section — headcount by gender, female %, DA count, wage parity."""
    if df_wf is None or len(df_wf) == 0:
        return _placeholder_section('P3', 'Workforce & Training',
            'BRSR_Workforce_Dataset not yet collected.')

    rows = [['Metric', 'Unit'] + all_fys]
    metrics = [
        ('Permanent Male', 'PermanentMale', 'count'),
        ('Permanent Female', 'PermanentFemale', 'count'),
        ('Permanent Other', 'PermanentOther', 'count'),
        ('Contractual Male', 'ContractualMale', 'count'),
        ('Contractual Female', 'ContractualFemale', 'count'),
        ('Differently-Abled (Permanent)', 'DifferentlyAbledPermanent', 'count'),
        ('Differently-Abled (Contractual)', 'DifferentlyAbledContractual', 'count'),
        ('Avg Wage — Permanent Male', 'AvgWagePermanentMaleINR', '₹/yr'),
        ('Avg Wage — Permanent Female', 'AvgWagePermanentFemaleINR', '₹/yr'),
    ]
    for label, col, unit in metrics:
        if col not in df_wf.columns:
            continue
        vals = []
        for fy in all_fys:
            sl = df_wf[df_wf['FY'] == fy]
            v = round(float(sl[col].sum()), 0) if len(sl) > 0 else '—'
            vals.append(v)
        rows.append([label, unit] + vals)

    # Female %
    fem_pct_row = ['Female % (all workers)', '%']
    for fy in all_fys:
        sl = df_wf[df_wf['FY'] == fy]
        if len(sl) == 0:
            fem_pct_row.append('—')
            continue
        total_f = float(sl['PermanentFemale'].sum() + sl['ContractualFemale'].sum())
        total   = float(sl[['PermanentMale','PermanentFemale','PermanentOther',
                              'ContractualMale','ContractualFemale','ContractualOther']].sum().sum())
        fem_pct_row.append(round(total_f / total * 100, 1) if total > 0 else '—')
    rows.insert(4, fem_pct_row)

    col_w = [200] + [65] + [max(45, int((AVAIL_W - 265) / max(len(all_fys), 1)))] * len(all_fys)
    def cell(v, hdr=False):
        sty = ST['th_l'] if hdr else ST['td_l']
        return Paragraph(str(v), sty)

    tbl_rows = [[cell(c, hdr=(i == 0)) for c in r] for i, r in enumerate(rows)]
    t = Table(tbl_rows, colWidths=col_w)
    ts = list(_BASE_TS) + [
        ('BACKGROUND', (0, 0), (-1, 0), AMBER),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]
    t.setStyle(TableStyle(ts))
    return [_sub('Principle 3 — Workforce (P3-Essential): Headcount & Gender Parity'), t, _gap(4)]


def _build_p3_training(df_tr, all_fys):
    """P3 Training section — avg hours, coverage, skill upgrade %."""
    if df_tr is None or len(df_tr) == 0:
        return _placeholder_section('P3', 'Training',
            'BRSR_Training_Dataset not yet collected.')

    metrics = [
        ('Avg Training Hours / Employee', 'AvgTrainingHrsAllEmployees', 'hrs', 1),
        ('Avg Training Hours (Male)', 'AvgTrainingHrsPerEmployeeMale', 'hrs', 1),
        ('Avg Training Hours (Female)', 'AvgTrainingHrsPerEmployeeFemale', 'hrs', 1),
        ('Training Coverage', 'TrainingCoveragePct', '%', 1),
        ('Skill Upgrade Coverage', 'SkillUpgradePct', '%', 1),
        ('Performance Review Coverage', 'PerformanceReviewCoveragePct', '%', 1),
    ]

    rows = [['Metric', 'Unit'] + all_fys]
    for label, col, unit, dp in metrics:
        if col not in df_tr.columns:
            continue
        vals = []
        for fy in all_fys:
            sl = df_tr[df_tr['FY'] == fy]
            v = round(float(sl[col].mean()), dp) if len(sl) > 0 else '—'
            vals.append(v)
        rows.append([label, unit] + vals)

    col_w = [200] + [65] + [max(45, int((AVAIL_W - 265) / max(len(all_fys), 1)))] * len(all_fys)
    def cell(v, hdr=False):
        return Paragraph(str(v), ST['th_l'] if hdr else ST['td_l'])
    tbl_rows = [[cell(c, hdr=(i == 0)) for c in r] for i, r in enumerate(rows)]
    t = Table(tbl_rows, colWidths=col_w)
    ts = list(_BASE_TS) + [
        ('BACKGROUND', (0, 0), (-1, 0), AMBER),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]
    t.setStyle(TableStyle(ts))
    return [_sub('Principle 3 — Training & Development (P3-Essential): Hours per Employee & Coverage'), t, _gap(4)]


def _build_p8_csr(df_csr, all_fys):
    """P8 CSR section — obligation vs spend, category breakdown."""
    if df_csr is None or len(df_csr) == 0:
        return _placeholder_section('P8', 'CSR',
            'BRSR_CSR_Dataset not yet collected.')

    # Summary table
    rows = [['Metric', 'Unit'] + all_fys]
    for label, key in [('CSR Obligation', 'ObligationCrore'), ('CSR Expenditure', 'TotalSpentCrore'), ('Unspent Amount', 'UnspentCrore')]:
        vals = []
        for fy in all_fys:
            sl = df_csr[df_csr['FY'] == fy]
            v = round(float(sl[key].iloc[0]), 2) if len(sl) > 0 else '—'
            vals.append(v)
        rows.append([label, '₹ Cr'] + vals)

    # % utilization
    util_row = ['Utilization %', '%']
    for fy in all_fys:
        sl = df_csr[df_csr['FY'] == fy]
        if len(sl) == 0:
            util_row.append('—')
        else:
            ob = float(sl['ObligationCrore'].iloc[0])
            sp = float(sl['TotalSpentCrore'].iloc[0])
            util_row.append(round(sp / ob * 100, 1) if ob > 0 else '—')
    rows.append(util_row)

    # Total beneficiaries
    bene_row = ['Total Beneficiaries', 'ppl']
    for fy in all_fys:
        sl = df_csr[df_csr['FY'] == fy]
        bene_row.append(int(sl['BeneficiaryCount'].sum()) if len(sl) > 0 else '—')
    rows.append(bene_row)

    col_w = [200] + [65] + [max(45, int((AVAIL_W - 265) / max(len(all_fys), 1)))] * len(all_fys)
    def cell(v, hdr=False):
        return Paragraph(str(v), ST['th_l'] if hdr else ST['td_l'])
    tbl_rows = [[cell(c, hdr=(i == 0)) for c in r] for i, r in enumerate(rows)]
    t = Table(tbl_rows, colWidths=col_w)
    ts = list(_BASE_TS) + [
        ('BACKGROUND', (0, 0), (-1, 0), AMBER),
        ('TEXTCOLOR',  (0, 0), (-1, 0), WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]
    t.setStyle(TableStyle(ts))
    return [_sub('Principle 8 — CSR (P8-Essential): Obligation, Expenditure & Beneficiaries'), t, _gap(4)]


# ── PUBLIC ENTRY POINTS ───────────────────────────────────────────────────────

def generate_brsr_pdf(fy, plant, filter_by_fy_fn, filter_annual_by_fy_fn,
                      load_energy_fn, load_ghg_fn, load_water_fn,
                      load_waste_fn, load_safety_fn,
                      load_workforce_fn=None, load_training_fn=None, load_csr_fn=None,
                      principles=None):
    """Generate a BRSR Essential-indicators PDF. Returns BytesIO.

    principles: optional list of BRSR_REPORT_TEMPLATES ids (main.py) to
    restrict Section C to - None/empty means every principle (default)."""
    include = lambda key: not principles or key in principles
    cfg = _load_config()
    buf = io.BytesIO()
    fy_label = fy or 'Full Dataset'
    org_name = cfg.get('organization', {}).get('name', 'Organization')
    org_label = f'{org_name}  ·  BRSR Disclosure Index {fy_label}'
    doc = _build_doc(buf, org_label=org_label)

    # Load + filter datasets to the Indian FY window
    df_e  = filter_by_fy_fn(_filter_plant(load_energy_fn(), plant), fy)
    df_g  = filter_by_fy_fn(_filter_plant(load_ghg_fn(),    plant), fy)
    df_w  = filter_by_fy_fn(_filter_plant(load_water_fn(),  plant), fy)
    df_ws = filter_by_fy_fn(_filter_plant(load_waste_fn(),  plant), fy)
    df_s  = filter_by_fy_fn(_filter_plant(load_safety_fn(), plant), fy)

    # Annual datasets (FY-keyed, plant-filtered)
    ALL_FYS = ['FY2019-20', 'FY2020-21', 'FY2021-22', 'FY2022-23', 'FY2023-24', 'FY2024-25']
    df_wf = None
    df_tr = None
    df_csr = None
    if load_workforce_fn:
        _wf = _filter_plant(load_workforce_fn(), plant)
        df_wf = filter_annual_by_fy_fn(_wf, fy) if fy else _wf
    if load_training_fn:
        _tr = _filter_plant(load_training_fn(), plant)
        df_tr = filter_annual_by_fy_fn(_tr, fy) if fy else _tr
    if load_csr_fn:
        _csr = load_csr_fn()
        df_csr = filter_annual_by_fy_fn(_csr, fy) if fy else _csr

    # FY period labels for annual tables — use all FYs if no specific FY selected
    report_fys = [fy] if fy else ALL_FYS

    # Calendar years present in this FY window (for GRI monthly table headers)
    fy_years = sorted(df_e['ReportingYear'].unique().tolist()) if 'ReportingYear' in df_e.columns else []
    safety_years = sorted(df_s['FiscalReportingPeriod'].unique().tolist()) if 'FiscalReportingPeriod' in df_s.columns else []

    story = []

    # ── Cover ──────────────────────────────────────────────────────────────────
    story += [
        _gap(40),
        Table([[Paragraph(org_name, ParagraphStyle('cov_org', fontName='Helvetica-Bold', fontSize=16, textColor=AMBER))]], colWidths=[AVAIL_W]),
        _gap(8),
        Table([[Paragraph('BRSR Disclosure — Essential Indicators', ParagraphStyle('cov_h', fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor('#1e293b')))]], colWidths=[AVAIL_W]),
        _gap(4),
        Table([[Paragraph(f'Reporting Period: {fy_label}', ParagraphStyle('cov_s', fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#475569')))]], colWidths=[AVAIL_W]),
        _gap(4),
        Table([[Paragraph('Scope: Essential Indicators only (as per SEBI BRSR circular)', ParagraphStyle('cov_note', fontName='Helvetica-Oblique', fontSize=9, textColor=colors.HexColor('#64748b')))]], colWidths=[AVAIL_W]),
        PageBreak(),
    ]

    # ── Section A ──────────────────────────────────────────────────────────────
    story += _section_a(cfg, fy_label)
    story.append(PageBreak())

    # ── Section B ──────────────────────────────────────────────────────────────
    story += _section_b(cfg)
    story.append(PageBreak())

    # ── Section C ──────────────────────────────────────────────────────────────
    story.append(_sec_hdr('SECTION C — Principle-Wise Performance Disclosures'))
    story.append(_gap(8))

    # P6 Environment (live data)
    if include('brsr_p6'):
        if fy_years:
            story += _build_p6_energy(df_e, fy_years)
            story += _build_p6_water(df_w, fy_years)
            story += _build_p6_ghg(df_g, fy_years)
            story += _build_p6_waste(df_ws, fy_years)
        else:
            story += _placeholder_section('P6', 'Environment', 'No energy/GHG/water/waste records found for the selected FY window.')
        story.append(_gap(6))

    # P3 Safety (live from GRI403)
    if include('brsr_p3_safety'):
        if safety_years:
            story += _build_p3_safety(df_s, safety_years)
        else:
            story += _placeholder_section('P3', 'Safety', 'No OHS records found for the selected FY window.')
        story.append(_gap(6))

    # P3 Workforce & Training (live from BRSR datasets)
    if include('brsr_p3_workforce'):
        story += _build_p3_workforce(df_wf, report_fys)
        story.append(_gap(6))
        story += _build_p3_training(df_tr, report_fys)
        story.append(_gap(6))

    # P8 CSR (live from BRSR CSR dataset)
    if include('brsr_p8'):
        story += _build_p8_csr(df_csr, report_fys)
        story.append(_gap(6))

    if include('brsr_other'):
        story += _placeholder_section('P2/P4/P5/P7/P9', 'Remaining Principles',
            'Data collection templates for P2/P4/P5/P7/P9 not yet established.')

    story += _notes([
        'This report covers BRSR Essential indicators only (SEBI BRSR circular 2021).',
        'P6 quantitative data is sourced from GRI 302/303/305/306/403 datasets, aggregated over the Indian Financial Year window.',
        'P3 Workforce/Training and P8 CSR data use synthetic datasets generated for demonstration purposes.',
        'Values are unaudited; obtain third-party assurance before public disclosure.',
    ])

    doc.build(story)
    buf.seek(0)
    return buf


def generate_brsr_excel(fy, plant, filter_by_fy_fn, filter_annual_by_fy_fn,
                        load_energy_fn, load_ghg_fn, load_water_fn,
                        load_waste_fn, load_safety_fn,
                        load_workforce_fn=None, load_training_fn=None, load_csr_fn=None,
                        principles=None):
    """Generate a BRSR disclosure Excel workbook (SEBI-format). Returns BytesIO.

    principles: optional list of BRSR_REPORT_TEMPLATES ids (main.py) to
    restrict the principle-wise sheets to - None/empty means every principle
    (default). Section A/B sheets always render, same as the PDF's front matter."""
    include = lambda key: not principles or key in principles
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    cfg    = _load_config()
    fy_label = fy or 'Full Dataset'
    ALL_FYS = ['FY2019-20', 'FY2020-21', 'FY2021-22', 'FY2022-23', 'FY2023-24', 'FY2024-25']
    report_fys = [fy] if fy else ALL_FYS

    df_e  = filter_by_fy_fn(_filter_plant(load_energy_fn(), plant), fy)
    df_g  = filter_by_fy_fn(_filter_plant(load_ghg_fn(),    plant), fy)
    df_w  = filter_by_fy_fn(_filter_plant(load_water_fn(),  plant), fy)
    df_ws = filter_by_fy_fn(_filter_plant(load_waste_fn(),  plant), fy)
    df_s  = filter_by_fy_fn(_filter_plant(load_safety_fn(), plant), fy)

    fy_years     = sorted(df_e['ReportingYear'].unique().tolist()) if 'ReportingYear' in df_e.columns else []
    safety_years = sorted(df_s['FiscalReportingPeriod'].unique().tolist()) if 'FiscalReportingPeriod' in df_s.columns else []

    # Annual datasets
    df_wf = None
    df_tr = None
    df_csr = None
    if load_workforce_fn:
        _wf = _filter_plant(load_workforce_fn(), plant)
        df_wf = filter_annual_by_fy_fn(_wf, fy) if fy else _wf
    if load_training_fn:
        _tr = _filter_plant(load_training_fn(), plant)
        df_tr = filter_annual_by_fy_fn(_tr, fy) if fy else _tr
    if load_csr_fn:
        _csr = load_csr_fn()
        df_csr = filter_annual_by_fy_fn(_csr, fy) if fy else _csr

    wb = Workbook()
    wb.remove(wb.active)

    amber_fill = PatternFill(fill_type='solid', fgColor='F59E0B')
    hdr_font   = Font(bold=True, color='FFFFFF')
    thin       = Side(border_style='thin', color='CCCCCC')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_sheet(title, rows_of_rows):
        ws = wb.create_sheet(title=title[:31])
        for r_idx, row in enumerate(rows_of_rows, 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                if r_idx == 1:
                    cell.fill = amber_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal='center')
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)
        return ws

    # Sheet: Section A
    org = cfg.get('organization', {})
    add_sheet('Section A — General', [
        ['Field', 'Value'],
        ['Legal Name',      org.get('name', '')],
        ['CIN',             org.get('cin', '')],
        ['Address',         org.get('registered_address', '')],
        ['Reporting FY',    fy_label],
        ['SEBI Category',   cfg.get('sebi_category', 'Listed')],
        ['Website',         org.get('website', '')],
        ['Contact Person',  org.get('contact_person', '')],
        ['Contact Email',   org.get('contact_email', '')],
    ])

    # Sheet: Section B — Policy status
    b_rows = [['Principle', 'Title', 'Policy Status', 'Committee']]
    for key in ['p1', 'p2', 'p3', 'p4', 'p5', 'p6', 'p7', 'p8', 'p9']:
        p = cfg.get('principles', {}).get(key, {})
        b_rows.append([
            key.upper(),
            p.get('policy_title', ''),
            p.get('policy_status', 'not-adopted'),
            p.get('committee') or '',
        ])
    add_sheet('Section B — Policies', b_rows)

    # Sheet: P6 Energy
    if include('brsr_p6') and fy_years:
        e_rows = [['Metric', 'Unit'] + [str(y) for y in fy_years]]
        for label, col in [('Total Energy Consumed', 'TotalEnergyConsumedGJ'),
                           ('Electricity — Renewables', 'ElectricityRenewableGJ'),
                           ('Electricity — Non-Renewables', 'ElectricityNonRenewableGJ')]:
            vals = [round(float(df_e[df_e['ReportingYear'] == y][col].sum()), 1) for y in fy_years]
            e_rows.append([label, 'GJ'] + vals)
        add_sheet('P6 Energy', e_rows)

        # Sheet: P6 Water
        w_rows = [['Metric', 'Unit'] + [str(y) for y in fy_years]]
        for label, col in [('Total Water Withdrawn', 'TotalWaterWithdrawn'),
                           ('Water Consumed', 'WaterConsumed'),
                           ('Withdrawn — Stress Areas', 'TotalWaterWithdrawnStressArea')]:
            vals = [round(float(df_w[df_w['ReportingYear'] == y][col].sum()), 1) for y in fy_years]
            w_rows.append([label, 'KL'] + vals)
        add_sheet('P6 Water', w_rows)

        # Sheet: P6 GHG
        g_rows = [['Metric', 'Unit'] + [str(y) for y in fy_years]]
        for label, col in [('Scope 1 GHG', 'Scope1TotaltCO2e'),
                           ('Scope 2 GHG (location)', 'Scope2LocationBasedtCO2e')]:
            vals = [round(float(df_g[df_g['ReportingYear'] == y][col].sum()), 1) for y in fy_years]
            g_rows.append([label, 'tCO2e'] + vals)
        add_sheet('P6 GHG', g_rows)

        # Sheet: P6 Waste (long-format dataset uses HazardousFlag + ValueNumber)
        ws_rows = [['Metric', 'Unit'] + [str(y) for y in fy_years]]
        if 'HazardousFlag' in df_ws.columns and 'ValueNumber' in df_ws.columns:
            for label, flag in [('Hazardous Waste Generated', 'Hazardous'),
                                 ('Non-Hazardous Waste Generated', 'Non-hazardous')]:
                sub_df = df_ws[df_ws['HazardousFlag'] == flag]
                vals   = [round(float(sub_df[sub_df['ReportingYear'] == y]['ValueNumber'].sum()), 2) for y in fy_years]
                ws_rows.append([label, 't'] + vals)
        else:
            for label, col in [('Hazardous Waste Generated', 'TotalHazardousWasteGenerated'),
                               ('Non-Hazardous Waste Generated', 'TotalNonHazardousWasteGenerated')]:
                if col in df_ws.columns:
                    vals = [round(float(df_ws[df_ws['ReportingYear'] == y][col].sum()), 2) for y in fy_years]
                    ws_rows.append([label, 't'] + vals)
        add_sheet('P6 Waste', ws_rows)

    # Sheet: P3 Safety — use correct GRI403 column aliases
    if include('brsr_p3_safety') and safety_years:
        _h   = 'HoursWorked' if 'HoursWorked' in df_s.columns else 'TotalHoursWorked'
        _rec = 'RecordableInjuries' if 'RecordableInjuries' in df_s.columns else 'TotalRecordableIncidents'
        _fat = 'FatalitiesInjury' if 'FatalitiesInjury' in df_s.columns else 'Fatalities'

        def yr_rate(df, y, num, den):
            sub = df[df['FiscalReportingPeriod'] == y]
            n = float(sub[num].sum()) if num in sub.columns else 0.0
            d = float(sub[den].sum()) if den in sub.columns else 0.0
            return _rate(n, d)

        s_rows = [['Metric', 'Unit'] + [str(y) for y in safety_years]]
        s_rows.append(['LTIFR', 'per 200k hrs'] +
                      [yr_rate(df_s, y, 'LostTimeInjuries', _h) for y in safety_years])
        s_rows.append(['TRIR', 'per 200k hrs'] +
                      [yr_rate(df_s, y, _rec, _h) for y in safety_years])
        if _fat in df_s.columns:
            s_rows.append(['Fatalities', 'cases'] +
                          [int(df_s[df_s['FiscalReportingPeriod'] == y][_fat].sum()) for y in safety_years])
        add_sheet('P3 Safety', s_rows)

    # Sheet: P3 Workforce
    if include('brsr_p3_workforce') and df_wf is not None and len(df_wf) > 0:
        wf_cols = [
            ('Permanent Male', 'PermanentMale', 'count'),
            ('Permanent Female', 'PermanentFemale', 'count'),
            ('Permanent Other', 'PermanentOther', 'count'),
            ('Contractual Male', 'ContractualMale', 'count'),
            ('Contractual Female', 'ContractualFemale', 'count'),
            ('Differently-Abled Permanent', 'DifferentlyAbledPermanent', 'count'),
            ('Differently-Abled Contractual', 'DifferentlyAbledContractual', 'count'),
            ('Avg Wage Permanent Male (₹)', 'AvgWagePermanentMaleINR', '₹/yr'),
            ('Avg Wage Permanent Female (₹)', 'AvgWagePermanentFemaleINR', '₹/yr'),
        ]
        wf_rows = [['Metric', 'Unit'] + report_fys]
        for label, col, unit in wf_cols:
            if col not in df_wf.columns:
                continue
            vals = [round(float(df_wf[df_wf['FY'] == f][col].sum()), 0) if len(df_wf[df_wf['FY'] == f]) > 0 else '' for f in report_fys]
            wf_rows.append([label, unit] + vals)
        # Female %
        fem_row = ['Female % (all workers)', '%']
        for f in report_fys:
            sl = df_wf[df_wf['FY'] == f]
            if len(sl) == 0:
                fem_row.append('')
            else:
                tot_f = float(sl['PermanentFemale'].sum() + sl['ContractualFemale'].sum())
                tot   = float(sl[['PermanentMale','PermanentFemale','PermanentOther','ContractualMale','ContractualFemale','ContractualOther']].sum().sum())
                fem_row.append(round(tot_f / tot * 100, 1) if tot > 0 else '')
        wf_rows.insert(4, fem_row)
        add_sheet('P3 Workforce', wf_rows)

    # Sheet: P3 Training
    if include('brsr_p3_workforce') and df_tr is not None and len(df_tr) > 0:
        tr_cols = [
            ('Avg Training Hrs / Employee', 'AvgTrainingHrsAllEmployees', 'hrs'),
            ('Avg Hrs (Male)', 'AvgTrainingHrsPerEmployeeMale', 'hrs'),
            ('Avg Hrs (Female)', 'AvgTrainingHrsPerEmployeeFemale', 'hrs'),
            ('Training Coverage', 'TrainingCoveragePct', '%'),
            ('Skill Upgrade Coverage', 'SkillUpgradePct', '%'),
            ('Performance Review Coverage', 'PerformanceReviewCoveragePct', '%'),
            ('Training Spend / Employee (₹)', 'TrainingSpendPerEmployeeINR', '₹'),
        ]
        tr_rows = [['Metric', 'Unit'] + report_fys]
        for label, col, unit in tr_cols:
            if col not in df_tr.columns:
                continue
            vals = [round(float(df_tr[df_tr['FY'] == f][col].mean()), 1) if len(df_tr[df_tr['FY'] == f]) > 0 else '' for f in report_fys]
            tr_rows.append([label, unit] + vals)
        add_sheet('P3 Training', tr_rows)

    # Sheet: P8 CSR
    if include('brsr_p8') and df_csr is not None and len(df_csr) > 0:
        csr_rows = [['Metric', 'Unit'] + report_fys]
        for label, col in [('CSR Obligation (₹ Cr)', 'ObligationCrore'),
                           ('CSR Expenditure (₹ Cr)', 'TotalSpentCrore'),
                           ('Unspent Amount (₹ Cr)', 'UnspentCrore')]:
            vals = [round(float(df_csr[df_csr['FY'] == f][col].iloc[0]), 2) if len(df_csr[df_csr['FY'] == f]) > 0 else '' for f in report_fys]
            csr_rows.append([label, '₹ Cr'] + vals)
        # beneficiaries
        bene_row = ['Total Beneficiaries', 'ppl']
        for f in report_fys:
            sl = df_csr[df_csr['FY'] == f]
            bene_row.append(int(sl['BeneficiaryCount'].sum()) if len(sl) > 0 else '')
        csr_rows.append(bene_row)
        # Category breakdown for last FY
        last_fy_csr = df_csr[df_csr['FY'] == report_fys[-1]]
        if len(last_fy_csr) > 0:
            csr_rows.append([])
            csr_rows.append([f'Category Breakdown — {report_fys[-1]}', 'Spend (₹ Cr)', 'Beneficiaries'])
            for _, row in last_fy_csr.iterrows():
                csr_rows.append([row['ProjectCategory'], round(float(row['SpentCrore']), 2), int(row['BeneficiaryCount'])])
        add_sheet('P8 CSR', csr_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
