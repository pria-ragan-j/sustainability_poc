import time
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, Response
import pandas as pd
import numpy as np
import io
import os
import json
from typing import Optional
from pydantic import BaseModel
from dotenv import load_dotenv
import httpx
import uuid

import storage

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

app = FastAPI(title="ESG Dashboard API")
storage.init_db()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    # Browsers only expose a small safelisted set of response headers to
    # JS fetch() for cross-origin requests (frontend :3000 -> backend :8000) -
    # Content-Disposition is NOT in that safelist by default, so without this,
    # res.headers.get('Content-Disposition') in api/client.js's generateReport()
    # always returns null and silently falls back to the generic "esg_report"
    # filename, regardless of what the server actually sends.
    expose_headers=["Content-Disposition"],
)

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")

MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY", "")
MISTRAL_API_URL = "https://api.mistral.ai/v1/chat/completions"
MISTRAL_MODEL = os.getenv("MISTRAL_MODEL", "mistral-large-latest")
MISTRAL_TEMP = float(os.getenv("MISTRAL_TEMPERATURE", "0.3"))
MISTRAL_MAX_TOKENS = int(os.getenv("MISTRAL_MAX_TOKENS", "2048"))

RATE_HOURS_BASIS = 200000  # GRI 403-9 standard incident-rate basis (per 200,000 hours worked)

# ─── DATA LOADERS ───────────────────────────────────────────────────────────
# All 5 source files share the same shape: a merged title row above the real
# header (header=1), data living in one sheet, and multi-line column names
# (e.g. "SurfaceWater\nWithdrawn") - stripped here so downstream code can use
# plain identifiers like "SurfaceWaterWithdrawn".

# In-memory cache keyed by filename. Stores the last-read DataFrame and the
# file's mtime at the moment of the read. Automatically refreshes when the
# source Excel file is saved/updated on disk (mtime change detection).
_CACHE: dict = {}

def _load_sheet(filename, sheet_name):
    path = os.path.join(DATA_DIR, filename)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0
    entry = _CACHE.get(filename)
    if entry and entry["mtime"] == mtime:
        return entry["df"]
    df = pd.read_excel(path, sheet_name=sheet_name, header=1)
    df.columns = [str(c).replace("\n", "") for c in df.columns]
    _CACHE[filename] = {"df": df, "mtime": mtime, "loaded_at": time.time()}
    return df

def load_water_data():
    df = _load_sheet("GRI303_Water_Dataset_2019_2025.xlsx", "Raw_Monthly_Data").copy()
    df["ReportingPeriod"] = pd.to_datetime(df["ReportingPeriod"])
    df["FiscalReportingPeriod"] = df["FiscalReportingPeriod"].astype(int)
    return df

def load_waste_data():
    df = _load_sheet("GRI306_Waste_Dataset_2019_2025.xlsx", "Raw_Monthly_Data").copy()
    df["ReportingPeriod"] = pd.to_datetime(df["ReportingPeriod"])
    df["FiscalReportingPeriod"] = df["FiscalReportingPeriod"].astype(int)
    return df

def load_safety_data():
    df = _load_sheet("GRI403_OHS_Dataset_2019_2025.xlsx", "Plant_Monthly_Data").copy()
    df["ReportingPeriod"] = pd.to_datetime(df["ReportingPeriod"])
    df["FiscalReportingPeriod"] = df["FiscalReportingPeriod"].astype(int)
    return df

def load_energy_data():
    df = _load_sheet("GRI302_Energy_Dataset_2019_2025.xlsx", "Raw_Monthly_Data").copy()
    df["ReportingPeriod"] = pd.to_datetime(df["ReportingPeriod"])
    df["FiscalReportingPeriod"] = df["FiscalReportingPeriod"].astype(int)
    return df

def load_ghg_data():
    df = _load_sheet("GRI305_GHG_Dataset_2019_2025.xlsx", "Raw_Monthly_Data").copy()
    df["ReportingPeriod"] = pd.to_datetime(df["ReportingPeriod"])
    df["FiscalReportingPeriod"] = df["FiscalReportingPeriod"].astype(int)
    return df

def load_workforce_data():
    df = _load_sheet("BRSR_Workforce_Dataset_FY2020_FY2025.xlsx", "Raw_Annual_Data").copy()
    return df

def load_training_data():
    df = _load_sheet("BRSR_Training_Dataset_FY2020_FY2025.xlsx", "Raw_Annual_Data").copy()
    return df

def load_csr_data():
    df = _load_sheet("BRSR_CSR_Dataset_FY2020_FY2025.xlsx", "Raw_Annual_Data").copy()
    return df

def load_process_safety_data():
    df = _load_sheet("GRI_RTCH540a_ProcessSafety_Dataset_2019_2025.xlsx", "Raw_Monthly_Data").copy()
    df["ReportingPeriod"] = pd.to_datetime(df["ReportingPeriod"])
    df["FiscalReportingPeriod"] = df["FiscalReportingPeriod"].astype(int)
    return df

def safe_val(v):
    if isinstance(v, (np.integer,)): return int(v)
    if isinstance(v, (np.floating,)): return round(float(v), 2) if not np.isnan(v) else 0
    if isinstance(v, float) and np.isnan(v): return 0
    return v

def clean_dict(d):
    return {k: safe_val(v) for k, v in d.items()}

def safe_filter_plant(df, plant, col="PlantName"):
    if plant and plant.lower() not in ("all", "all plants") and plant in df[col].unique():
        return df[df[col] == plant]
    return df

def safe_filter_region(df, region):
    if region and region.lower() not in ("all", "all regions") and "Region" in df.columns and region in df["Region"].unique():
        return df[df["Region"] == region]
    return df

def rate_per_basis(count_sum, hours_sum, basis=RATE_HOURS_BASIS):
    return round(float(count_sum) / float(hours_sum) * basis, 2) if hours_sum else 0

# GRI 302/305 "...pert" columns are per-tonne ratios pre-computed per row in the
# source data. Averaging those per-row ratios across plants with uneven production
# volumes isn't the standard GRI methodology - intensity must be the weighted ratio
# of totals (sum of the absolute metric / sum of production), which is also what
# this dashboard's own report disclosure notes already describe.
INTENSITY_NUMERATOR = {
    "EnergyIntensityGJpert": "TotalEnergyConsumedGJ",
    "UpstreamEnergyIntensityGJpert": "UpstreamEnergyGJ",
    "DownstreamEnergyIntensityGJpert": "DownstreamEnergyGJ",
    "Scope1IntensitytCO2epert": "Scope1TotaltCO2e",
    "Scope2IntensitytCO2epert": "Scope2LocationBasedtCO2e",
    "Scope3IntensitytCO2epert": "Scope3TotaltCO2e",
}

def weighted_ratio(df_slice, numerator_col, production_col="CarbonBlackProductiont"):
    if len(df_slice) == 0:
        return 0.0
    production = float(df_slice[production_col].sum())
    if production <= 0:
        return 0.0
    return float(df_slice[numerator_col].sum()) / production

# ─── SASB calculation helpers ───────────────────────────────────────────────
# These reuse the exact same columns the GRI report builders already read
# (HazardousFlag from pdf_report.py's pivot_haz(), WorkerType, FatalitiesInjury/
# FatalitiesIllHealth) - no new dataset, just a different aggregation framing
# for SASB RT-CH disclosures. See SASB_INTEGRATION_PLAN.md Section 3.

def hazardous_waste_breakdown(df):
    """Splits a waste dataframe slice by HazardousFlag (Hazardous/Non-hazardous),
    reusing the same WasteCategory/ValueNumber columns the GRI 306 report uses."""
    haz = df[df["HazardousFlag"] == "Hazardous"]
    nonhaz = df[df["HazardousFlag"] == "Non-hazardous"]
    haz_diverted = float(haz[haz["WasteCategory"] == "Diverted"]["ValueNumber"].sum())
    haz_disposed = float(haz[haz["WasteCategory"] == "Disposed"]["ValueNumber"].sum())
    haz_generated = haz_diverted + haz_disposed
    nonhaz_generated = float(nonhaz["ValueNumber"].sum())
    haz_recycled_pct = round(haz_diverted / haz_generated * 100, 1) if haz_generated > 0 else 0
    return {
        "haz_generated": haz_generated,
        "nonhaz_generated": nonhaz_generated,
        "haz_diverted": haz_diverted,
        "haz_disposed": haz_disposed,
        "haz_recycled_pct": haz_recycled_pct,
    }

def disposal_method_breakdown(df):
    """Collapses the 12 granular IndicatorKey rows (Reuse/Recycling/Other x
    Hazardous/Non-hazardous, Incineration/IncinRecovery/Landfill x Hazardous/
    Non-hazardous) into 6 GRI 306-4/306-5 disposal-and-recovery methods,
    summed across both hazardous flags - this level of detail sits in the raw
    rows today but was never aggregated for the dashboard."""
    by_method = df.groupby("IndicatorName")["ValueNumber"].sum()
    methods = ["Reuse", "Recycling", "Other", "Incineration", "Incineration (with recovery)", "Landfill"]
    breakdown = {m: round(float(by_method.get(m, 0)), 1) for m in methods}
    generated = sum(breakdown.values())
    recycling_rate = round(breakdown["Recycling"] / generated * 100, 1) if generated > 0 else 0
    return breakdown, recycling_rate

def safety_rate_split(df, worker_type=None):
    """TRIR for all workers, or filtered to a single WorkerType (Employee/Contractor) -
    SASB RT-CH-320a.1 requires this split; GRI 403 reporting today sums across both."""
    d = df if worker_type is None else df[df["WorkerType"] == worker_type]
    return rate_per_basis(d["RecordableInjuries"].sum(), d["HoursWorked"].sum())

def fatality_rate(df):
    """SASB RT-CH-320a.1 fatality rate - same rate_per_basis() convention as TRIR/LTIFR,
    summing the two existing fatality columns (work-related injury + ill health)."""
    fatalities = df["FatalitiesInjury"].sum() + df["FatalitiesIllHealth"].sum()
    return rate_per_basis(fatalities, df["HoursWorked"].sum())

def process_safety_rates(df):
    """SASB RT-CH-540a Process Safety rates - PSTIR (Tier1+Tier2 incidents
    per 200k hrs) and PSISR (severity-weighted rate), same rate_per_basis()
    convention as TRIR/LTIFR/fatality rate above."""
    incidents = float(df["Tier1Incidents"].sum() + df["Tier2Incidents"].sum())
    hours = float(df["HoursWorked"].sum())
    return {
        "incidents": round(incidents, 0),
        "tier1": round(float(df["Tier1Incidents"].sum()), 0),
        "tier2": round(float(df["Tier2Incidents"].sum()), 0),
        "pstir": rate_per_basis(incidents, hours),
        "psisr": rate_per_basis(df["SeverityScore"].sum(), hours),
    }


def explode_weighted(df, label_col, weight_col, sep=";"):
    """Splits a semicolon-joined label column (e.g. MainInjuryType) into
    individual labels, distributing each row's weight (e.g. RecordableInjuries)
    evenly across the labels present in that row, then sums by label."""
    rows = []
    for _, r in df.iterrows():
        val = r.get(label_col)
        weight = r.get(weight_col, 0)
        if pd.isna(val) or not str(val).strip() or not weight:
            continue
        parts = [p.strip() for p in str(val).split(sep) if p.strip()]
        if not parts:
            continue
        share = float(weight) / len(parts)
        for p in parts:
            rows.append((p, share))
    if not rows:
        return pd.Series(dtype=float)
    return pd.DataFrame(rows, columns=["label", "weight"]).groupby("label")["weight"].sum().sort_values(ascending=False)

# ─── WATER ENDPOINTS ────────────────────────────────────────────────────────

@app.get("/api/water/kpis")
def get_water_kpis(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_water_data()
    if year:
        df = df[df["ReportingYear"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    total_withdrawn = df["TotalWaterWithdrawn"].sum()
    total_discharged = df["TotalWaterDischarged"].sum()
    total_consumed = df["WaterConsumed"].sum()
    total_stress = df["TotalWaterWithdrawnStressArea"].sum()
    stress_pct = round((total_stress / total_withdrawn * 100), 1) if total_withdrawn > 0 else 0
    return {
        "totalWithdrawn": round(float(total_withdrawn), 1),
        "totalDischarged": round(float(total_discharged), 1),
        "totalConsumed": round(float(total_consumed), 1),
        "waterFromStressArea": round(float(total_stress), 1),
        "stressAreaPct": stress_pct,
        "groundwater": round(float(df["GroundWater"].sum()), 1),
        "surfaceWater": round(float(df["SurfaceWaterWithdrawn"].sum()), 1),
        "thirdParty": round(float(df["ThirdPartyWaterWithdrawn"].sum()), 1),
        "municipalRainwater": round(float(df["MunicipalRainwater"].sum()), 1),
        "recycledReused": round(float(df["WaterRecycledReused"].sum()), 1),
    }

@app.get("/api/water/trends")
def get_water_trends(plant: Optional[str] = None):
    df = load_water_data()
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    yearly = df.groupby("ReportingYear").agg(
        withdrawn=("TotalWaterWithdrawn", "sum"),
        discharged=("TotalWaterDischarged", "sum"),
        consumed=("WaterConsumed", "sum"),
        stressArea=("TotalWaterWithdrawnStressArea", "sum"),
    ).reset_index()
    return yearly.rename(columns={"ReportingYear": "year"}).apply(
        lambda col: col.round(1) if col.dtype == float else col
    ).to_dict(orient="records")

@app.get("/api/water/monthly")
def get_water_monthly(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_water_data()
    if year:
        df = df[df["ReportingYear"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    monthly = df.groupby(["ReportingYear", "ReportingMonthNum", "ReportingMonth"]).agg(
        withdrawn=("TotalWaterWithdrawn", "sum"),
        discharged=("TotalWaterDischarged", "sum"),
        consumed=("WaterConsumed", "sum"),
    ).reset_index().sort_values(["ReportingYear", "ReportingMonthNum"])
    return monthly.to_dict(orient="records")

@app.get("/api/water/by-source")
def get_water_by_source(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_water_data()
    if year:
        df = df[df["ReportingYear"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    return [
        {"source": "Groundwater", "value": round(float(df["GroundWater"].sum()), 1)},
        {"source": "Third-party / Municipal", "value": round(float(df["ThirdPartyWaterWithdrawn"].sum()), 1)},
        {"source": "Surface Water", "value": round(float(df["SurfaceWaterWithdrawn"].sum()), 1)},
        {"source": "Seawater", "value": round(float(df["SeawaterWithdrawn"].sum()), 1)},
        {"source": "Produced Water", "value": round(float(df["ProducedWaterWithdrawn"].sum()), 1)},
    ]

@app.get("/api/water/by-plant")
def get_water_by_plant(year: Optional[int] = None):
    df = load_water_data()
    if year:
        df = df[df["ReportingYear"] == year]
    by_plant = df.groupby("PlantName").agg(
        withdrawn=("TotalWaterWithdrawn", "sum"),
        consumed=("WaterConsumed", "sum"),
    ).reset_index()
    return by_plant.to_dict(orient="records")

@app.get("/api/water/plants")
def get_water_plants():
    df = load_water_data()
    return sorted(df["PlantName"].unique().tolist())

@app.get("/api/water/years")
def get_water_years():
    df = load_water_data()
    return sorted(df["ReportingYear"].unique().tolist())

# ─── WASTE ENDPOINTS ────────────────────────────────────────────────────────
# GRI 306-4/306-5 raw data is row-per-(plant, month, indicator) with a
# WasteCategory of "Diverted" or "Disposed" - "generated" is the sum of both.

@app.get("/api/waste/kpis")
def get_waste_kpis(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_waste_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    diverted = df[df["WasteCategory"] == "Diverted"]["ValueNumber"].sum()
    disposed = df[df["WasteCategory"] == "Disposed"]["ValueNumber"].sum()
    generated = diverted + disposed
    diversion_rate = round((diverted / generated * 100), 1) if generated > 0 else 0
    disposal_rate = round((disposed / generated * 100), 1) if generated > 0 else 0
    return {
        "totalGenerated": round(float(generated), 1),
        "totalDiverted": round(float(diverted), 1),
        "totalDisposed": round(float(disposed), 1),
        "diversionRate": diversion_rate,
        "disposalRate": disposal_rate,
    }

@app.get("/api/waste/trends")
def get_waste_trends(plant: Optional[str] = None):
    df = load_waste_data()
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    pivot = df.groupby(["FiscalReportingPeriod", "WasteCategory"])["ValueNumber"].sum().unstack(fill_value=0).reset_index()
    pivot.columns.name = None
    result = []
    for _, row in pivot.iterrows():
        diverted = float(row.get("Diverted", 0))
        disposed = float(row.get("Disposed", 0))
        result.append({
            "year": int(row["FiscalReportingPeriod"]),
            "TotalWasteGenerated": round(diverted + disposed, 1),
            "WasteDiverted": round(diverted, 1),
            "WasteDisposed": round(disposed, 1),
        })
    return result

@app.get("/api/waste/monthly")
def get_waste_monthly(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_waste_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    df = df.copy()
    df["month"] = df["ReportingPeriod"].dt.month
    pivot = df.groupby(["month", "WasteCategory"])["ValueNumber"].sum().unstack(fill_value=0).reset_index()
    pivot.columns.name = None
    MONTHS = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    result = []
    for _, row in pivot.iterrows():
        diverted = float(row.get("Diverted", 0))
        disposed = float(row.get("Disposed", 0))
        result.append({
            "month": MONTHS.get(int(row["month"]), str(int(row["month"]))),
            "TotalWasteGenerated": round(diverted + disposed, 1),
            "WasteDiverted": round(diverted, 1),
            "WasteDisposed": round(disposed, 1),
        })
    return sorted(result, key=lambda x: list(MONTHS.values()).index(x["month"]))

@app.get("/api/waste/by-plant")
def get_waste_by_plant(year: Optional[int] = None):
    df = load_waste_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    div = df[df["WasteCategory"] == "Diverted"].groupby("PlantName")["ValueNumber"].sum()
    dis = df[df["WasteCategory"] == "Disposed"].groupby("PlantName")["ValueNumber"].sum()
    plants = sorted(set(div.index) | set(dis.index))
    result = []
    for plant in plants:
        d = float(div.get(plant, 0))
        s = float(dis.get(plant, 0))
        g = d + s
        result.append({"plant": plant, "generated": round(g, 1), "diverted": round(d, 1), "diversionRate": round(d/g*100, 1) if g > 0 else 0})
    return result

@app.get("/api/waste/plants")
def get_waste_plants():
    df = load_waste_data()
    return sorted(df["PlantName"].unique().tolist())

@app.get("/api/waste/years")
def get_waste_years():
    df = load_waste_data()
    return sorted(df["FiscalReportingPeriod"].unique().tolist())

# ─── SAFETY / OHS ENDPOINTS ─────────────────────────────────────────────────
# The new GRI 403 file is wide-format raw counts (Headcount, HoursWorked,
# RecordableInjuries, LostTimeInjuries, FatalitiesInjury, ...) split by
# WorkerType (Employee/Contractor) rather than precomputed rates. LTIFR/TRIR
# are now derived as sum(incidents)/sum(hours) x 200,000 per the GRI 403-9
# convention, rather than averaging a precomputed column.

@app.get("/api/safety/kpis")
def get_safety_kpis(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_safety_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]

    ltifr = rate_per_basis(df["LostTimeInjuries"].sum(), df["HoursWorked"].sum())
    trir = rate_per_basis(df["RecordableInjuries"].sum(), df["HoursWorked"].sum())
    fatal = df["FatalitiesInjury"].sum() + df["FatalitiesIllHealth"].sum()
    injuries = df["RecordableInjuries"].sum()
    ill_health = df["RecordableIllHealth"].sum()

    top_causes = explode_weighted(df, "MainInjuryType", "RecordableInjuries").head(5)
    top_causes = [{"cause": c, "count": round(float(v), 1)} for c, v in top_causes.items()]

    return {
        "ltifr": ltifr,
        "trir": trir,
        "fatalIncidents": round(float(fatal), 1),
        "totalInjuries": round(float(injuries), 1),
        "illHealthCases": round(float(ill_health), 1),
        "topCauses": top_causes,
    }

@app.get("/api/safety/trends")
def get_safety_trends(plant: Optional[str] = None):
    df = load_safety_data()
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    yearly = df.groupby("FiscalReportingPeriod").agg(
        hours=("HoursWorked", "sum"),
        lostTime=("LostTimeInjuries", "sum"),
        recordable=("RecordableInjuries", "sum"),
        fatal=("FatalitiesInjury", "sum"),
        illHealth=("RecordableIllHealth", "sum"),
    ).reset_index()
    result = []
    for _, row in yearly.iterrows():
        result.append({
            "year": int(row["FiscalReportingPeriod"]),
            "LTIFR": rate_per_basis(row["lostTime"], row["hours"]),
            "TRIR": rate_per_basis(row["recordable"], row["hours"]),
            "WorkRelatedInjuries": round(float(row["recordable"]), 1),
            "SeverityFatal": round(float(row["fatal"]), 1),
            "WorkRelatedIllHealth": round(float(row["illHealth"]), 1),
        })
    return sorted(result, key=lambda x: x["year"])

@app.get("/api/safety/by-plant")
def get_safety_by_plant(year: Optional[int] = None):
    df = load_safety_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    grouped = df.groupby("PlantName").agg(
        hours=("HoursWorked", "sum"),
        lostTime=("LostTimeInjuries", "sum"),
        recordable=("RecordableInjuries", "sum"),
    ).reset_index()
    return [
        {
            "plant": row["PlantName"],
            "ltifr": rate_per_basis(row["lostTime"], row["hours"]),
            "trir": rate_per_basis(row["recordable"], row["hours"]),
            "injuries": round(float(row["recordable"]), 1),
        }
        for _, row in grouped.iterrows()
    ]

@app.get("/api/safety/injury-types")
def get_safety_injury_types(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_safety_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    types = explode_weighted(df, "MainInjuryType", "RecordableInjuries")
    return [{"type": t, "value": round(float(v), 1)} for t, v in types.items()]

@app.get("/api/safety/severity")
def get_safety_severity(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_safety_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    return [
        {"severity": "Fatal", "value": round(float(df["FatalitiesInjury"].sum()), 1)},
        {"severity": "High Consequence", "value": round(float(df["HighConsequenceInjuries"].sum()), 1)},
        {"severity": "Recordable", "value": round(float(df["RecordableInjuries"].sum()), 1)},
        {"severity": "Lost Time", "value": round(float(df["LostTimeInjuries"].sum()), 1)},
        {"severity": "First Aid", "value": round(float(df["FirstAidCases"].sum()), 1)},
    ]

@app.get("/api/safety/causes")
def get_safety_causes(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_safety_data()
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    causes = explode_weighted(df, "MainInjuryType", "RecordableInjuries").head(10)
    return [{"cause": c, "count": round(float(v), 1)} for c, v in causes.items()]

@app.get("/api/safety/plants")
def get_safety_plants():
    df = load_safety_data()
    return sorted(df["PlantName"].unique().tolist())

@app.get("/api/safety/years")
def get_safety_years():
    df = load_safety_data()
    return sorted(df["FiscalReportingPeriod"].unique().tolist())

# ─── ENERGY ENDPOINTS (GRI 302) ─────────────────────────────────────────────

@app.get("/api/energy/kpis")
def get_energy_kpis(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_energy_data()
    if year:
        df = df[df["ReportingYear"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    consumed = df["TotalEnergyConsumedGJ"].sum()
    renewable = df["ElectricityRenewableGJ"].sum()
    non_renewable = df["ElectricityNonRenewableGJ"].sum()
    renewable_elec = renewable + non_renewable
    return {
        "totalConsumed": round(float(consumed), 1),
        "totalSold": round(float(df["TotalEnergySoldGJ"].sum()), 1),
        "netEnergy": round(float(df["NetEnergyGJ"].sum()), 1),
        "renewableGJ": round(float(renewable), 1),
        "nonRenewableGJ": round(float(non_renewable), 1),
        "renewablePct": round(renewable / renewable_elec * 100, 1) if renewable_elec > 0 else 0,
        "intensity": round(weighted_ratio(df, "TotalEnergyConsumedGJ"), 2),
    }

@app.get("/api/energy/plants")
def get_energy_plants():
    df = load_energy_data()
    return sorted(df["PlantName"].unique().tolist())

@app.get("/api/energy/years")
def get_energy_years():
    df = load_energy_data()
    return sorted(df["ReportingYear"].unique().tolist())

# ─── GHG / EMISSIONS ENDPOINTS (GRI 305) ────────────────────────────────────

@app.get("/api/ghg/kpis")
def get_ghg_kpis(year: Optional[int] = None, plant: Optional[str] = None):
    df = load_ghg_data()
    if year:
        df = df[df["ReportingYear"] == year]
    if plant and plant != "All":
        df = df[df["PlantName"] == plant]
    return {
        "scope1": round(float(df["Scope1TotaltCO2e"].sum()), 1),
        "scope2": round(float(df["Scope2LocationBasedtCO2e"].sum()), 1),
        "scope3": round(float(df["Scope3TotaltCO2e"].sum()), 1),
        "intensity": round(weighted_ratio(df, "Scope1TotaltCO2e"), 4),
        "nox": round(float(df["GrossNOxt"].sum()), 2),
        "sox": round(float(df["GrossSOxt"].sum()), 2),
        "voc": round(float(df["GrossVOCt"].sum()), 2),
        "pm": round(float(df["GrossPMt"].sum()), 2),
    }

@app.get("/api/ghg/plants")
def get_ghg_plants():
    df = load_ghg_data()
    return sorted(df["PlantName"].unique().tolist())

@app.get("/api/ghg/years")
def get_ghg_years():
    df = load_ghg_data()
    return sorted(df["ReportingYear"].unique().tolist())

# ─── FILTERS ─────────────────────────────────────────────────────────────

@app.get("/api/filters")
def get_filters(domain: Optional[str] = "water"):
    loaders = {
        "waste": (load_waste_data, "FiscalReportingPeriod"),
        "safety": (load_safety_data, "FiscalReportingPeriod"),
        "energy": (load_energy_data, "ReportingYear"),
        "emissions": (load_ghg_data, "ReportingYear"),
        "water": (load_water_data, "ReportingYear"),
    }
    loader, year_col = loaders.get(domain, loaders["water"])
    df = loader()
    years = sorted(df[year_col].unique().tolist())
    plants = sorted(df["PlantName"].unique().tolist())
    regions = sorted(df["Region"].dropna().unique().tolist()) if "Region" in df.columns else []
    return {"years": [int(y) for y in years], "plants": plants, "regions": regions}

# ─── ENVIRONMENT (AGGREGATE) ────────────────────────────────────────────────

@app.get("/api/environment/kpis")
def get_environment_kpis(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    wdf_all = load_water_data()
    wsdf_all = load_waste_data()
    edf_all = load_energy_data()
    gdf_all = load_ghg_data()

    def slice_df(d, yr, year_col):
        if yr:
            d = d[d[year_col] == yr]
        d = safe_filter_plant(d, plant)
        return safe_filter_region(d, region)

    # ── Current period values (full filter: all years if year=None) ──────
    w_cur  = slice_df(wdf_all,  year, "ReportingYear")
    ws_cur = slice_df(wsdf_all, year, "FiscalReportingPeriod")
    e_cur  = slice_df(edf_all,  year, "ReportingYear")
    g_cur  = slice_df(gdf_all,  year, "ReportingYear")

    # Water
    withdrawn  = float(w_cur["TotalWaterWithdrawn"].sum())
    discharged = float(w_cur["TotalWaterDischarged"].sum())
    consumed   = float(w_cur["WaterConsumed"].sum())
    stress     = float(w_cur["TotalWaterWithdrawnStressArea"].sum())
    stress_pct = round(stress / withdrawn * 100, 1) if withdrawn > 0 else 0

    # Waste
    diverted       = float(ws_cur[ws_cur["WasteCategory"] == "Diverted"]["ValueNumber"].sum())
    disposed       = float(ws_cur[ws_cur["WasteCategory"] == "Disposed"]["ValueNumber"].sum())
    generated      = diverted + disposed
    diversion_rate = round(diverted / generated * 100, 1) if generated > 0 else 0

    # Energy
    energy_consumed = float(e_cur["TotalEnergyConsumedGJ"].sum())
    renewable       = float(e_cur["ElectricityRenewableGJ"].sum())
    non_renewable   = float(e_cur["ElectricityNonRenewableGJ"].sum())
    e_intensity     = weighted_ratio(e_cur, "TotalEnergyConsumedGJ")

    # GHG
    scope1    = float(g_cur["Scope1TotaltCO2e"].sum())
    scope2    = float(g_cur["Scope2LocationBasedtCO2e"].sum())
    scope3    = float(g_cur["Scope3TotaltCO2e"].sum())
    total_ghg = scope1 + scope2 + scope3

    # ── YoY trend — always computed ───────────────────────────────────────
    # Selected year vs year-1. If no year filter, use latest year vs previous.
    trend_year = year if year else int(wdf_all["ReportingYear"].max())

    w_t  = slice_df(wdf_all,  trend_year,     "ReportingYear")
    w_p  = slice_df(wdf_all,  trend_year - 1, "ReportingYear")
    ws_t = slice_df(wsdf_all, trend_year,     "FiscalReportingPeriod")
    ws_p = slice_df(wsdf_all, trend_year - 1, "FiscalReportingPeriod")
    e_t  = slice_df(edf_all,  trend_year,     "ReportingYear")
    e_p  = slice_df(edf_all,  trend_year - 1, "ReportingYear")
    g_t  = slice_df(gdf_all,  trend_year,     "ReportingYear")
    g_p  = slice_df(gdf_all,  trend_year - 1, "ReportingYear")

    def pct_change(curr, prev):
        if not prev:
            return None
        return round((curr - prev) / prev * 100, 1)

    # Water trend values
    wt_withdrawn  = float(w_t["TotalWaterWithdrawn"].sum())
    wp_withdrawn  = float(w_p["TotalWaterWithdrawn"].sum())
    wt_discharged = float(w_t["TotalWaterDischarged"].sum())
    wp_discharged = float(w_p["TotalWaterDischarged"].sum())
    wt_consumed   = float(w_t["WaterConsumed"].sum())
    wp_consumed   = float(w_p["WaterConsumed"].sum())
    wt_stress_raw = float(w_t["TotalWaterWithdrawnStressArea"].sum())
    wt_stress_pct = round(wt_stress_raw / wt_withdrawn * 100, 1) if wt_withdrawn > 0 else 0
    wp_stress_raw = float(w_p["TotalWaterWithdrawnStressArea"].sum())
    wp_stress_pct = round(wp_stress_raw / wp_withdrawn * 100, 1) if wp_withdrawn > 0 else 0

    # Waste trend values
    wst_div  = float(ws_t[ws_t["WasteCategory"] == "Diverted"]["ValueNumber"].sum())
    wst_dis  = float(ws_t[ws_t["WasteCategory"] == "Disposed"]["ValueNumber"].sum())
    wst_gen  = wst_div + wst_dis
    wst_rate = round(wst_div / wst_gen * 100, 1) if wst_gen > 0 else 0
    wsp_div  = float(ws_p[ws_p["WasteCategory"] == "Diverted"]["ValueNumber"].sum())
    wsp_dis  = float(ws_p[ws_p["WasteCategory"] == "Disposed"]["ValueNumber"].sum())
    wsp_gen  = wsp_div + wsp_dis
    wsp_rate = round(wsp_div / wsp_gen * 100, 1) if wsp_gen > 0 else 0

    # Energy trend values
    et_consumed     = float(e_t["TotalEnergyConsumedGJ"].sum())
    ep_consumed     = float(e_p["TotalEnergyConsumedGJ"].sum())
    et_renewable    = float(e_t["ElectricityRenewableGJ"].sum())
    ep_renewable    = float(e_p["ElectricityRenewableGJ"].sum())
    et_non_ren      = float(e_t["ElectricityNonRenewableGJ"].sum())
    ep_non_ren      = float(e_p["ElectricityNonRenewableGJ"].sum())
    et_intensity    = weighted_ratio(e_t, "TotalEnergyConsumedGJ")
    ep_intensity    = weighted_ratio(e_p, "TotalEnergyConsumedGJ")

    # GHG trend values
    gt_s1    = float(g_t["Scope1TotaltCO2e"].sum())
    gp_s1    = float(g_p["Scope1TotaltCO2e"].sum())
    gt_s2    = float(g_t["Scope2LocationBasedtCO2e"].sum())
    gp_s2    = float(g_p["Scope2LocationBasedtCO2e"].sum())
    gt_s3    = float(g_t["Scope3TotaltCO2e"].sum())
    gp_s3    = float(g_p["Scope3TotaltCO2e"].sum())
    gt_total = gt_s1 + gt_s2 + gt_s3
    gp_total = gp_s1 + gp_s2 + gp_s3

    trend = {
        "withdrawn":    pct_change(wt_withdrawn,  wp_withdrawn),
        "discharged":   pct_change(wt_discharged, wp_discharged),
        "consumed":     pct_change(wt_consumed,   wp_consumed),
        "stress":       pct_change(wt_stress_pct, wp_stress_pct),
        "generated":    pct_change(wst_gen,  wsp_gen),
        "diverted":     pct_change(wst_div,  wsp_div),
        "disposed":     pct_change(wst_dis,  wsp_dis),
        "diversion":    pct_change(wst_rate, wsp_rate),
        "energy":       pct_change(et_consumed,  ep_consumed),
        "renewable":    pct_change(et_renewable, ep_renewable),
        "non_renewable":pct_change(et_non_ren,   ep_non_ren),
        "e_intensity":  pct_change(et_intensity, ep_intensity),
        "scope1":       pct_change(gt_s1,    gp_s1),
        "scope2":       pct_change(gt_s2,    gp_s2),
        "scope3":       pct_change(gt_s3,    gp_s3),
        "total_ghg":    pct_change(gt_total, gp_total),
    }

    # Convert water m³ → megaliters (1 ML = 1,000 m³)
    withdrawn_ml  = round(withdrawn  / 1000, 3)
    discharged_ml = round(discharged / 1000, 3)
    consumed_ml   = round(consumed   / 1000, 3)

    return [
        # Water (4)
        {"id": "water-withdrawn",      "label": "Total Water Withdrawn",   "value": withdrawn_ml,          "unit": "ML",     "trend": trend["withdrawn"],     "gri": "GRI 303-3", "status": "live"},
        {"id": "water-discharged",     "label": "Total Water Discharged",  "value": discharged_ml,         "unit": "ML",     "trend": trend["discharged"],    "gri": "GRI 303-4", "status": "live"},
        {"id": "water-consumed",       "label": "Total Water Consumed",    "value": consumed_ml,           "unit": "ML",     "trend": trend["consumed"],      "gri": "GRI 303-5", "status": "live"},
        {"id": "water-stress",         "label": "Water from Stress Areas", "value": stress_pct,            "unit": "%",      "trend": trend["stress"],        "gri": "GRI 303-3", "status": "live"},
        # Waste (4)
        {"id": "waste-generated",      "label": "Total Waste Generated",   "value": round(generated, 1),   "unit": "tonnes", "trend": trend["generated"],     "gri": "GRI 306-3", "status": "live"},
        {"id": "waste-diverted",       "label": "Total Waste Diverted",    "value": round(diverted, 1),    "unit": "tonnes", "trend": trend["diverted"],      "gri": "GRI 306-4", "status": "live"},
        {"id": "waste-disposed",       "label": "Total Waste Disposed",    "value": round(disposed, 1),    "unit": "tonnes", "trend": trend["disposed"],      "gri": "GRI 306-5", "status": "live"},
        {"id": "waste-diversion",      "label": "Waste Diversion Rate",    "value": diversion_rate,        "unit": "%",      "trend": trend["diversion"],     "gri": "GRI 306-4", "status": "live"},
        # Energy (4)
        {"id": "energy-consumed",      "label": "Total Energy Consumed",   "value": round(energy_consumed, 1), "unit": "GJ", "trend": trend["energy"],        "gri": "GRI 302-1", "status": "live"},
        {"id": "renewable-energy",     "label": "Renewable Energy",        "value": round(renewable, 1),   "unit": "GJ",     "trend": trend["renewable"],     "gri": "GRI 302-1", "status": "live"},
        {"id": "non-renewable-energy", "label": "Non-Renewable Energy",    "value": round(non_renewable, 1),"unit": "GJ",    "trend": trend["non_renewable"], "gri": "GRI 302-1", "status": "live"},
        {"id": "energy-intensity",     "label": "Energy Intensity",        "value": round(e_intensity, 2), "unit": "GJ/t",   "trend": trend["e_intensity"],   "gri": "GRI 302-3", "status": "live"},
        # Emissions (4)
        {"id": "total-ghg",            "label": "Total GHG Emissions",     "value": round(total_ghg, 1),   "unit": "tCO₂e",  "trend": trend["total_ghg"],     "gri": "GRI 305",   "status": "live"},
        {"id": "scope1-ghg",           "label": "Scope 1 Emissions",       "value": round(scope1, 1),      "unit": "tCO₂e",  "trend": trend["scope1"],        "gri": "GRI 305-1", "status": "live"},
        {"id": "scope2-ghg",           "label": "Scope 2 Emissions",       "value": round(scope2, 1),      "unit": "tCO₂e",  "trend": trend["scope2"],        "gri": "GRI 305-2", "status": "live"},
        {"id": "scope3-ghg",           "label": "Scope 3 Emissions",       "value": round(scope3, 1),      "unit": "tCO₂e",  "trend": trend["scope3"],        "gri": "GRI 305-3", "status": "live"},
    ]

@app.get("/api/environment/water")
def get_environment_water(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None, view: Optional[str] = "monthly"):
    df_all = load_water_data()
    df_all = safe_filter_plant(df_all, plant)
    df_all = safe_filter_region(df_all, region)

    def ml(v): return round(float(v) / 1000, 3)  # m³ → megalitres (GRI standard)

    def stress_by_source(d):
        # Per-source water stress % (GRI 303-3) - fraction of each withdrawal
        # source drawn from a water-stress area, not just the aggregate %.
        sources = {
            "groundwater":  ("GroundWater",             "GroundWaterStress"),
            "surface":      ("SurfaceWaterWithdrawn",   "SurfaceWaterStress"),
            "seawater":     ("SeawaterWithdrawn",       "SeawaterStress"),
            "produced":     ("ProducedWaterWithdrawn",  "ProducedWaterStress"),
            "third_party":  ("ThirdPartyWaterWithdrawn","ThirdPartyWaterStress"),
        }
        out = {}
        for key, (wcol, scol) in sources.items():
            withdrawn_v = float(d[wcol].sum())
            stress_v = float(d[scol].sum())
            out[key] = round(stress_v / withdrawn_v * 100, 1) if withdrawn_v > 0 else 0
        return out

    AGG = dict(
        withdrawn=("TotalWaterWithdrawn", "sum"),
        discharged=("TotalWaterDischarged", "sum"),
        groundwater=("GroundWater", "sum"),
        third_party=("ThirdPartyWaterWithdrawn", "sum"),
        municipal_rainwater=("MunicipalRainwater", "sum"),
        surface_withdrawn=("SurfaceWaterWithdrawn", "sum"),
        seawater=("SeawaterWithdrawn", "sum"),
        produced_water=("ProducedWaterWithdrawn", "sum"),
        industrial=("IndustrialTreatment", "sum"),
        sewage=("MunicipalSewage", "sum"),
        surface_discharged=("FreshSurfaceWaterDischarged", "sum"),
        recycled=("WaterRecycledReused", "sum"),
        withdrawal_fresh=("FreshwaterWithdrawnLT1000TDS", "sum"),
        withdrawal_other=("OtherWaterWithdrawnGT1000TDS", "sum"),
        discharge_fresh=("FreshwaterDischargedLT1000TDS", "sum"),
        discharge_other=("OtherWaterDischargedGT1000TDS", "sum"),
    )

    if view == "yearly":
        yearly = df_all.groupby("ReportingYear").agg(**AGG).reset_index().sort_values("ReportingYear")
        by_plant = df_all.groupby("PlantName")["TotalWaterWithdrawn"].sum().reset_index()
        recycled_pct = [round(r / w * 100, 1) if w > 0 else 0 for r, w in zip(yearly["recycled"], yearly["withdrawn"])]
        return {
            "view": "yearly",
            "labels":            [int(y) for y in yearly["ReportingYear"]],
            "withdrawn":         [ml(v) for v in yearly["withdrawn"]],
            "discharged":        [ml(v) for v in yearly["discharged"]],
            "groundwater":       [ml(v) for v in yearly["groundwater"]],
            "third_party":       [ml(v) for v in yearly["third_party"]],
            "municipal_rainwater":[ml(v) for v in yearly["municipal_rainwater"]],
            "surface_withdrawn": [ml(v) for v in yearly["surface_withdrawn"]],
            "seawater":          [ml(v) for v in yearly["seawater"]],
            "produced_water":    [ml(v) for v in yearly["produced_water"]],
            "industrial":        [ml(v) for v in yearly["industrial"]],
            "sewage":            [ml(v) for v in yearly["sewage"]],
            "surface_discharged":[ml(v) for v in yearly["surface_discharged"]],
            "recycled":          [ml(v) for v in yearly["recycled"]],
            "recycled_pct":      recycled_pct,
            "withdrawal_fresh":  [ml(v) for v in yearly["withdrawal_fresh"]],
            "withdrawal_other":  [ml(v) for v in yearly["withdrawal_other"]],
            "discharge_fresh":   [ml(v) for v in yearly["discharge_fresh"]],
            "discharge_other":   [ml(v) for v in yearly["discharge_other"]],
            "stress_by_source":  stress_by_source(df_all),
            "plants":            by_plant["PlantName"].tolist(),
            "withdrawn_by_plant":[ml(v) for v in by_plant["TotalWaterWithdrawn"]],
        }

    if not year:
        year = int(df_all["ReportingYear"].max())
    df = df_all[df_all["ReportingYear"] == year]

    monthly = df.groupby(["ReportingMonthNum", "ReportingMonth"]).agg(**AGG).reset_index().sort_values("ReportingMonthNum")
    recycled_pct = [round(r / w * 100, 1) if w > 0 else 0 for r, w in zip(monthly["recycled"], monthly["withdrawn"])]

    by_plant = df.groupby("PlantName")["TotalWaterWithdrawn"].sum().reset_index()

    return {
        "months": monthly["ReportingMonth"].tolist(),
        "withdrawn":         [ml(v) for v in monthly["withdrawn"]],
        "discharged":        [ml(v) for v in monthly["discharged"]],
        "groundwater":       [ml(v) for v in monthly["groundwater"]],
        "third_party":       [ml(v) for v in monthly["third_party"]],
        "municipal_rainwater":[ml(v) for v in monthly["municipal_rainwater"]],
        "surface_withdrawn": [ml(v) for v in monthly["surface_withdrawn"]],
        "seawater":          [ml(v) for v in monthly["seawater"]],
        "produced_water":    [ml(v) for v in monthly["produced_water"]],
        "industrial":        [ml(v) for v in monthly["industrial"]],
        "sewage":            [ml(v) for v in monthly["sewage"]],
        "surface_discharged":[ml(v) for v in monthly["surface_discharged"]],
        "recycled":          [ml(v) for v in monthly["recycled"]],
        "recycled_pct":      recycled_pct,
        "withdrawal_fresh":  [ml(v) for v in monthly["withdrawal_fresh"]],
        "withdrawal_other":  [ml(v) for v in monthly["withdrawal_other"]],
        "discharge_fresh":   [ml(v) for v in monthly["discharge_fresh"]],
        "discharge_other":   [ml(v) for v in monthly["discharge_other"]],
        "stress_by_source":  stress_by_source(df),
        "plants":            by_plant["PlantName"].tolist(),
        "withdrawn_by_plant":[ml(v) for v in by_plant["TotalWaterWithdrawn"]],
        "year": year,
    }

@app.get("/api/environment/waste")
def get_environment_waste(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None, view: Optional[str] = "monthly"):
    df_all = load_waste_data()
    df_all = safe_filter_plant(df_all, plant)
    df_all = safe_filter_region(df_all, region)
    MONTHS = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}

    if view == "yearly":
        yearly_pivot = df_all.groupby(["FiscalReportingPeriod", "WasteCategory"])["ValueNumber"].sum().unstack(fill_value=0)
        for col in ["Diverted", "Disposed"]:
            if col not in yearly_pivot.columns:
                yearly_pivot[col] = 0
        yearly_pivot["Generated"] = yearly_pivot["Diverted"] + yearly_pivot["Disposed"]
        yearly_pivot = yearly_pivot.reset_index().sort_values("FiscalReportingPeriod")
        by_plant_div = df_all[df_all["WasteCategory"] == "Diverted"].groupby("PlantName")["ValueNumber"].sum()
        by_plant_dis = df_all[df_all["WasteCategory"] == "Disposed"].groupby("PlantName")["ValueNumber"].sum()
        plants = sorted(set(by_plant_div.index) | set(by_plant_dis.index))
        total_diverted = float(df_all[df_all["WasteCategory"] == "Diverted"]["ValueNumber"].sum())
        total_disposed = float(df_all[df_all["WasteCategory"] == "Disposed"]["ValueNumber"].sum())
        disposal_methods, recycling_rate = disposal_method_breakdown(df_all)
        return {
            "view": "yearly",
            "labels": [int(y) for y in yearly_pivot["FiscalReportingPeriod"]],
            "generated": [round(float(v), 1) for v in yearly_pivot["Generated"]],
            "diverted": [round(float(v), 1) for v in yearly_pivot["Diverted"]],
            "disposed": [round(float(v), 1) for v in yearly_pivot["Disposed"]],
            "plants": plants,
            "generated_by_plant": [round(float(by_plant_div.get(p, 0)) + float(by_plant_dis.get(p, 0)), 1) for p in plants],
            "diverted_by_plant": [round(float(by_plant_div.get(p, 0)), 1) for p in plants],
            "disposed_by_plant": [round(float(by_plant_dis.get(p, 0)), 1) for p in plants],
            "total_generated": round(total_diverted + total_disposed, 1),
            "total_diverted": round(total_diverted, 1),
            "total_disposed": round(total_disposed, 1),
            "disposal_methods": disposal_methods,
            "recycling_rate": recycling_rate,
        }

    if not year:
        year = int(df_all["FiscalReportingPeriod"].max())
    df = df_all[df_all["FiscalReportingPeriod"] == year].copy()
    df["month"] = df["ReportingPeriod"].dt.month

    pivot = df.groupby(["month", "WasteCategory"])["ValueNumber"].sum().unstack(fill_value=0)
    for col in ["Diverted", "Disposed"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot["Generated"] = pivot["Diverted"] + pivot["Disposed"]
    pivot = pivot.reset_index().sort_values("month")

    by_plant_div = df[df["WasteCategory"] == "Diverted"].groupby("PlantName")["ValueNumber"].sum()
    by_plant_dis = df[df["WasteCategory"] == "Disposed"].groupby("PlantName")["ValueNumber"].sum()
    plants = sorted(set(by_plant_div.index) | set(by_plant_dis.index))

    total_diverted = float(df[df["WasteCategory"] == "Diverted"]["ValueNumber"].sum())
    total_disposed = float(df[df["WasteCategory"] == "Disposed"]["ValueNumber"].sum())
    total_generated = total_diverted + total_disposed
    disposal_methods, recycling_rate = disposal_method_breakdown(df)

    return {
        "months": [MONTHS.get(int(m), str(int(m))) for m in pivot["month"]],
        "generated": [round(float(v), 1) for v in pivot["Generated"]],
        "diverted": [round(float(v), 1) for v in pivot["Diverted"]],
        "disposed": [round(float(v), 1) for v in pivot["Disposed"]],
        "plants": plants,
        "generated_by_plant": [round(float(by_plant_div.get(p, 0)) + float(by_plant_dis.get(p, 0)), 1) for p in plants],
        "diverted_by_plant": [round(float(by_plant_div.get(p, 0)), 1) for p in plants],
        "disposed_by_plant": [round(float(by_plant_dis.get(p, 0)), 1) for p in plants],
        "total_generated": round(total_generated, 1),
        "total_diverted": round(total_diverted, 1),
        "total_disposed": round(total_disposed, 1),
        "disposal_methods": disposal_methods,
        "recycling_rate": recycling_rate,
        "year": year,
    }

@app.get("/api/environment/energy")
def get_environment_energy(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None, view: Optional[str] = "monthly"):
    df_all = load_energy_data()
    df_all = safe_filter_plant(df_all, plant)
    df_all = safe_filter_region(df_all, region)

    if view == "yearly":
        yearly = df_all.groupby("ReportingYear").agg(
            oil=("OilNonProdGJ", "sum"),
            naturalGas=("NaturalGasNonProdGJ", "sum"),
            electricity=("ElectricityConsumedGJ", "sum"),
            steam=("SteamConsumedGJ", "sum"),
            tailGas=("TailGasConsumedGJ", "sum"),
            compressedAir=("CompressedAirGJ", "sum"),
            hotWater=("HotWaterConsumedGJ", "sum"),
            total=("TotalEnergyConsumedGJ", "sum"),
            sold=("TotalEnergySoldGJ", "sum"),
            upstream=("UpstreamEnergyGJ", "sum"),
            downstream=("DownstreamEnergyGJ", "sum"),
            renewable=("ElectricityRenewableGJ", "sum"),
            nonRenewable=("ElectricityNonRenewableGJ", "sum"),
            production=("CarbonBlackProductiont", "sum"),
        ).reset_index().sort_values("ReportingYear")
        # NetEnergyGJ in the source dataset is a generation artifact that
        # duplicates TotalEnergyConsumedGJ rather than subtracting sold
        # energy as its name implies - compute net energy directly instead.
        yearly["netEnergy"] = yearly["total"] - yearly["sold"]
        yearly["intensity"] = (yearly["total"] / yearly["production"]).where(yearly["production"] > 0, 0)
        yearly["upstreamIntensity"] = (yearly["upstream"] / yearly["production"]).where(yearly["production"] > 0, 0)
        yearly["downstreamIntensity"] = (yearly["downstream"] / yearly["production"]).where(yearly["production"] > 0, 0)
        by_plant = df_all.groupby("PlantName")["TotalEnergyConsumedGJ"].sum().reset_index()
        return {
            "view": "yearly",
            "labels": [int(y) for y in yearly["ReportingYear"]],
            "oil": [round(float(v), 1) for v in yearly["oil"]],
            "naturalGas": [round(float(v), 1) for v in yearly["naturalGas"]],
            "electricity": [round(float(v), 1) for v in yearly["electricity"]],
            "steam": [round(float(v), 1) for v in yearly["steam"]],
            "tailGas": [round(float(v), 1) for v in yearly["tailGas"]],
            "compressedAir": [round(float(v), 1) for v in yearly["compressedAir"]],
            "hotWater": [round(float(v), 1) for v in yearly["hotWater"]],
            "totalConsumed": round(float(df_all["TotalEnergyConsumedGJ"].sum()), 1),
            "totalSold": round(float(df_all["TotalEnergySoldGJ"].sum()), 1),
            "netEnergy": round(float(df_all["TotalEnergyConsumedGJ"].sum() - df_all["TotalEnergySoldGJ"].sum()), 1),
            "trendYears": [int(y) for y in yearly["ReportingYear"]],
            "soldTrend": [round(float(v), 1) for v in yearly["sold"]],
            "netEnergyTrend": [round(float(v), 1) for v in yearly["netEnergy"]],
            "upstreamTrend": [round(float(v), 1) for v in yearly["upstream"]],
            "downstreamTrend": [round(float(v), 1) for v in yearly["downstream"]],
            "upstreamIntensityTrend": [round(float(v), 4) for v in yearly["upstreamIntensity"]],
            "downstreamIntensityTrend": [round(float(v), 4) for v in yearly["downstreamIntensity"]],
            "renewableTrend": [round(float(v), 1) for v in yearly["renewable"]],
            "nonRenewableTrend": [round(float(v), 1) for v in yearly["nonRenewable"]],
            "intensityTrend": [round(float(v), 2) for v in yearly["intensity"]],
            "plants": by_plant["PlantName"].tolist(),
            "consumed_by_plant": [round(float(v), 1) for v in by_plant["TotalEnergyConsumedGJ"]],
        }

    if not year:
        year = int(df_all["ReportingYear"].max())
    df = df_all[df_all["ReportingYear"] == year]

    monthly = df.groupby(["ReportingMonthNum", "ReportingMonth"]).agg(
        oil=("OilNonProdGJ", "sum"),
        naturalGas=("NaturalGasNonProdGJ", "sum"),
        electricity=("ElectricityConsumedGJ", "sum"),
        steam=("SteamConsumedGJ", "sum"),
        tailGas=("TailGasConsumedGJ", "sum"),
        compressedAir=("CompressedAirGJ", "sum"),
        hotWater=("HotWaterConsumedGJ", "sum"),
        total=("TotalEnergyConsumedGJ", "sum"),
    ).reset_index().sort_values("ReportingMonthNum")

    yearly = df_all.groupby("ReportingYear").agg(
        sold=("TotalEnergySoldGJ", "sum"),
        upstream=("UpstreamEnergyGJ", "sum"),
        downstream=("DownstreamEnergyGJ", "sum"),
        renewable=("ElectricityRenewableGJ", "sum"),
        nonRenewable=("ElectricityNonRenewableGJ", "sum"),
        total=("TotalEnergyConsumedGJ", "sum"),
        production=("CarbonBlackProductiont", "sum"),
    ).reset_index().sort_values("ReportingYear")
    yearly["netEnergy"] = yearly["total"] - yearly["sold"]
    yearly["intensity"] = (yearly["total"] / yearly["production"]).where(yearly["production"] > 0, 0)
    yearly["upstreamIntensity"] = (yearly["upstream"] / yearly["production"]).where(yearly["production"] > 0, 0)
    yearly["downstreamIntensity"] = (yearly["downstream"] / yearly["production"]).where(yearly["production"] > 0, 0)

    by_plant = df.groupby("PlantName")["TotalEnergyConsumedGJ"].sum().reset_index()

    return {
        "year": year,
        "months": monthly["ReportingMonth"].tolist(),
        "oil": [round(float(v), 1) for v in monthly["oil"]],
        "naturalGas": [round(float(v), 1) for v in monthly["naturalGas"]],
        "electricity": [round(float(v), 1) for v in monthly["electricity"]],
        "steam": [round(float(v), 1) for v in monthly["steam"]],
        "tailGas": [round(float(v), 1) for v in monthly["tailGas"]],
        "compressedAir": [round(float(v), 1) for v in monthly["compressedAir"]],
        "hotWater": [round(float(v), 1) for v in monthly["hotWater"]],
        "totalConsumed": round(float(df["TotalEnergyConsumedGJ"].sum()), 1),
        "totalSold": round(float(df["TotalEnergySoldGJ"].sum()), 1),
        "netEnergy": round(float(df["TotalEnergyConsumedGJ"].sum() - df["TotalEnergySoldGJ"].sum()), 1),
        "trendYears": [int(y) for y in yearly["ReportingYear"]],
        "soldTrend": [round(float(v), 1) for v in yearly["sold"]],
        "netEnergyTrend": [round(float(v), 1) for v in yearly["netEnergy"]],
        "upstreamTrend": [round(float(v), 1) for v in yearly["upstream"]],
        "downstreamTrend": [round(float(v), 1) for v in yearly["downstream"]],
        "upstreamIntensityTrend": [round(float(v), 4) for v in yearly["upstreamIntensity"]],
        "downstreamIntensityTrend": [round(float(v), 4) for v in yearly["downstreamIntensity"]],
        "renewableTrend": [round(float(v), 1) for v in yearly["renewable"]],
        "nonRenewableTrend": [round(float(v), 1) for v in yearly["nonRenewable"]],
        "intensityTrend": [round(float(v), 2) for v in yearly["intensity"]],
        "plants": by_plant["PlantName"].tolist(),
        "consumed_by_plant": [round(float(v), 1) for v in by_plant["TotalEnergyConsumedGJ"]],
    }

SCOPE3_CATEGORIES = {
    "Scope3Cat1PurchasedGoodstCO2e":    "Purchased Goods & Services",
    "Scope3Cat3FuelEnergytCO2e":        "Fuel & Energy (not in Scope 1/2)",
    "Scope3Cat4UpstreamTransporttCO2e": "Upstream Transportation",
    "Scope3Cat5WastetCO2e":             "Waste Generated in Operations",
    "Scope3Cat6BusinessTraveltCO2e":    "Business Travel",
    "Scope3Cat7CommutetCO2e":           "Employee Commuting",
    "Scope3Cat9DownstreamTransporttCO2e":"Downstream Transportation",
    "Scope3Cat11UseSoldProductstCO2e":  "Use of Sold Products",
    "Scope3Cat12EndOfLifetCO2e":        "End-of-Life Treatment",
    "Scope3Cat15InvestmentstCO2e":      "Investments",
}

SCOPE1_FUEL_SOURCES = {
    "Scope1TailGasCombustiontCO2e":    "Tail Gas",
    "Scope1NaturalGasCombustiontCO2e": "Natural Gas",
    "Scope1FuelOilCombustiontCO2e":    "Fuel Oil",
    "Scope1OtherFuelCombustiontCO2e":  "Other Fuel",
}

def ghg_breakdowns(d):
    # Scope 1 by emission source (GRI 305-1) and Scope 3 by GHG Protocol
    # category (GRI 305-3) - both already computed per-row in the source data
    # but never aggregated/charted before; reused here as a single snapshot
    # breakdown over whatever period the caller has already filtered to.
    scope1_source = {
        "process":    round(float(d["Scope1ProcessEmissionstCO2e"].sum()), 1),
        "stationary": round(float(d["Scope1StationaryCombustiontCO2e"].sum()), 1),
    }
    scope1_fuel = {label: round(float(d[col].sum()), 1) for col, label in SCOPE1_FUEL_SOURCES.items()}
    scope3_category = {label: round(float(d[col].sum()), 1) for col, label in SCOPE3_CATEGORIES.items()}
    return scope1_source, scope1_fuel, scope3_category

@app.get("/api/environment/ghg")
def get_environment_ghg(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None, view: Optional[str] = "monthly"):
    df_all = load_ghg_data()
    df_all = safe_filter_plant(df_all, plant)
    df_all = safe_filter_region(df_all, region)

    if view == "yearly":
        yearly = df_all.groupby("ReportingYear").agg(
            scope1=("Scope1TotaltCO2e", "sum"),
            scope2=("Scope2LocationBasedtCO2e", "sum"),
            scope2_market=("Scope2MarketBasedtCO2e", "sum"),
            scope3=("Scope3TotaltCO2e", "sum"),
            production=("CarbonBlackProductiont", "sum"),
            nox=("GrossNOxt", "sum"),
            sox=("GrossSOxt", "sum"),
            voc=("GrossVOCt", "sum"),
            pm=("GrossPMt", "sum"),
        ).reset_index().sort_values("ReportingYear")
        yearly["intensity"] = (yearly["scope1"] / yearly["production"]).where(yearly["production"] > 0, 0)
        yearly["intensity2"] = (yearly["scope2"] / yearly["production"]).where(yearly["production"] > 0, 0)
        yearly["intensity3"] = (yearly["scope3"] / yearly["production"]).where(yearly["production"] > 0, 0)
        by_plant = df_all.groupby("PlantName").agg(
            scope1_p=("Scope1TotaltCO2e", "sum"),
            scope2_p=("Scope2LocationBasedtCO2e", "sum"),
        ).reset_index()
        plants = sorted(by_plant["PlantName"].tolist())
        scope1_source, scope1_fuel, scope3_category = ghg_breakdowns(df_all)
        return {
            "view": "yearly",
            "labels": [int(y) for y in yearly["ReportingYear"]],
            "scope1": round(float(df_all["Scope1TotaltCO2e"].sum()), 1),
            "scope2": round(float(df_all["Scope2LocationBasedtCO2e"].sum()), 1),
            "scope2_market": round(float(df_all["Scope2MarketBasedtCO2e"].sum()), 1),
            "scope3": round(float(df_all["Scope3TotaltCO2e"].sum()), 1),
            "nox": [round(float(v), 2) for v in yearly["nox"]],
            "sox": [round(float(v), 2) for v in yearly["sox"]],
            "voc": [round(float(v), 2) for v in yearly["voc"]],
            "pm": [round(float(v), 2) for v in yearly["pm"]],
            "trendYears": [int(y) for y in yearly["ReportingYear"]],
            "scope1Trend": [round(float(v), 1) for v in yearly["scope1"]],
            "scope2Trend": [round(float(v), 1) for v in yearly["scope2"]],
            "scope2MarketTrend": [round(float(v), 1) for v in yearly["scope2_market"]],
            "scope3Trend": [round(float(v), 1) for v in yearly["scope3"]],
            "intensityTrend": [round(float(v), 4) for v in yearly["intensity"]],
            "intensity2Trend": [round(float(v), 4) for v in yearly["intensity2"]],
            "intensity3Trend": [round(float(v), 4) for v in yearly["intensity3"]],
            "scope1_source": scope1_source,
            "scope1_fuel": scope1_fuel,
            "scope3_category": scope3_category,
            "plants": plants,
            "scope1_by_plant": [round(float(by_plant.loc[by_plant["PlantName"] == p, "scope1_p"].values[0]), 1) for p in plants],
            "scope2_by_plant": [round(float(by_plant.loc[by_plant["PlantName"] == p, "scope2_p"].values[0]), 1) for p in plants],
        }

    if not year:
        year = int(df_all["ReportingYear"].max())
    df = df_all[df_all["ReportingYear"] == year]

    monthly = df.groupby(["ReportingMonthNum", "ReportingMonth"]).agg(
        nox=("GrossNOxt", "sum"),
        sox=("GrossSOxt", "sum"),
        voc=("GrossVOCt", "sum"),
        pm=("GrossPMt", "sum"),
    ).reset_index().sort_values("ReportingMonthNum")

    yearly = df_all.groupby("ReportingYear").agg(
        scope1=("Scope1TotaltCO2e", "sum"),
        scope2=("Scope2LocationBasedtCO2e", "sum"),
        scope2_market=("Scope2MarketBasedtCO2e", "sum"),
        scope3=("Scope3TotaltCO2e", "sum"),
        production=("CarbonBlackProductiont", "sum"),
    ).reset_index().sort_values("ReportingYear")
    yearly["intensity"] = (yearly["scope1"] / yearly["production"]).where(yearly["production"] > 0, 0)
    yearly["intensity2"] = (yearly["scope2"] / yearly["production"]).where(yearly["production"] > 0, 0)
    yearly["intensity3"] = (yearly["scope3"] / yearly["production"]).where(yearly["production"] > 0, 0)

    # Plant-level breakdown for cross-filtering
    by_plant = df.groupby("PlantName").agg(
        scope1_p=("Scope1TotaltCO2e", "sum"),
        scope2_p=("Scope2LocationBasedtCO2e", "sum"),
    ).reset_index()
    plants = sorted(by_plant["PlantName"].tolist())
    scope1_source, scope1_fuel, scope3_category = ghg_breakdowns(df)

    return {
        "year": year,
        "scope1": round(float(df["Scope1TotaltCO2e"].sum()), 1),
        "scope2": round(float(df["Scope2LocationBasedtCO2e"].sum()), 1),
        "scope2_market": round(float(df["Scope2MarketBasedtCO2e"].sum()), 1),
        "scope3": round(float(df["Scope3TotaltCO2e"].sum()), 1),
        "months": monthly["ReportingMonth"].tolist(),
        "nox": [round(float(v), 2) for v in monthly["nox"]],
        "sox": [round(float(v), 2) for v in monthly["sox"]],
        "voc": [round(float(v), 2) for v in monthly["voc"]],
        "pm": [round(float(v), 2) for v in monthly["pm"]],
        "trendYears": [int(y) for y in yearly["ReportingYear"]],
        "scope1Trend": [round(float(v), 1) for v in yearly["scope1"]],
        "scope2Trend": [round(float(v), 1) for v in yearly["scope2"]],
        "scope2MarketTrend": [round(float(v), 1) for v in yearly["scope2_market"]],
        "scope3Trend": [round(float(v), 1) for v in yearly["scope3"]],
        "intensityTrend": [round(float(v), 4) for v in yearly["intensity"]],
        "intensity2Trend": [round(float(v), 4) for v in yearly["intensity2"]],
        "intensity3Trend": [round(float(v), 4) for v in yearly["intensity3"]],
        "scope1_source": scope1_source,
        "scope1_fuel": scope1_fuel,
        "scope3_category": scope3_category,
        "plants": plants,
        "scope1_by_plant": [round(float(by_plant.loc[by_plant["PlantName"] == p, "scope1_p"].values[0]), 1) for p in plants],
        "scope2_by_plant": [round(float(by_plant.loc[by_plant["PlantName"] == p, "scope2_p"].values[0]), 1) for p in plants],
    }

# ─── SOCIAL (AGGREGATE) ──────────────────────────────────────────────────────

def _fy_str(start_year):
    """'FY2024-25' for a calendar start year like 2024 - same convention as
    the BRSR annual datasets' FY column."""
    return f"FY{start_year}-{str(start_year + 1)[-2:]}"

def _workforce_snapshot(wf):
    if len(wf) == 0:
        return None
    total_female = float(wf["PermanentFemale"].sum() + wf["ContractualFemale"].sum())
    total_all = float(
        wf["PermanentMale"].sum() + wf["PermanentFemale"].sum() + wf["PermanentOther"].sum() +
        wf["ContractualMale"].sum() + wf["ContractualFemale"].sum() + wf["ContractualOther"].sum()
    )
    new_hires = float(wf["NewHiresMale"].sum() + wf["NewHiresFemale"].sum() + wf["NewHiresOther"].sum())
    turnover = float(wf["TurnoverMale"].sum() + wf["TurnoverFemale"].sum())
    return {
        "employees": int(total_all),
        "female_pct": round(total_female / total_all * 100, 1) if total_all > 0 else 0,
        "new_hire_rate": round(new_hires / total_all * 100, 1) if total_all > 0 else 0,
        "turnover_rate": round(turnover / total_all * 100, 1) if total_all > 0 else 0,
    }

@app.get("/api/social/kpis")
def get_social_kpis(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    sdf_all = load_safety_data()

    # Apply plant/region filters across the full dataset (needed for trend slices)
    sdf_base = safe_filter_plant(sdf_all, plant)
    sdf_base = safe_filter_region(sdf_base, region)

    # Current period values (year=None means all years combined)
    sdf = sdf_base[sdf_base["FiscalReportingPeriod"] == year] if year else sdf_base

    ltifr      = rate_per_basis(sdf["LostTimeInjuries"].sum(), sdf["HoursWorked"].sum())
    trir       = rate_per_basis(sdf["RecordableInjuries"].sum(), sdf["HoursWorked"].sum())
    injuries   = float(sdf["RecordableInjuries"].sum())
    ill_health = float(sdf["RecordableIllHealth"].sum())

    # YoY trend — always computed; falls back to latest year when no year filter
    def pct_change(curr, prev):
        if not prev:
            return None
        return round((curr - prev) / prev * 100, 1)

    trend_year = year if year else int(sdf_base["FiscalReportingPeriod"].max())
    s_t = sdf_base[sdf_base["FiscalReportingPeriod"] == trend_year]
    s_p = sdf_base[sdf_base["FiscalReportingPeriod"] == trend_year - 1]

    trend_ltifr    = pct_change(rate_per_basis(s_t["LostTimeInjuries"].sum(),   s_t["HoursWorked"].sum()),
                                rate_per_basis(s_p["LostTimeInjuries"].sum(),   s_p["HoursWorked"].sum()))
    trend_trir     = pct_change(rate_per_basis(s_t["RecordableInjuries"].sum(), s_t["HoursWorked"].sum()),
                                rate_per_basis(s_p["RecordableInjuries"].sum(), s_p["HoursWorked"].sum()))
    trend_injuries = pct_change(float(s_t["RecordableInjuries"].sum()),  float(s_p["RecordableInjuries"].sum()))
    trend_ill      = pct_change(float(s_t["RecordableIllHealth"].sum()), float(s_p["RecordableIllHealth"].sum()))

    # Workforce/training KPIs (GRI 401/404) - reuse the BRSR annual datasets
    # already loaded for the BRSR tab; these tiles were explicit placeholders
    # because nothing fed them yet, not because the data doesn't exist.
    wfdf_all = safe_filter_region(safe_filter_plant(load_workforce_data(), plant), region)
    trdf_all = safe_filter_region(safe_filter_plant(load_training_data(), plant), region)

    trend_fy = _fy_str(trend_year)
    prev_fy = _fy_str(trend_year - 1)
    if year:
        wf_cur = filter_annual_by_fy(wfdf_all, trend_fy)
    else:
        wf_cur = wfdf_all[wfdf_all["StartYear"] == int(wfdf_all["StartYear"].max())] if len(wfdf_all) else wfdf_all
    wf_prev = filter_annual_by_fy(wfdf_all, prev_fy)

    snap = _workforce_snapshot(wf_cur)
    snap_prev = _workforce_snapshot(wf_prev)
    trend_employees = pct_change(snap["employees"], snap_prev["employees"]) if snap and snap_prev else None
    trend_female = pct_change(snap["female_pct"], snap_prev["female_pct"]) if snap and snap_prev else None
    trend_new_hire = pct_change(snap["new_hire_rate"], snap_prev["new_hire_rate"]) if snap and snap_prev else None
    trend_turnover = pct_change(snap["turnover_rate"], snap_prev["turnover_rate"]) if snap and snap_prev else None

    if year:
        tr_cur = filter_annual_by_fy(trdf_all, trend_fy)
    else:
        tr_cur = trdf_all[trdf_all["StartYear"] == int(trdf_all["StartYear"].max())] if len(trdf_all) else trdf_all
    tr_prev = filter_annual_by_fy(trdf_all, prev_fy)
    avg_hrs = round(float(tr_cur["AvgTrainingHrsAllEmployees"].mean()), 1) if len(tr_cur) else None
    avg_hrs_prev = round(float(tr_prev["AvgTrainingHrsAllEmployees"].mean()), 1) if len(tr_prev) else None
    trend_hrs = pct_change(avg_hrs, avg_hrs_prev) if avg_hrs is not None and avg_hrs_prev else None

    return [
        {"id": "employees",     "label": "Total Employees",           "value": snap["employees"] if snap else None,    "unit": "people",       "trend": trend_employees, "gri": "GRI 2-7",   "status": "live" if snap else "placeholder"},
        {"id": "female-pct",    "label": "Female Employee %",         "value": snap["female_pct"] if snap else None,   "unit": "%",            "trend": trend_female,    "gri": "GRI 2-7",   "status": "live" if snap else "placeholder"},
        {"id": "work-injuries", "label": "Work Related Injuries",     "value": round(injuries, 1), "unit": "",             "trend": trend_injuries, "gri": "GRI 403-9", "status": "live"},
        {"id": "ill-health",    "label": "Work Related Ill Health",   "value": round(ill_health, 1),"unit": "",            "trend": trend_ill,      "gri": "GRI 403-9", "status": "live"},
        {"id": "ltifr",         "label": "LTIFR",                     "value": ltifr,              "unit": "per 200k hrs", "trend": trend_ltifr,   "gri": "GRI 403-9", "status": "live"},
        {"id": "trir",          "label": "TRIR",                      "value": trir,               "unit": "per 200k hrs", "trend": trend_trir,    "gri": "GRI 403-9", "status": "live"},
        {"id": "new-hire-rate", "label": "New Hire Rate",             "value": snap["new_hire_rate"] if snap else None, "unit": "%",            "trend": trend_new_hire, "gri": "GRI 401-1", "status": "live" if snap else "placeholder"},
        {"id": "turnover-rate", "label": "Turnover Rate",             "value": snap["turnover_rate"] if snap else None, "unit": "%",            "trend": trend_turnover, "gri": "GRI 401-1", "status": "live" if snap else "placeholder"},
        {"id": "training-hours","label": "Training Hours / Employee", "value": avg_hrs,            "unit": "hrs",          "trend": trend_hrs,      "gri": "GRI 404-1", "status": "live" if avg_hrs is not None else "placeholder"},
    ]

@app.get("/api/social/safety")
def get_social_safety(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None, view: Optional[str] = "monthly"):
    df_all = load_safety_data()
    df_all = safe_filter_plant(df_all, plant)
    df_all = safe_filter_region(df_all, region)

    # In yearly mode aggregate across all years; otherwise filter to the selected year.
    df = df_all if view == "yearly" else (df_all[df_all["FiscalReportingPeriod"] == year] if year else df_all)

    by_plant = df.groupby("PlantName").agg(
        hours=("HoursWorked", "sum"),
        lostTime=("LostTimeInjuries", "sum"),
        recordable=("RecordableInjuries", "sum"),
    ).reset_index()
    plants = sorted(by_plant["PlantName"].unique().tolist())
    ltifr_by_plant = {row["PlantName"]: rate_per_basis(row["lostTime"], row["hours"]) for _, row in by_plant.iterrows()}
    trir_by_plant = {row["PlantName"]: rate_per_basis(row["recordable"], row["hours"]) for _, row in by_plant.iterrows()}

    by_year = df_all.groupby("FiscalReportingPeriod").agg(
        hours=("HoursWorked", "sum"),
        lostTime=("LostTimeInjuries", "sum"),
        recordable=("RecordableInjuries", "sum"),
    ).reset_index()
    years = sorted(int(y) for y in by_year["FiscalReportingPeriod"].unique().tolist())
    ltifr_trend_map = {int(row["FiscalReportingPeriod"]): rate_per_basis(row["lostTime"], row["hours"]) for _, row in by_year.iterrows()}
    trir_trend_map = {int(row["FiscalReportingPeriod"]): rate_per_basis(row["recordable"], row["hours"]) for _, row in by_year.iterrows()}

    injury_types = explode_weighted(df, "MainInjuryType", "RecordableInjuries")
    injury_result = [{"type": t, "value": round(float(v), 1)} for t, v in injury_types.items()]

    ill_health_types = explode_weighted(df, "MainIllHealthType", "RecordableIllHealth")
    ill_health_result = [{"type": t, "value": round(float(v), 1)} for t, v in ill_health_types.items()]

    severity_result = [
        {"severity": "Fatal", "value": round(float(df["FatalitiesInjury"].sum()), 1)},
        {"severity": "High Consequence", "value": round(float(df["HighConsequenceInjuries"].sum()), 1)},
        {"severity": "Recordable", "value": round(float(df["RecordableInjuries"].sum()), 1)},
        {"severity": "First Aid", "value": round(float(df["FirstAidCases"].sum()), 1)},
    ]

    # Safety pyramid (GRI 403-9/403-10) - near misses sit at the wide base of
    # the classic safety triangle, above fatalities/lost-time/recordable/first
    # aid which are already charted; NearMissesReported has never been used.
    safety_pyramid = {
        "near_miss":   round(float(df["NearMissesReported"].sum()), 1),
        "first_aid":   round(float(df["FirstAidCases"].sum()), 1),
        "recordable":  round(float(df["RecordableInjuries"].sum()), 1),
        "lost_time":   round(float(df["LostTimeInjuries"].sum()), 1),
        "fatal":       round(float(df["FatalitiesInjury"].sum()), 1),
    }

    # Leading indicators (proactive safety activity, GRI 403-5/403-7 context) -
    # SafetyObservations/ToolboxTalks/SafetyInspections are recorded every
    # month per plant but were never aggregated into the dashboard before.
    headcount_sum = float(df["Headcount"].sum())
    leading_indicators = {
        "observations":   round(float(df["SafetyObservations"].sum()), 1),
        "toolbox_talks":  round(float(df["ToolboxTalks"].sum()), 1),
        "inspections":    round(float(df["SafetyInspections"].sum()), 1),
        "training_hours": round(float(df["TrainingHoursSafety"].sum()), 1),
        "training_hours_per_employee": round(float(df["TrainingHoursSafety"].sum()) / headcount_sum, 2) if headcount_sum > 0 else 0,
    }
    leading_by_year = df_all.groupby("FiscalReportingPeriod").agg(
        observations=("SafetyObservations", "sum"),
        toolboxTalks=("ToolboxTalks", "sum"),
        inspections=("SafetyInspections", "sum"),
        trainingHours=("TrainingHoursSafety", "sum"),
        headcount=("Headcount", "sum"),
    ).reset_index()
    leading_trend = {
        "observations":   [round(float(v), 1) for v in leading_by_year["observations"]],
        "toolbox_talks":  [round(float(v), 1) for v in leading_by_year["toolboxTalks"]],
        "inspections":    [round(float(v), 1) for v in leading_by_year["inspections"]],
        "training_hours_per_employee": [
            round(float(h) / float(hc), 2) if hc > 0 else 0
            for h, hc in zip(leading_by_year["trainingHours"], leading_by_year["headcount"])
        ],
    }

    # OHS / audit coverage (GRI 403-8) - % of the workforce covered by the
    # OHS management system, internal audits, and external/third-party audits.
    ohs_coverage = {
        "ohs_pct":             round(float(df["CoveredOHS"].sum()) / headcount_sum * 100, 1) if headcount_sum > 0 else 0,
        "internal_audit_pct": round(float(df["CoveredInternalAudit"].sum()) / headcount_sum * 100, 1) if headcount_sum > 0 else 0,
        "external_audit_pct": round(float(df["CoveredExternalAudit"].sum()) / headcount_sum * 100, 1) if headcount_sum > 0 else 0,
    }

    return {
        "plants": plants,
        "ltifr_by_plant": [ltifr_by_plant.get(p, 0) for p in plants],
        "trir_by_plant": [trir_by_plant.get(p, 0) for p in plants],
        "years": years,
        "ltifr_trend": [ltifr_trend_map.get(y, 0) for y in years],
        "trir_trend": [trir_trend_map.get(y, 0) for y in years],
        "injury_types": injury_result,
        "ill_health_types": ill_health_result,
        "severity": severity_result,
        "safety_pyramid": safety_pyramid,
        "leading_indicators": leading_indicators,
        "leading_trend_years": years,
        "leading_trend": leading_trend,
        "ohs_coverage": ohs_coverage,
    }

@app.get("/api/social/development")
def get_social_development(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    """GRI 401-1 (new hires/turnover), GRI 401-3 (parental leave) and GRI 404-1
    (training hours) trend + by-plant view, built from the BRSR Workforce and
    Training annual datasets - same source data the BRSR tab's own charts use,
    just reframed for the Social tab's GRI Development sub-tab."""
    wfdf_all = safe_filter_region(safe_filter_plant(load_workforce_data(), plant), region)
    trdf_all = safe_filter_region(safe_filter_plant(load_training_data(), plant), region)

    fys = sorted(wfdf_all["FY"].unique().tolist(), key=lambda f: int(f[2:6])) if len(wfdf_all) else []

    new_hire_rate, turnover_rate, maternity_pct, paternity_pct = [], [], [], []
    for fy in fys:
        wf = wfdf_all[wfdf_all["FY"] == fy]
        snap = _workforce_snapshot(wf)
        new_hire_rate.append(snap["new_hire_rate"] if snap else 0)
        turnover_rate.append(snap["turnover_rate"] if snap else 0)
        maternity_pct.append(round(float(wf["MaternityLeavePct"].mean()), 1) if len(wf) else 0)
        paternity_pct.append(round(float(wf["PaternityLeavePct"].mean()), 1) if len(wf) else 0)

    tr_fys = sorted(trdf_all["FY"].unique().tolist(), key=lambda f: int(f[2:6])) if len(trdf_all) else []
    avg_training_hrs = [round(float(trdf_all[trdf_all["FY"] == fy]["AvgTrainingHrsAllEmployees"].mean()), 1) for fy in tr_fys]
    training_coverage_pct = [round(float(trdf_all[trdf_all["FY"] == fy]["TrainingCoveragePct"].mean()), 1) for fy in tr_fys]

    current_fy = fys[-1] if fys else None
    if year:
        candidate = _fy_str(year)
        if candidate in fys:
            current_fy = candidate

    wf_cur = wfdf_all[wfdf_all["FY"] == current_fy] if current_fy else wfdf_all.iloc[:0]
    tr_cur = trdf_all[trdf_all["FY"] == current_fy] if current_fy else trdf_all.iloc[:0]

    plants = sorted(wf_cur["PlantName"].unique().tolist()) if len(wf_cur) else []
    hire_by_plant, turnover_by_plant = [], []
    for p in plants:
        snap = _workforce_snapshot(wf_cur[wf_cur["PlantName"] == p])
        hire_by_plant.append(snap["new_hire_rate"] if snap else 0)
        turnover_by_plant.append(snap["turnover_rate"] if snap else 0)

    tr_plants = sorted(tr_cur["PlantName"].unique().tolist()) if len(tr_cur) else []
    avg_hrs_by_plant = [round(float(tr_cur[tr_cur["PlantName"] == p]["AvgTrainingHrsAllEmployees"].mean()), 1) for p in tr_plants]

    return {
        "fys": fys,
        "new_hire_rate": new_hire_rate,
        "turnover_rate": turnover_rate,
        "maternity_pct": maternity_pct,
        "paternity_pct": paternity_pct,
        "training_fys": tr_fys,
        "avg_training_hrs": avg_training_hrs,
        "training_coverage_pct": training_coverage_pct,
        "current_fy": current_fy,
        "plants": plants,
        "hire_rate_by_plant": hire_by_plant,
        "turnover_rate_by_plant": turnover_by_plant,
        "training_plants": tr_plants,
        "avg_hrs_by_plant": avg_hrs_by_plant,
    }

def filter_by_fy(df, fy: str, date_col: str = "ReportingPeriod"):
    """Filter a monthly DataFrame to an Indian Financial Year window (Apr–Mar).
    fy format: 'FY2024-25' → April 2024 – March 2025.
    Falls back to returning all rows if fy is None or malformed.
    """
    if not fy or not fy.startswith("FY"):
        return df
    try:
        start_year = int(fy[2:6])   # e.g. 2024
    except ValueError:
        return df
    start = pd.Timestamp(f"{start_year}-04-01")
    end   = pd.Timestamp(f"{start_year + 1}-03-31")
    if date_col not in df.columns:
        return df
    return df[(df[date_col] >= start) & (df[date_col] <= end)]


def filter_annual_by_fy(df, fy, fy_col="FY"):
    """Filter annual BRSR datasets (one row per plant per FY) by FY string like 'FY2024-25'."""
    if not fy:
        return df
    return df[df[fy_col] == fy]


def _load_brsr_config():
    config_path = os.path.join(os.path.dirname(__file__), "data", "brsr_config.json")
    if not os.path.exists(config_path):
        return {}
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)


def _build_workforce_kpis(wf, wfdf_all, fy, plant):
    """Compute 2 BRSR workforce KPIs from annual workforce dataset."""
    def pct_change(curr, prev_v):
        if not prev_v:
            return None
        return round((curr - prev_v) / prev_v * 100, 1)

    if len(wf) == 0:
        return [
            {"id": "brsr-female-pct",        "label": "Female Employees %",      "value": None, "unit": "%",   "trend": None, "principle": "P3", "status": "live"},
            {"id": "brsr-differently-abled", "label": "Differently-Abled Count", "value": None, "unit": "ppl", "trend": None, "principle": "P3", "status": "live"},
        ]

    total_female = float(wf["PermanentFemale"].sum() + wf["ContractualFemale"].sum())
    total_all = float(
        wf["PermanentMale"].sum() + wf["PermanentFemale"].sum() + wf["PermanentOther"].sum() +
        wf["ContractualMale"].sum() + wf["ContractualFemale"].sum() + wf["ContractualOther"].sum()
    )
    female_pct = round(total_female / total_all * 100, 1) if total_all > 0 else 0
    da_count = int(wf["DifferentlyAbledPermanent"].sum() + wf["DifferentlyAbledContractual"].sum())

    # YoY trend — previous FY
    def prev_fy_str(fy_str):
        if not fy_str or not fy_str.startswith("FY"):
            return None
        try:
            y = int(fy_str[2:6])
        except ValueError:
            return None
        return f"FY{y-1}-{str(y)[-2:]}"

    pfy = prev_fy_str(fy)
    wf_prev = safe_filter_plant(filter_annual_by_fy(wfdf_all, pfy), plant) if pfy else wfdf_all.iloc[:0]

    if len(wf_prev):
        prev_female = float(wf_prev["PermanentFemale"].sum() + wf_prev["ContractualFemale"].sum())
        prev_all = float(
            wf_prev["PermanentMale"].sum() + wf_prev["PermanentFemale"].sum() + wf_prev["PermanentOther"].sum() +
            wf_prev["ContractualMale"].sum() + wf_prev["ContractualFemale"].sum() + wf_prev["ContractualOther"].sum()
        )
        prev_pct = round(prev_female / prev_all * 100, 1) if prev_all > 0 else 0
        trend_female = pct_change(female_pct, prev_pct)
        trend_da = pct_change(da_count, int(wf_prev["DifferentlyAbledPermanent"].sum() + wf_prev["DifferentlyAbledContractual"].sum()))
    else:
        trend_female = trend_da = None

    return [
        {"id": "brsr-female-pct",        "label": "Female Employees %",      "value": female_pct, "unit": "%",   "trend": trend_female, "principle": "P3", "status": "live"},
        {"id": "brsr-differently-abled", "label": "Differently-Abled Count", "value": da_count,   "unit": "ppl", "trend": trend_da,     "principle": "P3", "status": "live"},
    ]


def _build_training_kpis(tr, trdf_all, fy, plant):
    """Compute 2 BRSR training KPIs from annual training dataset."""
    def pct_change(curr, prev_v):
        if not prev_v:
            return None
        return round((curr - prev_v) / prev_v * 100, 1)

    if len(tr) == 0:
        return [
            {"id": "brsr-training-hours",    "label": "Avg Training Hours/Employee", "value": None, "unit": "hrs", "trend": None, "principle": "P3", "status": "live"},
            {"id": "brsr-training-coverage", "label": "Training Coverage",           "value": None, "unit": "%",   "trend": None, "principle": "P3", "status": "live"},
        ]

    avg_hrs      = round(float(tr["AvgTrainingHrsAllEmployees"].mean()), 1)
    avg_coverage = round(float(tr["TrainingCoveragePct"].mean()), 1)

    def prev_fy_str(fy_str):
        if not fy_str or not fy_str.startswith("FY"):
            return None
        try:
            y = int(fy_str[2:6])
        except ValueError:
            return None
        return f"FY{y-1}-{str(y)[-2:]}"

    pfy = prev_fy_str(fy)
    tr_prev = safe_filter_plant(filter_annual_by_fy(trdf_all, pfy), plant) if pfy else trdf_all.iloc[:0]

    if len(tr_prev):
        prev_hrs = round(float(tr_prev["AvgTrainingHrsAllEmployees"].mean()), 1)
        prev_cov = round(float(tr_prev["TrainingCoveragePct"].mean()), 1)
        trend_hrs = pct_change(avg_hrs, prev_hrs)
        trend_cov = pct_change(avg_coverage, prev_cov)
    else:
        trend_hrs = trend_cov = None

    return [
        {"id": "brsr-training-hours",    "label": "Avg Training Hours/Employee", "value": avg_hrs,      "unit": "hrs", "trend": trend_hrs, "principle": "P3", "status": "live"},
        {"id": "brsr-training-coverage", "label": "Training Coverage",           "value": avg_coverage, "unit": "%",   "trend": trend_cov, "principle": "P3", "status": "live"},
    ]


def _build_csr_compliance_kpis(csr, csrdf_all, fy):
    """Compute CSR spend (P8) and complaint resolution rate (P1) KPIs."""
    def pct_change(curr, prev_v):
        if not prev_v:
            return None
        return round((curr - prev_v) / prev_v * 100, 1)

    # CSR spend — pick first row per FY (ObligationCrore / TotalSpentCrore are FY-level constants)
    if len(csr) > 0:
        csr_spent = round(float(csr["TotalSpentCrore"].iloc[0]), 2)
        # Trend: compare to previous FY
        def prev_fy_str(fy_str):
            if not fy_str or not fy_str.startswith("FY"):
                return None
            try:
                y = int(fy_str[2:6])
            except ValueError:
                return None
            return f"FY{y-1}-{str(y)[-2:]}"
        pfy = prev_fy_str(fy)
        csr_prev = filter_annual_by_fy(csrdf_all, pfy) if pfy else csrdf_all.iloc[:0]
        prev_spent = round(float(csr_prev["TotalSpentCrore"].iloc[0]), 2) if len(csr_prev) > 0 else None
        csr_trend = pct_change(csr_spent, prev_spent)
        csr_status = "live"
    else:
        csr_spent = csr_trend = None
        csr_status = "live"

    # Complaint resolution — from brsr_config.json
    cfg = _load_brsr_config()
    p1 = cfg.get("principles", {}).get("p1", {})
    recv = p1.get("complaints_received")
    resv = p1.get("complaints_resolved")
    if recv is not None and recv > 0:
        complaint_res = round(resv / recv * 100, 1)
        complaint_status = "live"
    else:
        complaint_res = None
        complaint_status = "live"

    return [
        {"id": "brsr-csr-spend",            "label": "CSR Spend",               "value": csr_spent,    "unit": "₹ Cr", "trend": csr_trend, "principle": "P8", "status": csr_status},
        {"id": "brsr-complaint-resolution", "label": "Complaint Resolution Rate","value": complaint_res, "unit": "%",    "trend": None,      "principle": "P1", "status": complaint_status},
    ]


@app.get("/api/brsr/kpis")
def get_brsr_kpis(fy: Optional[str] = None, plant: Optional[str] = None, region: Optional[str] = None):
    """BRSR Essential indicators KPI summary.
    P6 (Environment) — live from GRI datasets filtered by Indian FY window.
    P3 Safety — live from GRI403 filtered by FY window.
    All other principles — placeholder (data collection pending).
    """
    gdf_all  = load_ghg_data()
    edf_all  = load_energy_data()
    wdf_all  = load_water_data()
    wsdf_all = load_waste_data()
    sdf_all  = load_safety_data()
    wfdf_all = load_workforce_data()
    trdf_all = load_training_data()
    csrdf_all= load_csr_data()

    def prep(d, date_col="ReportingPeriod"):
        d = filter_by_fy(d, fy, date_col)
        d = safe_filter_plant(d, plant)
        return safe_filter_region(d, region)

    def prep_annual(d):
        d = filter_annual_by_fy(d, fy)
        return safe_filter_plant(d, plant)

    g   = prep(gdf_all)
    e   = prep(edf_all)
    w   = prep(wdf_all)
    ws  = prep(wsdf_all)
    s   = prep(sdf_all)
    wf  = prep_annual(wfdf_all)
    tr  = prep_annual(trdf_all)
    csr = filter_annual_by_fy(csrdf_all, fy) if fy else csrdf_all

    energy_total   = float(e["TotalEnergyConsumedGJ"].sum())
    ren            = float(e["ElectricityRenewableGJ"].sum())
    non_ren        = float(e["ElectricityNonRenewableGJ"].sum())
    renewable_pct  = round(ren / (ren + non_ren) * 100, 1) if (ren + non_ren) > 0 else 0

    water_withdrawn = float(w["TotalWaterWithdrawn"].sum())
    water_consumed  = float(w["WaterConsumed"].sum())

    scope1 = float(g["Scope1TotaltCO2e"].sum())
    scope2 = float(g["Scope2LocationBasedtCO2e"].sum())

    if "HazardousFlag" in ws.columns and "ValueNumber" in ws.columns:
        haz_gen    = float(ws[ws["HazardousFlag"] == "Hazardous"]["ValueNumber"].sum())
        nonhaz_gen = float(ws[ws["HazardousFlag"] == "Non-hazardous"]["ValueNumber"].sum())
    else:
        haz_gen    = float(ws["TotalHazardousWasteGenerated"].sum()) if "TotalHazardousWasteGenerated" in ws.columns else 0.0
        nonhaz_gen = float(ws["TotalNonHazardousWasteGenerated"].sum()) if "TotalNonHazardousWasteGenerated" in ws.columns else 0.0

    # GRI403 columns: LostTimeInjuries, RecordableInjuries, HoursWorked, FatalitiesInjury
    _hours = float(s["HoursWorked"].sum()) if "HoursWorked" in s.columns else 0
    ltifr_val = rate_per_basis(
        float(s["LostTimeInjuries"].sum()) if "LostTimeInjuries" in s.columns else 0,
        _hours,
    )
    _rec_col = "RecordableInjuries" if "RecordableInjuries" in s.columns else "TotalRecordableIncidents"
    trir_val = rate_per_basis(
        float(s[_rec_col].sum()) if _rec_col in s.columns else 0,
        _hours,
    )
    _fat_col = "FatalitiesInjury" if "FatalitiesInjury" in s.columns else "Fatalities"
    fatalities = int(s[_fat_col].sum()) if _fat_col in s.columns else 0

    # YoY trend — compare this FY to previous FY
    def prev_fy(fy_str):
        if not fy_str or not fy_str.startswith("FY"):
            return None
        try:
            y = int(fy_str[2:6])
        except ValueError:
            return None
        return f"FY{y-1}-{str(y)[-2:]}"

    pfy = prev_fy(fy)

    def prep_prev(d, date_col="ReportingPeriod"):
        d = filter_by_fy(d, pfy, date_col)
        d = safe_filter_plant(d, plant)
        return safe_filter_region(d, region)

    def pct_change(curr, prev_v):
        if not prev_v:
            return None
        return round((curr - prev_v) / prev_v * 100, 1)

    def trend_col(cur_df, prev_df, col, basis_col=None):
        c = float(cur_df[col].sum()) if col in cur_df.columns else 0.0
        p = float(prev_df[col].sum()) if col in prev_df.columns else 0.0
        if basis_col:
            cb = float(cur_df[basis_col].sum()) if basis_col in cur_df.columns else 0.0
            pb = float(prev_df[basis_col].sum()) if basis_col in prev_df.columns else 0.0
            c_rate = rate_per_basis(c, cb)
            p_rate = rate_per_basis(p, pb)
            return pct_change(c_rate, p_rate)
        return pct_change(c, p)

    gp  = prep_prev(gdf_all);  ep = prep_prev(edf_all)
    wp  = prep_prev(wdf_all);  wsp = prep_prev(wsdf_all)
    sp  = prep_prev(sdf_all)

    ep_ren     = float(ep["ElectricityRenewableGJ"].sum()) if "ElectricityRenewableGJ" in ep.columns else 0.0
    ep_non_ren = float(ep["ElectricityNonRenewableGJ"].sum()) if "ElectricityNonRenewableGJ" in ep.columns else 0.0
    ep_ren_pct = round(ep_ren / (ep_ren + ep_non_ren) * 100, 1) if (ep_ren + ep_non_ren) > 0 else 0

    return [
        # P6 Energy — BRSR Essential B1/B2
        {"id": "brsr-energy-consumed", "label": "Total Energy Consumed",  "value": round(energy_total, 1), "unit": "GJ",  "trend": trend_col(e, ep, "TotalEnergyConsumedGJ"),         "principle": "P6", "status": "live"},
        {"id": "brsr-renewable-pct",   "label": "Renewable Energy Share", "value": renewable_pct,          "unit": "%",   "trend": pct_change(renewable_pct, ep_ren_pct),             "principle": "P6", "status": "live"},
        # P6 Water — BRSR Essential C1/C2
        {"id": "brsr-water-withdrawn", "label": "Water Withdrawn",        "value": round(water_withdrawn, 1), "unit": "KL",  "trend": trend_col(w, wp, "TotalWaterWithdrawn"),       "principle": "P6", "status": "live"},
        {"id": "brsr-water-consumed",  "label": "Water Consumed",         "value": round(water_consumed,  1), "unit": "KL",  "trend": trend_col(w, wp, "WaterConsumed"),             "principle": "P6", "status": "live"},
        # P6 GHG — BRSR Essential D1/D2
        {"id": "brsr-scope1-ghg",      "label": "Scope 1 GHG Emissions",  "value": round(scope1, 1), "unit": "tCO₂e", "trend": trend_col(g, gp, "Scope1TotaltCO2e"),               "principle": "P6", "status": "live"},
        {"id": "brsr-scope2-ghg",      "label": "Scope 2 GHG Emissions",  "value": round(scope2, 1), "unit": "tCO₂e", "trend": trend_col(g, gp, "Scope2LocationBasedtCO2e"),       "principle": "P6", "status": "live"},
        # P6 Waste — BRSR Essential E1/E2
        {"id": "brsr-waste-hazardous",    "label": "Hazardous Waste Generated",     "value": round(haz_gen, 2),    "unit": "t", "trend": None, "principle": "P6", "status": "live"},
        {"id": "brsr-waste-nonhazardous", "label": "Non-Hazardous Waste Generated", "value": round(nonhaz_gen, 2), "unit": "t", "trend": None, "principle": "P6", "status": "live"},
        # P3 Safety — live from GRI403
        {"id": "brsr-ltifr",      "label": "LTIFR",       "value": ltifr_val,  "unit": "per 200k hrs", "trend": trend_col(s, sp, "LostTimeInjuries", "HoursWorked"),             "principle": "P3", "status": "live"},
        {"id": "brsr-trir",       "label": "TRIR",        "value": trir_val,   "unit": "per 200k hrs", "trend": trend_col(s, sp, _rec_col, "HoursWorked"),                       "principle": "P3", "status": "live"},
        {"id": "brsr-fatalities", "label": "Fatalities",  "value": fatalities, "unit": "cases",        "trend": trend_col(s, sp, _fat_col),                                      "principle": "P3", "status": "live"},
        # P3 Workforce — live from BRSR_Workforce_Dataset
        *_build_workforce_kpis(wf, wfdf_all, fy, plant),
        # P3 Training — live from BRSR_Training_Dataset
        *_build_training_kpis(tr, trdf_all, fy, plant),
        # P8 CSR / P1 Compliance — live from BRSR_CSR_Dataset + brsr_config.json
        *_build_csr_compliance_kpis(csr, csrdf_all, fy),
    ]


@app.get("/api/brsr/config")
def get_brsr_config():
    """Serve the brsr_config.json qualitative disclosure template to the frontend."""
    return _load_brsr_config()


@app.get("/api/brsr/workforce")
def get_brsr_workforce(fy: Optional[str] = None, plant: Optional[str] = None, region: Optional[str] = None):
    """BRSR P3 workforce chart data — headcount by gender, female % trend, DA count, benefits."""
    df_all = load_workforce_data()

    ALL_FYS = ["FY2019-20", "FY2020-21", "FY2021-22", "FY2022-23", "FY2023-24", "FY2024-25"]

    def _plant_filter(d):
        return safe_filter_plant(d, plant)

    headcount_male, headcount_female, headcount_other = [], [], []
    female_pct_trend, da_trend = [], []
    turnover_rate_male, turnover_rate_female, wage_gap_pct = [], [], []

    for f in ALL_FYS:
        sl = _plant_filter(filter_annual_by_fy(df_all, f))
        if len(sl) == 0:
            for lst in [headcount_male, headcount_female, headcount_other, female_pct_trend,
                        da_trend, turnover_rate_male, turnover_rate_female, wage_gap_pct]:
                lst.append(None)
            continue
        pm = float(sl["PermanentMale"].sum())
        pf = float(sl["PermanentFemale"].sum())
        po = float(sl["PermanentOther"].sum())
        cm_ = float(sl["ContractualMale"].sum())
        cf  = float(sl["ContractualFemale"].sum())
        co  = float(sl["ContractualOther"].sum())
        total = pm + pf + po + cm_ + cf + co
        total_female = pf + cf
        headcount_male.append(round(pm + cm_))
        headcount_female.append(round(total_female))
        headcount_other.append(round(po + co))
        female_pct_trend.append(round(total_female / total * 100, 1) if total > 0 else 0)
        da_trend.append(int(sl["DifferentlyAbledPermanent"].sum() + sl["DifferentlyAbledContractual"].sum()))
        # Turnover rates
        tm = float(sl["TurnoverMale"].sum())
        tf_ = float(sl["TurnoverFemale"].sum())
        turnover_rate_male.append(round(tm / (pm + cm_) * 100, 1) if (pm + cm_) > 0 else 0)
        turnover_rate_female.append(round(tf_ / (pf + cf) * 100, 1) if (pf + cf) > 0 else 0)
        # Wage gap
        wm = float(sl["AvgWagePermanentMaleINR"].mean())
        wf_ = float(sl["AvgWagePermanentFemaleINR"].mean())
        gap = round((1 - wf_ / wm) * 100, 1) if wm > 0 else 0
        wage_gap_pct.append(gap)

    # By-plant breakdown for latest FY
    latest_fy = fy if fy else ALL_FYS[-1]
    latest_sl = _plant_filter(filter_annual_by_fy(df_all, latest_fy))
    plants_list = sorted(latest_sl["PlantName"].unique().tolist()) if len(latest_sl) > 0 else []
    female_pct_by_plant = []
    for p in plants_list:
        row = latest_sl[latest_sl["PlantName"] == p]
        tot_f = float(row["PermanentFemale"].sum() + row["ContractualFemale"].sum())
        tot   = float((row[["PermanentMale","PermanentFemale","PermanentOther",
                             "ContractualMale","ContractualFemale","ContractualOther"]]).sum().sum())
        female_pct_by_plant.append(round(tot_f / tot * 100, 1) if tot > 0 else 0)

    # Benefits (latest FY, mean across plants)
    benefits = {}
    if len(latest_sl) > 0:
        benefits = {
            "health_insurance_perm":    100,
            "health_insurance_contract": round(float(latest_sl["HealthInsuranceContractualPct"].mean()), 1),
            "accident_insurance":       100,
            "maternity_leave":          100,
            "paternity_leave":          round(float(latest_sl["PaternityLeavePct"].mean()), 1),
            "pf_covered":               100,
        }

    return {
        "fys": ALL_FYS,
        "headcount_male": headcount_male,
        "headcount_female": headcount_female,
        "headcount_other": headcount_other,
        "female_pct": female_pct_trend,
        "differently_abled": da_trend,
        "turnover_rate_male": turnover_rate_male,
        "turnover_rate_female": turnover_rate_female,
        "wage_gap_pct": wage_gap_pct,
        "plants": plants_list,
        "female_pct_by_plant": female_pct_by_plant,
        "benefits": benefits,
        "current_fy": latest_fy,
    }


@app.get("/api/brsr/training")
def get_brsr_training(fy: Optional[str] = None, plant: Optional[str] = None, region: Optional[str] = None):
    """BRSR P3 training chart data — avg hours, coverage, skill upgrade, spend trend."""
    df_all = load_training_data()

    ALL_FYS = ["FY2019-20", "FY2020-21", "FY2021-22", "FY2022-23", "FY2023-24", "FY2024-25"]

    def _plant_filter(d):
        return safe_filter_plant(d, plant)

    avg_hrs_all, avg_hrs_male, avg_hrs_female = [], [], []
    coverage_trend, skill_upgrade_trend, perf_review_trend = [], [], []
    spend_per_emp_trend = []

    for f in ALL_FYS:
        sl = _plant_filter(filter_annual_by_fy(df_all, f))
        if len(sl) == 0:
            for lst in [avg_hrs_all, avg_hrs_male, avg_hrs_female, coverage_trend,
                        skill_upgrade_trend, perf_review_trend, spend_per_emp_trend]:
                lst.append(None)
            continue
        avg_hrs_all.append(round(float(sl["AvgTrainingHrsAllEmployees"].mean()), 1))
        avg_hrs_male.append(round(float(sl["AvgTrainingHrsPerEmployeeMale"].mean()), 1))
        avg_hrs_female.append(round(float(sl["AvgTrainingHrsPerEmployeeFemale"].mean()), 1))
        coverage_trend.append(round(float(sl["TrainingCoveragePct"].mean()), 1))
        skill_upgrade_trend.append(round(float(sl["SkillUpgradePct"].mean()), 1))
        perf_review_trend.append(round(float(sl["PerformanceReviewCoveragePct"].mean()), 1))
        spend_per_emp_trend.append(round(float(sl["TrainingSpendPerEmployeeINR"].mean())))

    # By-plant for latest FY
    latest_fy = fy if fy else ALL_FYS[-1]
    latest_sl = _plant_filter(filter_annual_by_fy(df_all, latest_fy))
    plants_list = sorted(latest_sl["PlantName"].unique().tolist()) if len(latest_sl) > 0 else []
    avg_hrs_by_plant   = [round(float(latest_sl[latest_sl["PlantName"] == p]["AvgTrainingHrsAllEmployees"].mean()), 1) for p in plants_list]
    coverage_by_plant  = [round(float(latest_sl[latest_sl["PlantName"] == p]["TrainingCoveragePct"].mean()), 1) for p in plants_list]

    # Training type breakdown (latest FY average across plants)
    training_type = {}
    if len(latest_sl) > 0:
        training_type = {
            "Safety":     round(float(latest_sl["SafetyTrainingHrsPermanent"].sum() / len(latest_sl)), 0),
            "Technical":  round(float(latest_sl["TechnicalSkillsHrsPerEmp"].mean()), 1),
            "Leadership": round(float(latest_sl["LeadershipHrsPerEmp"].mean()), 1),
            "Compliance": round(float(latest_sl["ComplianceHrsPerEmp"].mean()), 1),
            "Soft Skills":round(float(latest_sl["SoftSkillsHrsPerEmp"].mean()), 1),
        }

    return {
        "fys": ALL_FYS,
        "avg_training_hrs": avg_hrs_all,
        "avg_training_hrs_male": avg_hrs_male,
        "avg_training_hrs_female": avg_hrs_female,
        "training_coverage_pct": coverage_trend,
        "skill_upgrade_pct": skill_upgrade_trend,
        "performance_review_pct": perf_review_trend,
        "training_spend_per_employee": spend_per_emp_trend,
        "plants": plants_list,
        "avg_hrs_by_plant": avg_hrs_by_plant,
        "coverage_by_plant": coverage_by_plant,
        "training_type_breakdown": training_type,
        "current_fy": latest_fy,
    }


@app.get("/api/brsr/csr")
def get_brsr_csr(fy: Optional[str] = None):
    """BRSR P8 CSR chart data — obligation vs spend, category breakdown, beneficiaries."""
    df_all = load_csr_data()

    ALL_FYS = ["FY2019-20", "FY2020-21", "FY2021-22", "FY2022-23", "FY2023-24", "FY2024-25"]
    obligation_trend, spent_trend, beneficiary_trend = [], [], []

    for f in ALL_FYS:
        sl = df_all[df_all["FY"] == f]
        if len(sl) == 0:
            obligation_trend.append(None)
            spent_trend.append(None)
            beneficiary_trend.append(None)
        else:
            obligation_trend.append(round(float(sl["ObligationCrore"].iloc[0]), 2))
            spent_trend.append(round(float(sl["TotalSpentCrore"].iloc[0]), 2))
            beneficiary_trend.append(int(sl["BeneficiaryCount"].sum()))

    # Category breakdown for selected/latest FY
    latest_fy = fy if fy else ALL_FYS[-1]
    latest_sl = df_all[df_all["FY"] == latest_fy]

    categories, spent_by_cat, bene_by_cat = [], [], []
    if len(latest_sl) > 0:
        for _, row in latest_sl.iterrows():
            categories.append(row["ProjectCategory"])
            spent_by_cat.append(round(float(row["SpentCrore"]), 2))
            bene_by_cat.append(int(row["BeneficiaryCount"]))

    current_fy_data = {}
    if len(latest_sl) > 0:
        obligation = round(float(latest_sl["ObligationCrore"].iloc[0]), 2)
        spent = round(float(latest_sl["TotalSpentCrore"].iloc[0]), 2)
        current_fy_data = {
            "obligation": obligation,
            "spent": spent,
            "unspent": round(float(latest_sl["UnspentCrore"].iloc[0]), 2),
            "pct_spent": round(spent / obligation * 100, 1) if obligation > 0 else 0,
            "total_beneficiaries": int(latest_sl["BeneficiaryCount"].sum()),
            "states": latest_sl["LocationState"].iloc[0] if len(latest_sl) > 0 else "",
        }

    return {
        "fys": ALL_FYS,
        "obligation_crore": obligation_trend,
        "spent_crore": spent_trend,
        "total_beneficiaries": beneficiary_trend,
        "categories": categories,
        "spent_by_category": spent_by_cat,
        "beneficiaries_by_category": bene_by_cat,
        "current_fy": current_fy_data,
        "selected_fy": latest_fy,
    }


class BrsrReportRequest(BaseModel):
    fy: Optional[str] = None
    plant: Optional[str] = None
    format: Optional[str] = "pdf"


@app.post("/api/brsr/reports/generate")
def generate_brsr_report(req: BrsrReportRequest):
    """Generate a BRSR disclosure report (PDF or Excel) covering all 9 Principles."""
    common_kwargs = dict(
        fy=req.fy,
        plant=req.plant,
        filter_by_fy_fn=filter_by_fy,
        filter_annual_by_fy_fn=filter_annual_by_fy,
        load_energy_fn=load_energy_data,
        load_ghg_fn=load_ghg_data,
        load_water_fn=load_water_data,
        load_waste_fn=load_waste_data,
        load_safety_fn=load_safety_data,
        load_workforce_fn=load_workforce_data,
        load_training_fn=load_training_data,
        load_csr_fn=load_csr_data,
    )
    fy_label = req.fy or "all"
    if req.format == "pdf":
        from brsr_report import generate_brsr_pdf
        buf = generate_brsr_pdf(**common_kwargs)
        filename = f"esg_brsr_report_{fy_label}.pdf"
        content = buf.getvalue()
        storage.save_report(content, filename, framework="BRSR", format="pdf", fy=req.fy, plant=req.plant)
        return StreamingResponse(io.BytesIO(content), media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename={filename}"})
    elif req.format == "excel":
        from brsr_report import generate_brsr_excel
        buf = generate_brsr_excel(**common_kwargs)
        filename = f"esg_brsr_report_{fy_label}.xlsx"
        content = buf.getvalue()
        storage.save_report(content, filename, framework="BRSR", format="excel", fy=req.fy, plant=req.plant)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        raise HTTPException(status_code=400, detail="BRSR reports support 'pdf' and 'excel' formats only.")


@app.get("/api/sasb/kpis")
def get_sasb_kpis(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    """
    SASB RT-CH (Chemicals) aggregate KPI endpoint. Mirrors the response shape
    of /api/environment/kpis + /api/social/kpis (id/label/value/unit/trend/
    status) but grouped by SASB's own topic structure rather than GRI's.
    Computed from the SAME 5 GRI datasets wherever the topic overlaps
    (GHG, Air Quality, Energy, Water, Hazardous Waste, Workforce Safety) -
    see SASB_INTEGRATION_PLAN.md Section 3. Process Safety (RT-CH-540a) is
    sourced from a dedicated synthetic dataset (see
    generate_process_safety_dataset.py) since no real process-safety
    incident data exists for this demo.
    """
    gdf_all = load_ghg_data()
    edf_all = load_energy_data()
    wdf_all = load_water_data()
    wsdf_all = load_waste_data()
    sdf_all = load_safety_data()
    psdf_all = load_process_safety_data()

    def slice_df(d, yr, year_col):
        if yr:
            d = d[d[year_col] == yr]
        d = safe_filter_plant(d, plant)
        return safe_filter_region(d, region)

    g_cur  = slice_df(gdf_all,  year, "ReportingYear")
    e_cur  = slice_df(edf_all,  year, "ReportingYear")
    w_cur  = slice_df(wdf_all,  year, "ReportingYear")
    ws_cur = slice_df(wsdf_all, year, "FiscalReportingPeriod")
    s_cur  = slice_df(sdf_all,  year, "FiscalReportingPeriod")
    ps_cur = slice_df(psdf_all, year, "FiscalReportingPeriod")

    # GHG & Air Quality — RT-CH-110a, RT-CH-120a
    scope1 = float(g_cur["Scope1TotaltCO2e"].sum())
    nox = float(g_cur["GrossNOxt"].sum())
    sox = float(g_cur["GrossSOxt"].sum())
    voc = float(g_cur["GrossVOCt"].sum())
    pm  = float(g_cur["GrossPMt"].sum())

    # Energy Management — RT-CH-130a
    energy_consumed = float(e_cur["TotalEnergyConsumedGJ"].sum())
    renewable     = float(e_cur["ElectricityRenewableGJ"].sum())
    non_renewable = float(e_cur["ElectricityNonRenewableGJ"].sum())
    renewable_pct = round(renewable / (renewable + non_renewable) * 100, 1) if (renewable + non_renewable) > 0 else 0

    # Water Management — RT-CH-140a
    withdrawn = float(w_cur["TotalWaterWithdrawn"].sum())
    consumed  = float(w_cur["WaterConsumed"].sum())
    stress    = float(w_cur["TotalWaterWithdrawnStressArea"].sum())
    stress_pct = round(stress / withdrawn * 100, 1) if withdrawn > 0 else 0

    # Hazardous Waste Management — RT-CH-150a
    haz = hazardous_waste_breakdown(ws_cur)

    # GHG intensity + market-based Scope 2 + Scope 3 — production-normalized
    # disclosures SASB RT-CH-110a.2 calls for, reusing the same per-tonne
    # weighted-ratio convention as the GRI 305 intensity calculations.
    scope1_intensity = weighted_ratio(g_cur, "Scope1TotaltCO2e")
    scope2_market = float(g_cur["Scope2MarketBasedtCO2e"].sum())
    scope3_total = float(g_cur["Scope3TotaltCO2e"].sum())

    # Workforce Health & Safety — RT-CH-320a
    trir_total      = safety_rate_split(s_cur)
    trir_employee   = safety_rate_split(s_cur, "Employee")
    trir_contractor = safety_rate_split(s_cur, "Contractor")
    fat_rate        = fatality_rate(s_cur)

    # Process Safety — RT-CH-540a
    ps = process_safety_rates(ps_cur)

    # YoY trend — always computed (same convention as the GRI aggregate endpoints)
    trend_year = year if year else int(gdf_all["ReportingYear"].max())

    g_t, g_p   = slice_df(gdf_all,  trend_year, "ReportingYear"), slice_df(gdf_all,  trend_year - 1, "ReportingYear")
    e_t, e_p   = slice_df(edf_all,  trend_year, "ReportingYear"), slice_df(edf_all,  trend_year - 1, "ReportingYear")
    w_t, w_p   = slice_df(wdf_all,  trend_year, "ReportingYear"), slice_df(wdf_all,  trend_year - 1, "ReportingYear")
    ws_t, ws_p = slice_df(wsdf_all, trend_year, "FiscalReportingPeriod"), slice_df(wsdf_all, trend_year - 1, "FiscalReportingPeriod")
    s_t, s_p   = slice_df(sdf_all,  trend_year, "FiscalReportingPeriod"), slice_df(sdf_all,  trend_year - 1, "FiscalReportingPeriod")
    ps_t, ps_p = slice_df(psdf_all, trend_year, "FiscalReportingPeriod"), slice_df(psdf_all, trend_year - 1, "FiscalReportingPeriod")

    def pct_change(curr, prev):
        if not prev:
            return None
        return round((curr - prev) / prev * 100, 1)

    haz_t, haz_p = hazardous_waste_breakdown(ws_t), hazardous_waste_breakdown(ws_p)
    ps_t_rates, ps_p_rates = process_safety_rates(ps_t), process_safety_rates(ps_p)

    wt_withdrawn = float(w_t["TotalWaterWithdrawn"].sum())
    wp_withdrawn = float(w_p["TotalWaterWithdrawn"].sum())
    wt_consumed  = float(w_t["WaterConsumed"].sum())
    wp_consumed  = float(w_p["WaterConsumed"].sum())
    wt_stress_raw = float(w_t["TotalWaterWithdrawnStressArea"].sum())
    wt_stress_pct = round(wt_stress_raw / wt_withdrawn * 100, 1) if wt_withdrawn > 0 else 0
    wp_stress_raw = float(w_p["TotalWaterWithdrawnStressArea"].sum())
    wp_stress_pct = round(wp_stress_raw / wp_withdrawn * 100, 1) if wp_withdrawn > 0 else 0

    et_renewable     = float(e_t["ElectricityRenewableGJ"].sum())
    ep_renewable     = float(e_p["ElectricityRenewableGJ"].sum())
    et_non_renewable = float(e_t["ElectricityNonRenewableGJ"].sum())
    ep_non_renewable = float(e_p["ElectricityNonRenewableGJ"].sum())
    et_renewable_pct = round(et_renewable / (et_renewable + et_non_renewable) * 100, 1) if (et_renewable + et_non_renewable) > 0 else 0
    ep_renewable_pct = round(ep_renewable / (ep_renewable + ep_non_renewable) * 100, 1) if (ep_renewable + ep_non_renewable) > 0 else 0

    trend = {
        "scope1": pct_change(float(g_t["Scope1TotaltCO2e"].sum()), float(g_p["Scope1TotaltCO2e"].sum())),
        "nox":    pct_change(float(g_t["GrossNOxt"].sum()), float(g_p["GrossNOxt"].sum())),
        "sox":    pct_change(float(g_t["GrossSOxt"].sum()), float(g_p["GrossSOxt"].sum())),
        "voc":    pct_change(float(g_t["GrossVOCt"].sum()), float(g_p["GrossVOCt"].sum())),
        "pm":     pct_change(float(g_t["GrossPMt"].sum()), float(g_p["GrossPMt"].sum())),
        "energy":         pct_change(float(e_t["TotalEnergyConsumedGJ"].sum()), float(e_p["TotalEnergyConsumedGJ"].sum())),
        "renewable":      pct_change(et_renewable, ep_renewable),
        "non_renewable":  pct_change(et_non_renewable, ep_non_renewable),
        "renewable_pct":  pct_change(et_renewable_pct, ep_renewable_pct),
        "withdrawn":  pct_change(wt_withdrawn, wp_withdrawn),
        "consumed":   pct_change(wt_consumed, wp_consumed),
        "stress_pct": pct_change(wt_stress_pct, wp_stress_pct),
        "haz_generated":    pct_change(haz_t["haz_generated"], haz_p["haz_generated"]),
        "nonhaz_generated": pct_change(haz_t["nonhaz_generated"], haz_p["nonhaz_generated"]),
        "haz_recycled_pct": pct_change(haz_t["haz_recycled_pct"], haz_p["haz_recycled_pct"]),
        "scope1_intensity": pct_change(weighted_ratio(g_t, "Scope1TotaltCO2e"), weighted_ratio(g_p, "Scope1TotaltCO2e")),
        "scope2_market":    pct_change(float(g_t["Scope2MarketBasedtCO2e"].sum()), float(g_p["Scope2MarketBasedtCO2e"].sum())),
        "scope3_total":     pct_change(float(g_t["Scope3TotaltCO2e"].sum()), float(g_p["Scope3TotaltCO2e"].sum())),
        "trir_total":      pct_change(safety_rate_split(s_t), safety_rate_split(s_p)),
        "trir_employee":   pct_change(safety_rate_split(s_t, "Employee"), safety_rate_split(s_p, "Employee")),
        "trir_contractor": pct_change(safety_rate_split(s_t, "Contractor"), safety_rate_split(s_p, "Contractor")),
        "fatality_rate":   pct_change(fatality_rate(s_t), fatality_rate(s_p)),
        "ps_incidents": pct_change(ps_t_rates["incidents"], ps_p_rates["incidents"]),
        "pstir":        pct_change(ps_t_rates["pstir"], ps_p_rates["pstir"]),
        "psisr":        pct_change(ps_t_rates["psisr"], ps_p_rates["psisr"]),
    }

    withdrawn_ml = round(withdrawn / 1000, 3)
    consumed_ml  = round(consumed / 1000, 3)

    return [
        # GHG & Air Quality — RT-CH-110a, RT-CH-120a
        {"id": "sasb-scope1-ghg", "label": "Scope 1 GHG Emissions",   "value": round(scope1, 1), "unit": "tCO₂e", "trend": trend["scope1"], "sasb": "RT-CH-110a.1", "status": "live"},
        {"id": "sasb-nox",        "label": "NOx Emissions",           "value": round(nox, 2),    "unit": "t",     "trend": trend["nox"],    "sasb": "RT-CH-120a.1", "status": "live"},
        {"id": "sasb-sox",        "label": "SOx Emissions",           "value": round(sox, 2),    "unit": "t",     "trend": trend["sox"],    "sasb": "RT-CH-120a.1", "status": "live"},
        {"id": "sasb-voc",        "label": "VOC Emissions",           "value": round(voc, 2),    "unit": "t",     "trend": trend["voc"],    "sasb": "RT-CH-120a.1", "status": "live"},
        {"id": "sasb-pm",         "label": "Particulate Matter (PM)", "value": round(pm, 2),     "unit": "t",     "trend": trend["pm"],     "sasb": "RT-CH-120a.1", "status": "live"},
        {"id": "sasb-scope1-intensity", "label": "Scope 1 Intensity",          "value": round(scope1_intensity, 3), "unit": "tCO₂e/t product", "trend": trend["scope1_intensity"], "sasb": "RT-CH-110a.2", "status": "live"},
        # Scope2MarketBasedtCO2e is entirely unpopulated in the source GHG dataset (0/420 rows) -
        # genuine data gap, not a bug; flagged "placeholder" rather than reporting a misleading 0.
        {"id": "sasb-scope2-market",    "label": "Scope 2 (Market-Based)",     "value": round(scope2_market, 1) if g_cur["Scope2MarketBasedtCO2e"].notna().any() else None, "unit": "tCO₂e", "trend": trend["scope2_market"], "sasb": "RT-CH-110a.1", "status": "live" if g_cur["Scope2MarketBasedtCO2e"].notna().any() else "placeholder"},
        {"id": "sasb-scope3-total",     "label": "Scope 3 Emissions (Total)",  "value": round(scope3_total, 1),     "unit": "tCO₂e",            "trend": trend["scope3_total"],     "sasb": "RT-CH-110a.1", "status": "live"},
        # Energy Management — RT-CH-130a
        {"id": "sasb-energy-consumed",      "label": "Total Energy Consumed",  "value": round(energy_consumed, 1), "unit": "GJ", "trend": trend["energy"],        "sasb": "RT-CH-130a.1", "status": "live"},
        {"id": "sasb-renewable-energy",     "label": "Renewable Energy",       "value": round(renewable, 1),       "unit": "GJ", "trend": trend["renewable"],     "sasb": "RT-CH-130a.1", "status": "live"},
        {"id": "sasb-non-renewable-energy", "label": "Non-Renewable Energy",   "value": round(non_renewable, 1),   "unit": "GJ", "trend": trend["non_renewable"], "sasb": "RT-CH-130a.1", "status": "live"},
        {"id": "sasb-renewable-pct",        "label": "Renewable Energy %",     "value": renewable_pct,             "unit": "%",  "trend": trend["renewable_pct"], "sasb": "RT-CH-130a.1", "status": "live"},
        # Water Management — RT-CH-140a (stress-area % is self-reported, not yet confirmed WRI Aqueduct-aligned — see plan Section 2)
        {"id": "sasb-water-withdrawn",  "label": "Total Water Withdrawn", "value": withdrawn_ml, "unit": "ML", "trend": trend["withdrawn"],  "sasb": "RT-CH-140a.1", "status": "live"},
        {"id": "sasb-water-consumed",   "label": "Total Water Consumed",  "value": consumed_ml,  "unit": "ML", "trend": trend["consumed"],   "sasb": "RT-CH-140a.1", "status": "live"},
        {"id": "sasb-water-stress-pct", "label": "Water in Stress Areas", "value": stress_pct,   "unit": "%",  "trend": trend["stress_pct"], "sasb": "RT-CH-140a.1", "status": "partial"},
        # Hazardous Waste Management — RT-CH-150a
        {"id": "sasb-hazardous-waste",       "label": "Hazardous Waste Generated",     "value": round(haz["haz_generated"], 1),    "unit": "tonnes", "trend": trend["haz_generated"],    "sasb": "RT-CH-150a.1", "status": "live"},
        {"id": "sasb-nonhazardous-waste",    "label": "Non-Hazardous Waste Generated", "value": round(haz["nonhaz_generated"], 1), "unit": "tonnes", "trend": trend["nonhaz_generated"],  "sasb": "RT-CH-150a.1", "status": "live"},
        {"id": "sasb-hazardous-recycled-pct","label": "Hazardous Waste Recycled %",    "value": haz["haz_recycled_pct"],           "unit": "%",      "trend": trend["haz_recycled_pct"], "sasb": "RT-CH-150a.1", "status": "live"},
        # Workforce Health & Safety — RT-CH-320a
        {"id": "sasb-trir-total",      "label": "TRIR (All Workers)", "value": trir_total,      "unit": "per 200k hrs", "trend": trend["trir_total"],      "sasb": "RT-CH-320a.1", "status": "live"},
        {"id": "sasb-trir-employee",   "label": "TRIR (Employees)",   "value": trir_employee,   "unit": "per 200k hrs", "trend": trend["trir_employee"],   "sasb": "RT-CH-320a.1", "status": "live"},
        {"id": "sasb-trir-contractor", "label": "TRIR (Contractors)", "value": trir_contractor, "unit": "per 200k hrs", "trend": trend["trir_contractor"], "sasb": "RT-CH-320a.1", "status": "live"},
        {"id": "sasb-fatality-rate",   "label": "Fatality Rate",      "value": fat_rate,        "unit": "per 200k hrs", "trend": trend["fatality_rate"],   "sasb": "RT-CH-320a.1", "status": "live"},
        # Process Safety — RT-CH-540a (synthetic dataset — see generate_process_safety_dataset.py)
        {"id": "sasb-process-safety-incidents", "label": "Process Safety Incidents",              "value": ps["incidents"], "unit": "count",        "trend": trend["ps_incidents"], "sasb": "RT-CH-540a.1", "status": "live"},
        {"id": "sasb-pstir",                     "label": "Process Safety Total Incident Rate",    "value": ps["pstir"],     "unit": "per 200k hrs", "trend": trend["pstir"],        "sasb": "RT-CH-540a.1", "status": "live"},
        {"id": "sasb-psisr",                     "label": "Process Safety Incident Severity Rate", "value": ps["psisr"],     "unit": "per 200k hrs", "trend": trend["psisr"],        "sasb": "RT-CH-540a.2", "status": "live"},
    ]


@app.get("/api/sasb/hazardous-waste")
def get_sasb_hazardous_waste(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None, view: Optional[str] = "monthly"):
    """
    Hazardous/Non-hazardous waste breakdown for the SASB RT-CH-150a dashboard.
    Reuses the same HazardousFlag column already used by pdf_report.py's
    pivot_haz() for the GRI 306-4/306-5 hazardous split tables - this is a
    new framing of existing data, not a new dataset.
    """
    df_all = load_waste_data()
    df_all = safe_filter_plant(df_all, plant)
    df_all = safe_filter_region(df_all, region)
    MONTHS = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

    if view == "yearly":
        years = sorted(int(y) for y in df_all["FiscalReportingPeriod"].unique())
        haz_by_year, nonhaz_by_year = [], []
        for y in years:
            b = hazardous_waste_breakdown(df_all[df_all["FiscalReportingPeriod"] == y])
            haz_by_year.append(round(b["haz_generated"], 1))
            nonhaz_by_year.append(round(b["nonhaz_generated"], 1))
        plants = sorted(df_all["PlantName"].unique().tolist())
        by_plant = {p: hazardous_waste_breakdown(df_all[df_all["PlantName"] == p]) for p in plants}
        total = hazardous_waste_breakdown(df_all)
        haz_disposal_methods, _ = disposal_method_breakdown(df_all[df_all["HazardousFlag"] == "Hazardous"])
        return {
            "view": "yearly",
            "labels": years,
            "hazardous": haz_by_year,
            "nonHazardous": nonhaz_by_year,
            "plants": plants,
            "hazardous_by_plant": [round(by_plant[p]["haz_generated"], 1) for p in plants],
            "nonhazardous_by_plant": [round(by_plant[p]["nonhaz_generated"], 1) for p in plants],
            "total_hazardous": round(total["haz_generated"], 1),
            "total_nonhazardous": round(total["nonhaz_generated"], 1),
            "hazardous_recycled_pct": total["haz_recycled_pct"],
            "hazardous_disposal_methods": haz_disposal_methods,
        }

    if not year:
        year = int(df_all["FiscalReportingPeriod"].max())
    df = df_all[df_all["FiscalReportingPeriod"] == year].copy()
    df["month"] = df["ReportingPeriod"].dt.month

    months_present = sorted(df["month"].unique())
    haz_by_month, nonhaz_by_month = [], []
    for m in months_present:
        b = hazardous_waste_breakdown(df[df["month"] == m])
        haz_by_month.append(round(b["haz_generated"], 1))
        nonhaz_by_month.append(round(b["nonhaz_generated"], 1))

    plants = sorted(df["PlantName"].unique().tolist())
    by_plant = {p: hazardous_waste_breakdown(df[df["PlantName"] == p]) for p in plants}
    total = hazardous_waste_breakdown(df)
    haz_disposal_methods, _ = disposal_method_breakdown(df[df["HazardousFlag"] == "Hazardous"])

    return {
        "year": year,
        "months": [MONTHS.get(int(m), str(int(m))) for m in months_present],
        "hazardous": haz_by_month,
        "nonHazardous": nonhaz_by_month,
        "plants": plants,
        "hazardous_by_plant": [round(by_plant[p]["haz_generated"], 1) for p in plants],
        "nonhazardous_by_plant": [round(by_plant[p]["nonhaz_generated"], 1) for p in plants],
        "total_hazardous": round(total["haz_generated"], 1),
        "total_nonhazardous": round(total["nonhaz_generated"], 1),
        "hazardous_recycled_pct": total["haz_recycled_pct"],
        "hazardous_disposal_methods": haz_disposal_methods,
    }


@app.get("/api/sasb/process-safety")
def get_sasb_process_safety(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None, view: Optional[str] = "monthly"):
    """Process Safety trend + by-plant view for the SASB RT-CH-540a dashboard.
    Sourced from the synthetic GRI_RTCH540a_ProcessSafety_Dataset (see
    generate_process_safety_dataset.py) since no real process-safety
    incident data exists for this demo."""
    df_all = load_process_safety_data()
    df_all = safe_filter_plant(df_all, plant)
    df_all = safe_filter_region(df_all, region)
    MONTHS = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

    if view == "yearly":
        years = sorted(int(y) for y in df_all["FiscalReportingPeriod"].unique())
        tier1_by_year, tier2_by_year, pstir_by_year, psisr_by_year = [], [], [], []
        for y in years:
            r = process_safety_rates(df_all[df_all["FiscalReportingPeriod"] == y])
            tier1_by_year.append(r["tier1"])
            tier2_by_year.append(r["tier2"])
            pstir_by_year.append(r["pstir"])
            psisr_by_year.append(r["psisr"])
        plants = sorted(df_all["PlantName"].unique().tolist())
        by_plant = {p: process_safety_rates(df_all[df_all["PlantName"] == p]) for p in plants}
        total = process_safety_rates(df_all)
        return {
            "view": "yearly",
            "labels": years,
            "tier1": tier1_by_year,
            "tier2": tier2_by_year,
            "pstir": pstir_by_year,
            "psisr": psisr_by_year,
            "plants": plants,
            "pstir_by_plant": [by_plant[p]["pstir"] for p in plants],
            "psisr_by_plant": [by_plant[p]["psisr"] for p in plants],
            "total_incidents": total["incidents"],
            "total_tier1": total["tier1"],
            "total_tier2": total["tier2"],
            "total_pstir": total["pstir"],
            "total_psisr": total["psisr"],
        }

    if not year:
        year = int(df_all["FiscalReportingPeriod"].max())
    df = df_all[df_all["FiscalReportingPeriod"] == year].copy()
    df["month"] = df["ReportingPeriod"].dt.month

    months_present = sorted(df["month"].unique())
    tier1_by_month, tier2_by_month, pstir_by_month, psisr_by_month = [], [], [], []
    for m in months_present:
        r = process_safety_rates(df[df["month"] == m])
        tier1_by_month.append(r["tier1"])
        tier2_by_month.append(r["tier2"])
        pstir_by_month.append(r["pstir"])
        psisr_by_month.append(r["psisr"])

    plants = sorted(df["PlantName"].unique().tolist())
    by_plant = {p: process_safety_rates(df[df["PlantName"] == p]) for p in plants}
    total = process_safety_rates(df)

    return {
        "year": year,
        "months": [MONTHS.get(int(m), str(int(m))) for m in months_present],
        "tier1": tier1_by_month,
        "tier2": tier2_by_month,
        "pstir": pstir_by_month,
        "psisr": psisr_by_month,
        "plants": plants,
        "pstir_by_plant": [by_plant[p]["pstir"] for p in plants],
        "psisr_by_plant": [by_plant[p]["psisr"] for p in plants],
        "total_incidents": total["incidents"],
        "total_tier1": total["tier1"],
        "total_tier2": total["tier2"],
        "total_pstir": total["pstir"],
        "total_psisr": total["psisr"],
    }


@app.get("/api/admin/cache-status")
def get_cache_status():
    result = {}
    for key, entry in _CACHE.items():
        path = os.path.join(DATA_DIR, key)
        try:
            current_mtime = os.path.getmtime(path)
        except OSError:
            current_mtime = 0
        result[key] = {
            "cached_mtime": entry["mtime"],
            "current_mtime": current_mtime,
            "is_stale": current_mtime != entry["mtime"],
            "loaded_at": entry.get("loaded_at"),
        }
    return result

# ─── INSIGHTS ────────────────────────────────────────────────────────────────
# Deterministic, rule-based 3-bullet summaries per sub-tab - computed directly
# from the filtered data so they're instant and don't depend on an AI key.

@app.get("/api/insights/water")
def get_water_insights(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    df_all = safe_filter_region(safe_filter_plant(load_water_data(), plant), region)
    years = sorted(int(y) for y in df_all["ReportingYear"].unique().tolist())

    if year:
        df = df_all[df_all["ReportingYear"] == year]
        prev_df = df_all[df_all["ReportingYear"] == year - 1]
        curr_label, compare_label = str(year), str(year - 1)
    else:
        df = df_all
        prev_df = df_all[df_all["ReportingYear"] == years[0]] if years else df_all.iloc[0:0]
        df = df_all[df_all["ReportingYear"] == years[-1]] if len(years) >= 2 else df_all
        curr_label = str(years[-1]) if years else ""
        compare_label = str(years[0]) if len(years) >= 2 else None

    def _ml(v_m3): return round(float(v_m3) / 1000, 1)

    insights = []
    curr_withdrawn = float(df["TotalWaterWithdrawn"].sum())
    prev_withdrawn = float(prev_df["TotalWaterWithdrawn"].sum()) if len(prev_df) else None
    if prev_withdrawn:
        change = round((curr_withdrawn - prev_withdrawn) / prev_withdrawn * 100, 1)
        direction = "up" if change >= 0 else "down"
        insights.append(f"Water withdrawal in {curr_label} is **{abs(change)}% {direction}** vs {compare_label}, totaling **{_ml(curr_withdrawn):,} ML**.")
    else:
        insights.append(f"Total water withdrawal for the selected period is **{_ml(curr_withdrawn):,} ML** ({curr_label} is the earliest year on record).")

    stress = float(df["TotalWaterWithdrawnStressArea"].sum())
    stress_pct = round(stress / curr_withdrawn * 100, 1) if curr_withdrawn > 0 else 0
    insights.append(f"**{stress_pct}%** of withdrawal (**{_ml(stress):,} ML**) comes from high-water-stress areas.")

    by_plant = df.groupby("PlantName")["TotalWaterWithdrawn"].sum().sort_values(ascending=False)
    if len(by_plant) > 1:
        insights.append(f"**{by_plant.index[0]}** withdraws the most water (**{_ml(by_plant.iloc[0]):,} ML**), while **{by_plant.index[-1]}** withdraws the least (**{_ml(by_plant.iloc[-1]):,} ML**).")
    elif len(by_plant) == 1:
        insights.append(f"**{by_plant.index[0]}** is the only plant in this view, withdrawing **{_ml(by_plant.iloc[0]):,} ML**.")
    else:
        insights.append("No plant-level data available for the selected filters.")

    return {"insights": insights[:3]}

@app.get("/api/insights/waste")
def get_waste_insights(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    df_all = safe_filter_region(safe_filter_plant(load_waste_data(), plant), region)
    years = sorted(int(y) for y in df_all["FiscalReportingPeriod"].unique().tolist())
    df = df_all[df_all["FiscalReportingPeriod"] == year] if year else df_all
    curr_label = str(year) if year else (str(years[-1]) if years else "")

    diverted = float(df[df["WasteCategory"] == "Diverted"]["ValueNumber"].sum())
    disposed = float(df[df["WasteCategory"] == "Disposed"]["ValueNumber"].sum())
    generated = diverted + disposed
    diversion_rate = round(diverted / generated * 100, 1) if generated > 0 else 0

    insights = [f"Current waste diversion rate is **{diversion_rate}%** (**{round(diverted):,}** of **{round(generated):,} tons** diverted from disposal)."]

    by_plant_div = df[df["WasteCategory"] == "Diverted"].groupby("PlantName")["ValueNumber"].sum()
    by_plant_dis = df[df["WasteCategory"] == "Disposed"].groupby("PlantName")["ValueNumber"].sum()
    plants = sorted(set(by_plant_div.index) | set(by_plant_dis.index))
    rates = {}
    for p in plants:
        g = float(by_plant_div.get(p, 0)) + float(by_plant_dis.get(p, 0))
        if g > 0:
            rates[p] = round(float(by_plant_div.get(p, 0)) / g * 100, 1)
    if len(rates) > 1:
        best, worst = max(rates, key=rates.get), min(rates, key=rates.get)
        insights.append(f"**{best}** leads in diversion rate at **{rates[best]}%**, while **{worst}** lags at **{rates[worst]}%**.")
    elif len(rates) == 1:
        p = next(iter(rates))
        insights.append(f"**{p}** is the only plant in this view, with a diversion rate of **{rates[p]}%**.")
    else:
        insights.append("No plant-level diversion data available for the selected filters.")

    disposal_pct = round(disposed / generated * 100, 1) if generated > 0 else 0
    insights.append(f"**{round(disposed):,} tons ({disposal_pct}%)** of generated waste was sent to disposal rather than diverted in {curr_label}.")

    return {"insights": insights[:3]}

@app.get("/api/insights/safety")
def get_safety_insights(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    df_all = safe_filter_region(safe_filter_plant(load_safety_data(), plant), region)
    years = sorted(int(y) for y in df_all["FiscalReportingPeriod"].unique().tolist())
    df = df_all[df_all["FiscalReportingPeriod"] == year] if year else df_all
    curr_label = str(year) if year else (str(years[-1]) if years else "")

    hours = df["HoursWorked"].sum()
    trir = rate_per_basis(df["RecordableInjuries"].sum(), hours) if hours else None
    ltifr = rate_per_basis(df["LostTimeInjuries"].sum(), hours) if hours else None
    if trir is not None:
        risk_note = "an elevated" if trir > 3 else "a moderate" if trir > 1 else "a low"
        insights = [f"TRIR for {curr_label} is **{trir}** per 200,000 hours worked (LTIFR: **{ltifr}**), indicating **{risk_note}** injury frequency rate."]
    else:
        insights = ["No injury-rate data available for the selected filters."]

    causes = explode_weighted(df, "MainInjuryType", "RecordableInjuries")
    if len(causes):
        insights.append(f"**\"{causes.index[0]}\"** is the leading recordable-injury cause, accounting for **{round(float(causes.iloc[0]), 1)}** weighted case(s).")
    else:
        insights.append("No incident cause data available for the selected filters.")

    by_plant = df.groupby("PlantName").agg(hours=("HoursWorked", "sum"), recordable=("RecordableInjuries", "sum"))
    trir_by_plant = {p: rate_per_basis(row["recordable"], row["hours"]) for p, row in by_plant.iterrows()}
    if len(trir_by_plant) > 1:
        worst_plant = max(trir_by_plant, key=trir_by_plant.get)
        insights.append(f"**{worst_plant}** has the highest TRIR at **{trir_by_plant[worst_plant]}**, warranting closer safety review.")
    elif len(trir_by_plant) == 1:
        p = next(iter(trir_by_plant))
        insights.append(f"**{p}** is the only plant in this view, with a TRIR of **{trir_by_plant[p]}**.")
    else:
        insights.append("No plant-level TRIR data available for the selected filters.")

    return {"insights": insights[:3]}

@app.get("/api/insights/energy")
def get_energy_insights(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    df_all = safe_filter_region(safe_filter_plant(load_energy_data(), plant), region)
    years = sorted(int(y) for y in df_all["ReportingYear"].unique().tolist())

    if year:
        df = df_all[df_all["ReportingYear"] == year]
        prev_df = df_all[df_all["ReportingYear"] == year - 1]
        curr_label, compare_label = str(year), str(year - 1)
    else:
        df = df_all[df_all["ReportingYear"] == years[-1]] if years else df_all
        prev_df = df_all[df_all["ReportingYear"] == years[0]] if len(years) >= 2 else df_all.iloc[0:0]
        curr_label = str(years[-1]) if years else ""
        compare_label = str(years[0]) if len(years) >= 2 else None

    consumed = float(df["TotalEnergyConsumedGJ"].sum())
    prev_consumed = float(prev_df["TotalEnergyConsumedGJ"].sum()) if len(prev_df) else None
    if prev_consumed:
        change = round((consumed - prev_consumed) / prev_consumed * 100, 1)
        direction = "up" if change >= 0 else "down"
        insights = [f"Energy consumption in {curr_label} is **{abs(change)}% {direction}** vs {compare_label}, totaling **{round(consumed):,} GJ**."]
    else:
        insights = [f"Total energy consumption for the selected period is **{round(consumed):,} GJ** ({curr_label} is the earliest year on record)."]

    renewable = float(df["ElectricityRenewableGJ"].sum())
    non_renewable = float(df["ElectricityNonRenewableGJ"].sum())
    total_elec = renewable + non_renewable
    renewable_pct = round(renewable / total_elec * 100, 1) if total_elec > 0 else 0
    insights.append(f"**{renewable_pct}%** of electricity consumed (**{round(renewable):,} GJ**) comes from renewable sources.")

    by_plant = df.groupby("PlantName")["TotalEnergyConsumedGJ"].sum().sort_values(ascending=False)
    if len(by_plant) > 1:
        insights.append(f"**{by_plant.index[0]}** consumes the most energy (**{round(by_plant.iloc[0]):,} GJ**), while **{by_plant.index[-1]}** consumes the least (**{round(by_plant.iloc[-1]):,} GJ**).")
    elif len(by_plant) == 1:
        insights.append(f"**{by_plant.index[0]}** is the only plant in this view, consuming **{round(by_plant.iloc[0]):,} GJ**.")
    else:
        insights.append("No plant-level data available for the selected filters.")

    return {"insights": insights[:3]}

@app.get("/api/insights/emissions")
def get_emissions_insights(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    df_all = safe_filter_region(safe_filter_plant(load_ghg_data(), plant), region)
    years = sorted(int(y) for y in df_all["ReportingYear"].unique().tolist())

    if year:
        df = df_all[df_all["ReportingYear"] == year]
        prev_df = df_all[df_all["ReportingYear"] == year - 1]
        curr_label, compare_label = str(year), str(year - 1)
    else:
        df = df_all[df_all["ReportingYear"] == years[-1]] if years else df_all
        prev_df = df_all[df_all["ReportingYear"] == years[0]] if len(years) >= 2 else df_all.iloc[0:0]
        curr_label = str(years[-1]) if years else ""
        compare_label = str(years[0]) if len(years) >= 2 else None

    scope1 = float(df["Scope1TotaltCO2e"].sum())
    prev_scope1 = float(prev_df["Scope1TotaltCO2e"].sum()) if len(prev_df) else None
    if prev_scope1:
        change = round((scope1 - prev_scope1) / prev_scope1 * 100, 1)
        direction = "up" if change >= 0 else "down"
        insights = [f"Scope 1 emissions in {curr_label} are **{abs(change)}% {direction}** vs {compare_label}, totaling **{round(scope1):,} tCO₂e**."]
    else:
        insights = [f"Total Scope 1 emissions for the selected period are **{round(scope1):,} tCO₂e** ({curr_label} is the earliest year on record)."]

    scope2 = float(df["Scope2LocationBasedtCO2e"].sum())
    scope3 = float(df["Scope3TotaltCO2e"].sum())
    total = scope1 + scope2 + scope3
    scope1_pct = round(scope1 / total * 100, 1) if total > 0 else 0
    insights.append(f"Scope 1 makes up **{scope1_pct}%** of total tracked emissions (Scope 1+2+3: **{round(total):,} tCO₂e**).")

    by_plant = df.groupby("PlantName")["Scope1TotaltCO2e"].sum().sort_values(ascending=False)
    if len(by_plant) > 1:
        insights.append(f"**{by_plant.index[0]}** has the highest Scope 1 footprint (**{round(by_plant.iloc[0]):,} tCO₂e**), while **{by_plant.index[-1]}** has the lowest (**{round(by_plant.iloc[-1]):,} tCO₂e**).")
    elif len(by_plant) == 1:
        insights.append(f"**{by_plant.index[0]}** is the only plant in this view, with Scope 1 emissions of **{round(by_plant.iloc[0]):,} tCO₂e**.")
    else:
        insights.append("No plant-level emissions data available for the selected filters.")

    return {"insights": insights[:3]}

# ─── OUTLIER DETECTION ───────────────────────────────────────────────────────
# IQR-based year-over-year outlier detection per domain.
# Returns up to 5 most significant anomalies with severity and a description.

_OUTLIER_CONFIG = {
    "water":     [("TotalWaterWithdrawn",  "Total Water Withdrawn",  "ML",     1/1000),
                  ("WaterConsumed",         "Water Consumed",         "ML",     1/1000)],
    "waste":     [("_generated",           "Waste Generated",        "tonnes", 1)],
    "safety":    [("RecordableInjuries",   "Recordable Injuries",    "count",  1),
                  ("LostTimeInjuries",      "Lost Time Injuries",     "count",  1)],
    "energy":    [("TotalEnergyConsumedGJ","Energy Consumed",        "GJ",     1),
                  ("EnergyIntensityGJpert", "Energy Intensity",       "GJ/t",   1)],
    "emissions": [("Scope1TotaltCO2e",     "Scope 1 GHG Emissions",  "tCO₂e",  1),
                  ("GrossNOxt",             "NOx Emissions",          "tonnes", 1)],
}

def _compute_outliers(domain: str, year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None,
                      low_pct: float = 12.0, medium_pct: float = 25.0, high_pct: float = 50.0,
                      limit: Optional[int] = 6):
    configs = _OUTLIER_CONFIG.get(domain, [])
    if not configs:
        return {"anomalies": [], "latest_trend": []}

    loaders = {
        "water":     (load_water_data,    "ReportingYear"),
        "waste":     (load_waste_data,     "FiscalReportingPeriod"),
        "safety":    (load_safety_data,    "FiscalReportingPeriod"),
        "energy":    (load_energy_data,    "ReportingYear"),
        "emissions": (load_ghg_data,       "ReportingYear"),
    }
    loader_fn, year_col = loaders[domain]
    df = safe_filter_region(safe_filter_plant(loader_fn(), plant), region)

    anomalies = []
    latest_trend = []
    for col, label, unit, scale in configs:
        # Build yearly-plant totals (or means for intensity)
        use_mean = "Intensity" in col or "intensity" in col
        agg_fn = "mean" if use_mean else "sum"

        if col == "_generated":
            # Waste generated = diverted + disposed
            yearly = (df.groupby(["FiscalReportingPeriod", "PlantName"])["ValueNumber"]
                        .sum().reset_index()
                        .rename(columns={"FiscalReportingPeriod": "yr", "ValueNumber": "val"}))
        else:
            yearly = (df.groupby([year_col, "PlantName"])[col]
                        .agg(agg_fn).reset_index()
                        .rename(columns={year_col: "yr", col: "val"}))

        plants_list = sorted(yearly["PlantName"].unique())
        years_list  = sorted(yearly["yr"].unique())
        if len(years_list) < 2:
            continue

        # Anomaly detection looks back over the 5 most recent reporting years.
        # Historical years remain in 'yearly' as baselines but are never flagged.
        recent_window = set(years_list[-5:])

        for p in plants_list:
            pseries = yearly[yearly["PlantName"] == p].set_index("yr")["val"].sort_index()
            yrs = pseries.index.tolist()
            for i in range(1, len(yrs)):
                y_cur  = int(yrs[i])
                y_prev = int(yrs[i - 1])
                if y_cur not in recent_window:
                    continue  # skip years outside the 5-year window
                v_cur  = float(pseries[y_cur])
                v_prev = float(pseries[y_prev])
                if v_prev == 0:
                    continue
                chg = round((v_cur - v_prev) / abs(v_prev) * 100, 1)
                if abs(chg) < low_pct:
                    continue
                severity = "high" if abs(chg) >= high_pct else "medium" if abs(chg) >= medium_pct else "low"
                direction = "up" if chg > 0 else "down"
                v_display   = round(v_cur  * scale, 3)
                v_prev_disp = round(v_prev * scale, 3)
                anomalies.append({
                    "metric": label,
                    "unit": unit,
                    "plant": p,
                    "year": y_cur,
                    "value": v_display,
                    "prev_value": v_prev_disp,
                    "change_pct": abs(chg),
                    "direction": direction,
                    "severity": severity,
                    "description": (
                        f"**{label}** at {p} {direction} **{abs(chg)}%** "
                        f"({y_prev}→{y_cur}: {v_prev_disp:,} → {v_display:,} {unit})"
                    ),
                })

        # Latest YoY trend: organization-wide (all plants combined), always
        # shown regardless of the anomaly threshold, so every domain has a
        # visible most-recent-year comparison even when nothing is anomalous.
        yearly_total = yearly.groupby("yr")["val"].agg(agg_fn).sort_index()
        yrs_total = yearly_total.index.tolist()
        if len(yrs_total) >= 2:
            yc, yp = yrs_total[-1], yrs_total[-2]
            vc, vp = float(yearly_total[yc]), float(yearly_total[yp])
            if vp != 0:
                chg = round((vc - vp) / abs(vp) * 100, 1)
                latest_trend.append({
                    "metric": label,
                    "unit": unit,
                    "year": int(yc),
                    "prev_year": int(yp),
                    "value": round(vc * scale, 3),
                    "prev_value": round(vp * scale, 3),
                    "change_pct": abs(chg),
                    "direction": "up" if chg > 0 else "down",
                })

    # Sort by year (most recent first), then severity, then change magnitude
    severity_rank = {"high": 0, "medium": 1, "low": 2}
    anomalies.sort(key=lambda x: (-x["year"], severity_rank[x["severity"]], -x["change_pct"]))
    return {"anomalies": anomalies[:limit] if limit else anomalies, "latest_trend": latest_trend}

@app.get("/api/outliers/{domain}")
def get_outliers(domain: str, year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    t = storage.get_domain_thresholds(domain)
    return _compute_outliers(domain, year, plant, region, t["low"], t["medium"], t["high"])

@app.get("/api/outliers")
def get_all_outliers(year: Optional[int] = None, plant: Optional[str] = None, region: Optional[str] = None):
    """Aggregate anomaly feed across every GRI domain with live data, for the
    Alerts screen. SASB/BRSR aren't given separate outlier configs since their
    underlying datasets substantially overlap these 5 GRI domains already."""
    all_anomalies = []
    all_trend = []
    for domain in _OUTLIER_CONFIG:
        t = storage.get_domain_thresholds(domain)
        result = _compute_outliers(domain, year, plant, region, t["low"], t["medium"], t["high"])
        for a in result["anomalies"]:
            all_anomalies.append({**a, "domain": domain})
        for t2 in result["latest_trend"]:
            all_trend.append({**t2, "domain": domain})

    severity_rank = {"high": 0, "medium": 1, "low": 2}
    all_anomalies.sort(key=lambda x: (-x["year"], severity_rank[x["severity"]], -x["change_pct"]))
    return {"anomalies": all_anomalies, "latest_trend": all_trend}

# ─── AI ASSISTANT ───────────────────────────────────────────────────────────

class ChatRequest(BaseModel):
    message: str
    context: Optional[str] = "general"
    history: Optional[list] = []

# Domains with a live dataset behind them vs. placeholder-only domains.
LIVE_DOMAINS = {"water", "waste", "safety", "energy", "emissions"}
PLACEHOLDER_DOMAINS = {"workforce", "development"}
DOMAIN_LABELS = {
    "water": "Water", "waste": "Waste", "safety": "Safety", "energy": "Energy",
    "emissions": "Emissions", "workforce": "Workforce", "development": "Development",
}

def build_system_prompt(context: str, scope: Optional[str] = None, framework: str = "GRI"):
    # scope=None/"all" -> include every live domain (whole-context awareness).
    # scope=<domain> -> include only that domain's summary, scoped deep-dive mode.
    include_water = scope in (None, "all", "water")
    include_waste = scope in (None, "all", "waste")
    include_safety = scope in (None, "all", "safety")
    include_energy = scope in (None, "all", "energy")
    include_emissions = scope in (None, "all", "emissions")

    water_summary = ""
    waste_summary = ""
    safety_summary = ""
    energy_summary = ""
    emissions_summary = ""
    if include_water:
        try:
            wdf = load_water_data()
            water_summary = f"""
Water Data (GRI 303):
- Years: {sorted(wdf['ReportingYear'].unique().tolist())}
- Plants: {sorted(wdf['PlantName'].unique().tolist())}
- Total Withdrawn (all time): {round(float(wdf['TotalWaterWithdrawn'].sum()),0)} m3
- Total Consumed (all time): {round(float(wdf['WaterConsumed'].sum()),0)} m3
- Water from Stress Areas: {round(float(wdf['TotalWaterWithdrawnStressArea'].sum()),0)} m3
"""
        except: pass
    if include_waste:
        try:
            wsdf = load_waste_data()
            waste_summary = f"""
Waste Data (GRI 306):
- Years: {sorted(wsdf['FiscalReportingPeriod'].unique().tolist())}
- Plants: {sorted(wsdf['PlantName'].unique().tolist())}
- Total Diverted (all time): {round(float(wsdf[wsdf['WasteCategory']=='Diverted']['ValueNumber'].sum()),0)} tons
- Total Disposed (all time): {round(float(wsdf[wsdf['WasteCategory']=='Disposed']['ValueNumber'].sum()),0)} tons
"""
        except: pass
    if include_safety:
        try:
            sdf = load_safety_data()
            safety_summary = f"""
Safety Data (GRI 403):
- Years: {sorted(sdf['FiscalReportingPeriod'].unique().tolist())}
- Plants: {sorted(sdf['PlantName'].unique().tolist())}
- Worker Types: {sorted(sdf['WorkerType'].unique().tolist())}
- TRIR (all time, per 200,000 hrs): {rate_per_basis(sdf['RecordableInjuries'].sum(), sdf['HoursWorked'].sum())}
- LTIFR (all time, per 200,000 hrs): {rate_per_basis(sdf['LostTimeInjuries'].sum(), sdf['HoursWorked'].sum())}
- Top injury causes: {explode_weighted(sdf, 'MainInjuryType', 'RecordableInjuries').head(5).round(1).to_dict()}
"""
        except: pass
    if include_energy:
        try:
            edf = load_energy_data()
            energy_summary = f"""
Energy Data (GRI 302):
- Years: {sorted(edf['ReportingYear'].unique().tolist())}
- Plants: {sorted(edf['PlantName'].unique().tolist())}
- Total Consumed (all time): {round(float(edf['TotalEnergyConsumedGJ'].sum()),0)} GJ
- Energy Intensity (GRI-weighted, total GJ / total tonnes): {round(weighted_ratio(edf, 'TotalEnergyConsumedGJ'),2)} GJ/tonne
"""
        except: pass
    if include_emissions:
        try:
            gdf = load_ghg_data()
            emissions_summary = f"""
GHG & Air Emissions Data (GRI 305):
- Years: {sorted(gdf['ReportingYear'].unique().tolist())}
- Plants: {sorted(gdf['PlantName'].unique().tolist())}
- Scope 1 (all time): {round(float(gdf['Scope1TotaltCO2e'].sum()),0)} tCO2e
- Scope 2 location-based (all time): {round(float(gdf['Scope2LocationBasedtCO2e'].sum()),0)} tCO2e
- Scope 3 (all time): {round(float(gdf['Scope3TotaltCO2e'].sum()),0)} tCO2e
"""
        except: pass

    no_data_note = ""
    focus_instruction = ""
    if scope and scope != "all":
        label = DOMAIN_LABELS.get(scope, scope.capitalize())
        focus_instruction = f"\nFocus this entire response on {label} only. Do not discuss other domains unless the user explicitly asks for a comparison.\n"
        if scope in PLACEHOLDER_DOMAINS:
            no_data_note = f"\nNote: No dataset is currently connected for {label} (this GRI disclosure isn't tracked yet). Say so plainly, then base any analysis on general ESG best practice rather than fabricating numbers.\n"

    # SASB RT-CH (Chemicals) awareness — always included so the assistant can
    # answer SASB-framed questions even though SASB has its own separate
    # dashboard (see SASB_INTEGRATION_PLAN.md Section 6). Computed from the
    # same datasets summarized above, just reframed with SASB's calculations
    # (hazardous-waste split, fatality rate, Employee/Contractor TRIR split).
    sasb_summary = ""
    try:
        wsdf = load_waste_data()
        sdf = load_safety_data()
        haz = hazardous_waste_breakdown(wsdf)
        sasb_summary = f"""
SASB RT-CH (Chemicals) framework is also available on this dashboard, alongside GRI, as a separate set of dashboards and reports:
- Hazardous Waste Generated (RT-CH-150a, all time): {round(haz['haz_generated'])} tonnes, {haz['haz_recycled_pct']}% recycled
- Non-Hazardous Waste Generated (all time): {round(haz['nonhaz_generated'])} tonnes
- TRIR split by worker type (RT-CH-320a, all time): Employees {safety_rate_split(sdf, 'Employee')}, Contractors {safety_rate_split(sdf, 'Contractor')}, per 200,000 hrs
- Fatality Rate (RT-CH-320a, all time): {fatality_rate(sdf)} per 200,000 hrs
- Process Safety Incidents (RT-CH-540a): not tracked — no process-safety dataset exists yet (genuine data gap, not a placeholder oversight)
If asked about SASB metrics, use these figures rather than claiming SASB isn't supported.
"""
    except Exception:
        pass

    # BRSR (Business Responsibility & Sustainability Reporting) — SEBI/MCA
    # mandate for Indian-listed companies, organized around NGRBC 9 Principles.
    # P6 environment data comes from the same GRI datasets already summarized;
    # P3 workforce/training and P8 CSR use dedicated BRSR annual datasets.
    brsr_summary = ""
    brsr_focus_instruction = ""
    try:
        wfdf = load_workforce_data()
        trdf = load_training_data()
        csrdf = load_csr_data()
        cfg = _load_brsr_config()

        total_emp_count = (
            wfdf[["PermanentMale","PermanentFemale","PermanentOther",
                   "ContractualMale","ContractualFemale","ContractualOther"]].sum().sum()
        )
        total_female = wfdf[["PermanentFemale","ContractualFemale"]].sum().sum()
        female_pct_all = round(float(total_female) / float(total_emp_count) * 100, 1) if total_emp_count > 0 else 0
        da_count_all = int(wfdf[["DifferentlyAbledPermanent","DifferentlyAbledContractual"]].sum().sum())

        avg_training_hrs = round(float(trdf["AvgTrainingHrsAllEmployees"].mean()), 1)
        avg_coverage = round(float(trdf["TrainingCoveragePct"].mean()), 1)

        fys_available = sorted(wfdf["FY"].unique().tolist()) if "FY" in wfdf.columns else []
        csr_total = round(float(csrdf["TotalSpentCrore"].sum()), 2) if "TotalSpentCrore" in csrdf.columns else 0

        p1 = cfg.get("principles", {}).get("p1", {})
        recv = p1.get("complaints_received") or 0
        resv = p1.get("complaints_resolved") or 0
        complaint_str = f"{round(resv/recv*100,1)}% resolution rate ({resv}/{recv} complaints)" if recv > 0 else "data not yet in config"

        brsr_summary = f"""
BRSR (Business Responsibility & Sustainability Reporting) is the third framework on this dashboard, mandated by SEBI for Indian-listed companies. It maps to NGRBC's 9 Principles (P1–P9):
- P6 — Environment: Same GRI datasets above (energy, water, GHG, waste) reported under Indian Financial Year windows (Apr–Mar). FYs available: {fys_available if fys_available else 'FY2020-21 to FY2024-25'}
- P3 — Workforce (Principle 3, Essential): Female representation {female_pct_all}% of total workforce; Differently-abled employees: {da_count_all}
- P3 — Training (Principle 3): Avg {avg_training_hrs} hrs/employee training; Training coverage: {avg_coverage}%
- P8 — CSR (Principle 8): Total CSR spend ₹{csr_total} Cr across all FYs
- P1 — Ethics & Compliance (Principle 1): Complaint resolution — {complaint_str}
- P2, P4, P5, P7, P9: Qualitative disclosures tracked in brsr_config.json (policy/stakeholder data)
When answering BRSR questions, map metrics to the correct Principle number and specify whether it is an Essential or Leadership indicator.
"""
        if framework == "BRSR":
            brsr_focus_instruction = "\nThe user is currently viewing the BRSR dashboard. Frame all responses using BRSR/NGRBC terminology. Map every metric to its NGRBC Principle (P1–P9) and state whether it is an Essential or Leadership indicator. For environmental metrics, present them as P6 disclosures using Indian FY conventions.\n"
    except Exception:
        pass

    return f"""You are an expert ESG (Environmental, Social & Governance) analyst assistant for an executive dashboard.
You specialize in three reporting frameworks tracked by this dashboard:
1. GRI Standards: GRI 302 (Energy), GRI 303 (Water & Effluents), GRI 305 (Emissions), GRI 306 (Waste), GRI 403 (Occupational Health & Safety).
2. SASB Sustainability Accounting Standards (Chemicals industry, RT-CH) — a separate set of dashboards and reports.
3. BRSR (Business Responsibility & Sustainability Reporting) — SEBI/MCA mandate for Indian-listed companies, organized around NGRBC 9 Principles (P1–P9).
All three frameworks draw from the same underlying operational datasets; the framing, periodicity, and metric calculations differ by framework.

You have access to the following organizational data:
{water_summary}
{waste_summary}
{safety_summary}
{energy_summary}
{emissions_summary}
{sasb_summary}
{brsr_summary}
{no_data_note}
Your capabilities:
1. Answer questions about KPIs and data trends across GRI, SASB, and BRSR frameworks
2. Generate GRI-compliant or BRSR-compliant narrative report text
3. Identify anomalies and provide insights
4. Explain GRI, SASB RT-CH, and BRSR/NGRBC disclosure requirements
5. Recommend improvements based on data patterns
{focus_instruction}{brsr_focus_instruction}
Always be concise, data-driven, and actionable. When generating report narratives, follow the disclosure structure of the framework in use.

Format every response in Markdown so it renders cleanly in a chat UI:
- Open with one direct-answer sentence - no preamble.
- Use "- " bullet points for supporting details and breakdowns.
- Use "1. " numbered lists for sequential or prioritized recommendations.
- Use **bold** for key numbers, metrics, and plant/category names.
- Use short "### " headers only if the answer has multiple distinct sections.
- Keep paragraphs to 1-2 sentences. Never return a single unbroken wall of text."""

@app.post("/api/chat")
async def chat(req: ChatRequest):
    if not MISTRAL_API_KEY:
        raise HTTPException(status_code=500, detail="MISTRAL_API_KEY not configured")

    system_prompt = build_system_prompt(req.context)
    messages = [{"role": "system", "content": system_prompt}]
    for h in (req.history or [])[-6:]:
        messages.append({"role": h["role"], "content": h["content"]})
    messages.append({"role": "user", "content": req.message})

    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.post(
            MISTRAL_API_URL,
            headers={"Authorization": f"Bearer {MISTRAL_API_KEY}", "Content-Type": "application/json"},
            json={"model": "mistral-large-latest", "messages": messages, "max_tokens": 1024, "temperature": 0.3},
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail=resp.text)
        data = resp.json()
        return {"reply": data["choices"][0]["message"]["content"]}

class AiChatRequest(BaseModel):
    message: str
    tab: Optional[str] = "environment"
    sub_tab: Optional[str] = None
    filters: Optional[dict] = None
    chat_history: Optional[list] = []
    domain_hint: Optional[str] = None
    thread_id: Optional[str] = None
    framework: Optional[str] = "GRI"

def _sse(data: dict) -> str:
    return f"data: {json.dumps(data)}\n\n"

# Questions using these words call for deeper reasoning (recommendations,
# predictions, root-cause analysis) where answer quality is much better if
# scoped to one domain instead of dumping water+waste+safety context at once.
DEEP_REASONING_KEYWORDS = [
    "recommend", "recommendation", "predict", "prediction", "forecast",
    "analyse", "analyze", "analysis", "suggest", "strategy", "optimi",
    "root cause", "deep dive", "deep-dive", "improve", "improvement",
]
DOMAIN_KEYWORDS = {
    "water": ["water"],
    "waste": ["waste"],
    "safety": ["safety", "ltifr", "trir", "injury", "incident"],
    "energy": ["energy"],
    "emissions": ["emission", "ghg", "carbon", "scope 1", "scope 2"],
    "workforce": ["workforce", "headcount", "employee", "gender"],
    "development": ["training", "development", "turnover", "hire", "hiring"],
}
ALL_SCOPE_KEYWORDS = [
    "all tabs", "every tab", "everything", "all areas", "all domains",
    "across the board", "entire organization", "every area", "whole dashboard", "all of it",
]

def detect_domain_intent(message: str) -> Optional[str]:
    """Returns None for normal questions (use full context), 'all' for an
    explicit full-picture request, a domain key if one is named, or
    'clarify' when the question needs deep reasoning but names no domain."""
    lower = message.lower()
    if not any(k in lower for k in DEEP_REASONING_KEYWORDS):
        return None
    if any(k in lower for k in ALL_SCOPE_KEYWORDS):
        return "all"
    for domain, kws in DOMAIN_KEYWORDS.items():
        if any(k in lower for k in kws):
            return domain
    return "clarify"

@app.post("/api/ai/chat")
async def ai_chat_stream(req: AiChatRequest):
    thread_id = req.thread_id
    if thread_id and not storage.get_thread(thread_id):
        thread_id = None  # caller passed a stale/unknown id - fall back to ephemeral

    async def event_gen():
        if thread_id:
            storage.add_message(thread_id, "user", req.message)
            thread = storage.get_thread(thread_id)
            # First user message in a thread becomes its title, mirroring how
            # ChatGPT/Claude derive thread titles from the opening question.
            if thread["title"] == "New chat" and len([m for m in thread["messages"] if m["role"] == "user"]) == 1:
                storage.rename_thread(thread_id, req.message[:60])

        if not MISTRAL_API_KEY:
            msg = "AI Assistant is not configured. Set MISTRAL_API_KEY on the backend to enable this feature."
            yield _sse({"token": msg})
            if thread_id:
                storage.add_message(thread_id, "assistant", msg)
            yield _sse({"done": True})
            return

        scope = req.domain_hint or detect_domain_intent(req.message)
        if scope == "clarify":
            msg = "That's the kind of question I can answer much more sharply if I focus on one area - which would you like me to dig into?"
            options = [{"id": d, "label": label} for d, label in DOMAIN_LABELS.items()] + [{"id": "all", "label": "All Areas"}]
            yield _sse({"token": msg})
            yield _sse({"clarify": True, "options": options})
            if thread_id:
                storage.add_message(thread_id, "assistant", msg)
            yield _sse({"done": True})
            return

        context_label = f"{req.tab} > {req.sub_tab}" if req.sub_tab else req.tab
        system_prompt = build_system_prompt(req.tab, scope=scope, framework=req.framework or "GRI") + (
            f"\n\nCurrent dashboard context: {context_label}. Active framework: {req.framework or 'GRI'}. Active filters: {json.dumps(req.filters or {})}"
        )
        messages = [{"role": "system", "content": system_prompt}]
        for h in (req.chat_history or [])[-6:]:
            messages.append({"role": h["role"], "content": h["content"]})
        messages.append({"role": "user", "content": req.message})

        full_reply = []
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                async with client.stream(
                    "POST", MISTRAL_API_URL,
                    headers={"Authorization": f"Bearer {MISTRAL_API_KEY}", "Content-Type": "application/json"},
                    json={"model": MISTRAL_MODEL, "messages": messages, "max_tokens": MISTRAL_MAX_TOKENS, "temperature": MISTRAL_TEMP, "stream": True},
                ) as resp:
                    if resp.status_code != 200:
                        body = await resp.aread()
                        yield _sse({"error": f"AI service error {resp.status_code}: {body.decode(errors='ignore')[:200]}"})
                        yield _sse({"done": True})
                        return
                    async for line in resp.aiter_lines():
                        if not line or not line.startswith("data: "):
                            continue
                        payload = line[len("data: "):].strip()
                        if payload == "[DONE]":
                            break
                        try:
                            chunk = json.loads(payload)
                            delta = chunk["choices"][0]["delta"].get("content", "")
                        except Exception:
                            continue
                        if delta:
                            full_reply.append(delta)
                            yield _sse({"token": delta})
        except Exception as e:
            yield _sse({"error": str(e)})
            yield _sse({"done": True})
            return

        if thread_id and full_reply:
            storage.add_message(thread_id, "assistant", "".join(full_reply))
        yield _sse({"done": True})

    return StreamingResponse(event_gen(), media_type="text/event-stream")

# ─── CHAT THREADS (persisted) ───────────────────────────────────────────────

class RenameThreadRequest(BaseModel):
    title: str

@app.get("/api/chats")
def list_chat_threads():
    return storage.list_threads()

@app.post("/api/chats")
def create_chat_thread():
    return storage.create_thread()

@app.delete("/api/chats")
def clear_chat_threads():
    storage.clear_all_threads()
    return {"ok": True}

@app.get("/api/chats/{thread_id}")
def get_chat_thread(thread_id: str):
    thread = storage.get_thread(thread_id)
    if not thread:
        raise HTTPException(status_code=404, detail="Thread not found")
    return thread

@app.patch("/api/chats/{thread_id}")
def rename_chat_thread(thread_id: str, req: RenameThreadRequest):
    if not storage.rename_thread(thread_id, req.title):
        raise HTTPException(status_code=404, detail="Thread not found")
    return {"ok": True}

@app.delete("/api/chats/{thread_id}")
def delete_chat_thread(thread_id: str):
    if not storage.delete_thread(thread_id):
        raise HTTPException(status_code=404, detail="Thread not found")
    return {"ok": True}

# ─── REPORTS ─────────────────────────────────────────────────────────────────

REPORT_TEMPLATES = [
    {"id": "gri_303", "name": "GRI 303 — Water", "gri": "303", "hasData": True},
    {"id": "gri_306", "name": "GRI 306 — Waste", "gri": "306", "hasData": True},
    {"id": "gri_403", "name": "GRI 403 — Safety", "gri": "403", "hasData": True},
    {"id": "gri_302", "name": "GRI 302 — Energy", "gri": "302", "hasData": True},
    {"id": "gri_305", "name": "GRI 305 — Emissions", "gri": "305", "hasData": True},
]

# SASB RT-CH (Chemicals) report templates — same file formats (CSV/Excel/PDF)
# as GRI, selected via the new `framework` field on ReportRequest. See
# SASB_INTEGRATION_PLAN.md Section 5 (confirmed: file formats unchanged,
# only section content/disclosure labeling differs by framework).
SASB_REPORT_TEMPLATES = [
    {"id": "sasb_ghg_air",       "name": "RT-CH-110/120 — GHG & Air Quality",     "sasb": "RT-CH-110a/120a", "hasData": True},
    {"id": "sasb_energy",        "name": "RT-CH-130 — Energy Management",         "sasb": "RT-CH-130a",      "hasData": True},
    {"id": "sasb_water",         "name": "RT-CH-140 — Water Management",          "sasb": "RT-CH-140a",      "hasData": True},
    {"id": "sasb_waste",         "name": "RT-CH-150 — Hazardous Waste Management","sasb": "RT-CH-150a",      "hasData": True},
    {"id": "sasb_safety",        "name": "RT-CH-320 — Workforce Health & Safety", "sasb": "RT-CH-320a",      "hasData": True},
    {"id": "sasb_process_safety","name": "RT-CH-540 — Process Safety",            "sasb": "RT-CH-540a",      "hasData": False},
]

@app.get("/api/reports/templates")
def get_report_templates(framework: Optional[str] = "GRI"):
    return SASB_REPORT_TEMPLATES if (framework or "GRI").upper() == "SASB" else REPORT_TEMPLATES

class ReportRequest(BaseModel):
    templates: list
    year: Optional[int] = None
    plant: Optional[str] = None
    format: Optional[str] = "csv"
    framework: Optional[str] = "GRI"

# GRI 303 source data now tracks Seawater, Produced Water, third-party water,
# TDS water-quality breakdown, and discharge into groundwater/seawater - the
# only category genuinely absent is a water-quality (TDS) breakdown of
# discharge by destination, which the dataset doesn't separately report.
ML_PER_M3 = 1 / 1000.0

def _ml(value_m3):
    return round(float(value_m3) * ML_PER_M3, 3)

def build_water_gri_sections(df, year=None, plant=None):
    df = safe_filter_plant(df, plant)
    if year:
        df = df[df["ReportingYear"] == year]
    years = sorted(int(y) for y in df["ReportingYear"].unique().tolist())

    def yearly_sum(col):
        return {y: _ml(df[df["ReportingYear"] == y][col].sum()) for y in years}

    # ── GRI 303-1 & 303-2: Management Approach (qualitative narratives) ────────
    ma_303_1 = {
        "type": "narrative",
        "subtitle": "Disclosure 303-1: Interactions with water as a shared resource",
        "row_header": "Disclosure Element",
        "years": [],
        "rows": [
            {"label": "a. How the organization interacts with water",
             "guidance": "[Required — manual input] Describe how and where water is withdrawn, consumed, and discharged, including any water-related impacts the organization has caused, contributed to, or is directly linked to through its business relationships."},
            {"label": "b. Approach used to identify water-related impacts",
             "guidance": "[Required — manual input] Describe the approach used to identify water-related impacts across the value chain, including scope, timeframe, and tools/methodologies applied (e.g., WRI Aqueduct Water Risk Atlas, WWF Water Risk Filter)."},
            {"label": "c. How water-related impacts are addressed",
             "guidance": "[Required — manual input] Describe how the organization works with stakeholders — communities, regulators, and value-chain partners — to steward water as a shared resource."},
            {"label": "d. Process for setting water-related goals and targets",
             "guidance": "[Required — manual input] Describe the process for setting water-related goals and targets, and explain how they relate to public policy and the local context of each water-stressed area."},
        ],
        "notes": [
            "GRI 303-1 is a management approach disclosure requiring qualitative organizational narrative.",
            "This section cannot be auto-generated from dataset values and must be completed by the reporting organization.",
            "Reference: GRI 303: Water and Effluents 2018, Disclosure 303-1.",
        ],
    }

    ma_303_2 = {
        "type": "narrative",
        "subtitle": "Disclosure 303-2: Management of water discharge-related impacts",
        "row_header": "Disclosure Element",
        "years": [],
        "rows": [
            {"label": "a. Minimum standards for effluent quality",
             "guidance": "[Required — manual input] Describe minimum standards set for effluent discharge quality and how they were determined, covering: (i) standards for facilities with no local discharge requirements; (ii) any internally developed water quality standards; (iii) sector-specific standards applied; (iv) whether the receiving waterbody profile was considered."},
        ],
        "notes": [
            "GRI 303-2 is a management approach disclosure requiring qualitative organizational narrative.",
            "This section cannot be auto-generated from dataset values and must be completed by the reporting organization.",
            "Reference: GRI 303: Water and Effluents 2018, Disclosure 303-2.",
        ],
    }

    # ── GRI 303-3: Water Withdrawal ───────────────────────────────────────────
    withdrawal_all_areas = {
        "subtitle": "Total water withdrawal from all areas (GRI 303-3-a)",
        "row_header": "Source",
        "unit": "megaliters",
        "years": years,
        "rows": [
            {"label": "Surface Water", "values": yearly_sum("SurfaceWaterWithdrawn")},
            {"label": "Groundwater", "values": yearly_sum("GroundWater")},
            {"label": "Seawater", "values": yearly_sum("SeawaterWithdrawn")},
            {"label": "Produced Water", "values": yearly_sum("ProducedWaterWithdrawn")},
            {"label": "Third-party Water", "values": yearly_sum("ThirdPartyWaterWithdrawn")},
        ],
        "notes": [
            "Surface Water, Groundwater, Seawater, and Produced Water are sourced directly from metered withdrawal data.",
            "Third-party Water reflects water purchased from municipal suppliers (e.g. DJB, MCGM, CMWSSB).",
            "Per GRI 303 guidance, harvested (collected) rainwater is classified as surface water and is included in the Surface Water withdrawal total.",
            "Values converted from cubic meters (m³) to megaliters (ML) by dividing by 1,000 (1 ML = 1,000 m³).",
        ],
    }

    withdrawal_stress = {
        "subtitle": "Total water withdrawal from areas with water stress (GRI 303-3-b)",
        "row_header": "Source",
        "unit": "megaliters",
        "years": years,
        "rows": [
            {"label": "Surface Water", "values": yearly_sum("SurfaceWaterStress")},
            {"label": "Groundwater", "values": yearly_sum("GroundWaterStress")},
            {"label": "Seawater", "values": yearly_sum("SeawaterStress")},
            {"label": "Produced Water", "values": yearly_sum("ProducedWaterStress")},
            {"label": "Third-party Water", "values": yearly_sum("ThirdPartyWaterStress")},
            {"label": "Total — all sources (stress areas)", "values": yearly_sum("TotalWaterWithdrawnStressArea"), "is_total": True},
        ],
        "notes": [
            "Stress-area volumes are sourced directly from the dataset's per-source water-stress columns.",
            "Water stress classification methodology: [Required — specify the tool used, e.g., WRI Aqueduct Water Risk Atlas at high (40–80%) or extremely high (>80%) baseline water stress, or WWF Water Risk Filter at moderate/high/very high water depletion.]",
            "GRI 303 compilation requirement 2.1 (SHALL): Organizations must use publicly available and credible tools to assess water stress.",
            "GRI 303-3-b-v: For Third-party water in stress areas, the standard additionally requires a breakdown by the source (surface, groundwater, seawater, produced water) from which the third-party supplier draws. This sub-breakdown is not currently available in the dataset.",
        ],
    }

    withdrawal_quality = {
        "subtitle": "Total water withdrawal by water quality (GRI 303-3-c)",
        "row_header": "Quality",
        "unit": "megaliters",
        "years": years,
        "rows": [
            {"label": "Freshwater (≤1,000 mg/L TDS)", "values": yearly_sum("FreshwaterWithdrawnLT1000TDS")},
            {"label": "Other Water (>1,000 mg/L TDS)", "values": yearly_sum("OtherWaterWithdrawnGT1000TDS")},
        ],
        "notes": [
            "TDS threshold: freshwater ≤1,000 mg/L Total Dissolved Solids; other water >1,000 mg/L (GRI 303 Glossary).",
            "GRI 303-3-c requires this TDS quality breakdown per source (Surface Water, Groundwater, etc.) not as an aggregate total.",
            "The current dataset tracks aggregate freshwater/other-water volumes across all sources. Per-source TDS breakdown requires additional dataset columns (e.g., SurfaceWaterFreshwater, GroundwaterFreshwater).",
        ],
    }

    # ── GRI 303-4: Water Discharge ────────────────────────────────────────────
    discharge_all_areas = {
        "subtitle": "Total water discharge to all areas (GRI 303-4-a)",
        "row_header": "Destination",
        "unit": "megaliters",
        "years": years,
        "rows": [
            {"label": "Surface Water", "values": yearly_sum("FreshSurfaceWaterDischarged")},
            {"label": "Groundwater", "values": yearly_sum("GroundwaterDischarged")},
            {"label": "Seawater", "values": yearly_sum("SeawaterBrackishDischarged")},
            {"label": "Third-party Water", "values": yearly_sum("ThirdPartyDischarged")},
        ],
        "notes": [
            "Third-party Water reflects discharge routed to municipal sewage / third-party treatment facilities.",
            "Per GRI 303 definitions, discharge destination 'Seawater' covers marine environments. Brackish/estuarine discharge is reported under this category with this note.",
        ],
    }

    discharge_quality = {
        "subtitle": "Total water discharge by water quality (GRI 303-4-b)",
        "row_header": "Quality",
        "unit": "megaliters",
        "years": years,
        "rows": [
            {"label": "Freshwater (≤1,000 mg/L TDS)", "values": yearly_sum("FreshwaterDischargedLT1000TDS")},
            {"label": "Other Water (>1,000 mg/L TDS)", "values": yearly_sum("OtherWaterDischargedGT1000TDS")},
        ],
        "notes": [
            "TDS threshold: freshwater ≤1,000 mg/L; other water >1,000 mg/L (GRI 303 Glossary).",
            "GRI 303-4-d: Priority substances of concern — defining priority substances, setting discharge limits, and reporting non-compliance incidents — requires qualitative input and effluent quality monitoring data not currently in the dataset.",
        ],
    }

    discharge_stress = {
        "subtitle": "Total water discharge to areas with water stress (GRI 303-4-c)",
        "row_header": "Quality Category",
        "unit": "megaliters",
        "years": years,
        "rows": [
            {"label": "Freshwater (≤1,000 mg/L TDS) — to water stress areas",
             "values": {y: "N/A" for y in years}},
            {"label": "Other Water (>1,000 mg/L TDS) — to water stress areas",
             "values": {y: "N/A" for y in years}},
        ],
        "notes": [
            "GRI 303-4-c requires total water discharge to water stress areas broken down by freshwater/other water quality.",
            "DATA GAP: The current dataset does not include discharge-to-stress-area classification by TDS quality.",
            "Action required: Add 'FreshwaterDischargedStressArea' and 'OtherWaterDischargedStressArea' columns to GRI303_Water_Dataset to complete this disclosure.",
        ],
    }

    # ── GRI 303-5: Water Consumption ──────────────────────────────────────────
    consumption = {
        "subtitle": "Water consumption (GRI 303-5)",
        "row_header": "Metric",
        "unit": "megaliters",
        "years": years,
        "rows": [
            {"label": "Water Consumed (withdrawn − discharged)", "values": yearly_sum("WaterConsumed")},
            {"label": "Water Recycled / Reused", "values": yearly_sum("WaterRecycledReused")},
            {"label": "Water Consumed in Stress Areas", "values": yearly_sum("WaterConsumedStressArea")},
        ],
        "notes": [
            "GRI 303-5 formula: Water Consumption = Total Water Withdrawn − Total Water Discharged.",
            "Water Recycled/Reused is reported as a positive efficiency indicator; it does not affect withdrawal or consumption totals.",
            "GRI 303-5-c Change in water storage: Assessed as not significant — no large water storage facilities are operated. If significant water storage is introduced, this metric must be added to this table.",
        ],
    }

    return [
        {"section_title": "GRI 303 Management Approach", "tables": [ma_303_1, ma_303_2]},
        {"section_title": "GRI 303-3 Water Withdrawal", "tables": [withdrawal_all_areas, withdrawal_stress, withdrawal_quality]},
        {"section_title": "GRI 303-4 Water Discharge", "tables": [discharge_all_areas, discharge_quality, discharge_stress]},
        {"section_title": "GRI 303-5 Water Consumption", "tables": [consumption]},
    ]

def write_gri_sheet(wb, sections, sheet_name="GRI Report"):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(sheet_name[:31])
    header_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    total_fill   = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True, size=12)
    sub_font     = Font(bold=True, italic=True, size=10)
    col_hdr_font = Font(bold=True, size=10)
    bold9        = Font(bold=True, size=9)
    reg9         = Font(size=9)
    note_lbl     = Font(bold=True, italic=True, size=8, color="555555")
    note_f       = Font(italic=True, size=8, color="555555")

    max_years = max((len(t["years"]) for sec in sections for t in sec["tables"]), default=0)

    row = 1
    for section in sections:
        for t_idx, table in enumerate(section["tables"]):
            table_cols = 2 + len(table["years"])

            if t_idx == 0:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(table_cols, 3))
                cell = ws.cell(row=row, column=1, value=section["section_title"])
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                row += 1

            ws.cell(row=row, column=1, value=table["subtitle"]).font = sub_font
            row += 1

            if table.get("type") == "narrative":
                # Management-approach / narrative table — two text columns, no year data
                ws.cell(row=row, column=1, value=table.get("row_header", "Disclosure Element")).font = col_hdr_font
                ws.cell(row=row, column=2, value="Required Content / Placeholder").font = col_hdr_font
                row += 1
                ital9 = Font(italic=True, size=9, color="444444")
                for r in table["rows"]:
                    ws.cell(row=row, column=1, value=r["label"]).font = reg9
                    gc = ws.cell(row=row, column=2, value=r.get("guidance", ""))
                    gc.font = ital9
                    gc.alignment = Alignment(wrap_text=True)
                    row += 1
            else:
                ws.cell(row=row, column=1, value=table["row_header"]).font = col_hdr_font
                ws.cell(row=row, column=2, value="Units").font = col_hdr_font
                for i, yr in enumerate(table["years"]):
                    ws.cell(row=row, column=3 + i, value=yr).font = col_hdr_font
                row += 1

                for r in table["rows"]:
                    is_total = r.get("is_total", False)
                    f = bold9 if is_total else reg9
                    unit = r.get("unit", table.get("unit", ""))
                    lc = ws.cell(row=row, column=1, value=r["label"]); lc.font = f
                    uc = ws.cell(row=row, column=2, value=unit);        uc.font = f
                    if is_total:
                        lc.fill = total_fill; uc.fill = total_fill
                    for i, yr in enumerate(table["years"]):
                        v = r["values"].get(yr)
                        vc = ws.cell(row=row, column=3 + i, value=v if v is not None else 0)
                        vc.font = f
                        if is_total:
                            vc.fill = total_fill
                    row += 1

            if table["notes"]:
                row += 1
                ws.cell(row=row, column=1, value="Notes:").font = note_lbl
                row += 1
                for note in table["notes"]:
                    ws.cell(row=row, column=1, value=note).font = note_f
                    row += 1
            row += 1

    has_narrative = any(t.get("type") == "narrative" for sec in sections for t in sec["tables"])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80 if has_narrative else 12
    for i in range(max_years):
        ws.column_dimensions[get_column_letter(3 + i)].width = 12

# Keep alias for backward compatibility
def write_water_gri_sheet(wb, sections, sheet_name="GRI 303 - Water"):
    write_gri_sheet(wb, sections, sheet_name=sheet_name)

def write_gri_csv(out, sections):
    for section in sections:
        for t_idx, table in enumerate(section["tables"]):
            if t_idx == 0:
                out.write(f"{section['section_title']}\n")
            out.write(f"{table['subtitle']}\n")
            if table.get("type") == "narrative":
                hdr = table.get("row_header", "Disclosure Element")
                out.write(f'"{hdr}","Required Content / Placeholder"\n')
                for r in table["rows"]:
                    label = r["label"].replace('"', '""')
                    guidance = r.get("guidance", "").replace('"', '""')
                    out.write(f'"{label}","{guidance}"\n')
            else:
                header = [table["row_header"], "Units"] + [str(y) for y in table["years"]]
                out.write(",".join(header) + "\n")
                for r in table["rows"]:
                    unit = r.get("unit", table.get("unit", ""))
                    row_vals = [f'"{r["label"]}"', unit] + [
                        "N/A" if r["values"].get(y) == "N/A"
                        else ("0" if r["values"].get(y) is None else str(r["values"].get(y)))
                        for y in table["years"]
                    ]
                    out.write(",".join(row_vals) + "\n")
            if table["notes"]:
                out.write("\nNotes:\n")
                for note in table["notes"]:
                    out.write(f"# {note}\n")
            out.write("\n")

def write_water_gri_csv(out, sections):
    write_gri_csv(out, sections)

def build_energy_gri_sections(df, year=None, plant=None):
    df = safe_filter_plant(df, plant)
    if year:
        df = df[df["ReportingYear"] == year]
    years = sorted(int(y) for y in df["ReportingYear"].unique().tolist())

    def ys(col):
        return {y: round(float(df[df["ReportingYear"] == y][col].sum()), 1) for y in years}
    def ym(col, dp=4):
        numerator_col = INTENSITY_NUMERATOR.get(col)
        result = {}
        for y in years:
            ydf = df[df["ReportingYear"] == y]
            if not len(ydf):
                result[y] = 0
            elif numerator_col:
                result[y] = round(weighted_ratio(ydf, numerator_col), dp)
            else:
                result[y] = round(float(ydf[col].mean()), dp)
        return result

    # ── GRI 302 Management Approach ───────────────────────────────────────────
    ma_302 = {
        "type": "narrative",
        "subtitle": "GRI 302 Energy — Management Approach",
        "row_header": "Disclosure Element",
        "years": [],
        "rows": [
            {"label": "Energy policy and commitments",
             "guidance": "[Required — manual input] Describe the organization's policies and commitments related to energy management, including any renewable energy targets or energy intensity reduction goals."},
            {"label": "Energy management responsibilities",
             "guidance": "[Required — manual input] Identify who holds responsibility for managing energy-related impacts and implementing the organization's energy policies at senior management level."},
            {"label": "Evaluating effectiveness of energy management",
             "guidance": "[Required — manual input] Describe how the organization evaluates the effectiveness of energy management actions, including KPIs monitored, internal audits, and senior management review frequency."},
        ],
        "notes": [
            "GRI 302 management approach must explain how the organization identifies and manages its material energy-related impacts.",
            "This section cannot be auto-generated from dataset values and must be completed by the reporting organization.",
            "Reference: GRI 302: Energy 2016.",
        ],
    }

    consumed_table = {
        "subtitle": "Energy consumed within the organization (GJ)", "row_header": "Source", "unit": "GJ", "years": years,
        "rows": [
            {"label": "Oil (Non-production)",     "values": ys("OilNonProdGJ")},
            {"label": "Natural Gas (Non-prod.)",  "values": ys("NaturalGasNonProdGJ")},
            {"label": "Electricity Consumed",     "values": ys("ElectricityConsumedGJ")},
            {"label": "Steam Consumed",           "values": ys("SteamConsumedGJ")},
            {"label": "Tail Gas Consumed",        "values": ys("TailGasConsumedGJ")},
            {"label": "Compressed Air",           "values": ys("CompressedAirGJ")},
            {"label": "Hot Water Consumed",       "values": ys("HotWaterConsumedGJ")},
            {"label": "Total Energy Consumed",    "values": ys("TotalEnergyConsumedGJ"), "is_total": True},
        ],
        "notes": [
            "Site-specific gross calorific values (GCV) used for natural gas conversions.",
            "Conversion factors applied: 780 MKCal for electricity and 4.187 MKCal to GJ for thermal energy.",
            "GRI 302-1-f: The standards, methodologies, and assumptions used for energy calculations are documented above. The GRI 302: Energy 2016 standard is followed.",
            "GRI 302-1-a: Fuel types listed include non-renewable fuels. Renewable fuels (biofuels, biomass, etc.) should be reported separately if consumed; if none, state 'None'.",
        ],
    }
    sold_table = {
        "subtitle": "Energy sold (GJ)", "row_header": "Source", "unit": "GJ", "years": years,
        "rows": [
            {"label": "Electricity Sold",  "values": ys("ElectricitySoldGJ")},
            {"label": "Tail Gas Sold",     "values": ys("TailGasSoldGJ")},
            {"label": "Hot Water Sold",    "values": ys("HotWaterSoldGJ")},
            {"label": "Steam Sold",        "values": ys("SteamSoldGJ")},
            {"label": "Total Energy Sold", "values": ys("TotalEnergySoldGJ"), "is_total": True},
            {"label": "Net Energy (Consumed − Sold)", "values": ys("NetEnergyGJ")},
        ],
        "notes": [],
    }
    # Renewable Electricity % - computed as sum(renewable)/sum(renewable+non-renewable)
    # per year, not a naive average of the pre-computed per-row RenewableElectricityPct
    # column (RenewableElectricityPct has no INTENSITY_NUMERATOR entry, so ym() would
    # otherwise fall back to .mean() here - the same naive-vs-weighted distinction
    # already fixed for KPI cards and matching the SASB report's calculation).
    renew_vals = ys("ElectricityRenewableGJ")
    nonrenew_vals = ys("ElectricityNonRenewableGJ")
    renew_pct_vals = {
        y: round(renew_vals[y] / (renew_vals[y] + nonrenew_vals[y]) * 100, 1) if (renew_vals[y] + nonrenew_vals[y]) > 0 else 0
        for y in years
    }
    intensity_table = {
        "subtitle": "Renewable electricity and energy intensity (GRI 302-3)", "row_header": "Metric", "unit": "", "years": years,
        "rows": [
            {"label": "Renewable Electricity",    "unit": "GJ",    "values": renew_vals},
            {"label": "Non-Renewable Electricity","unit": "GJ",    "values": nonrenew_vals},
            {"label": "Renewable Electricity %",  "unit": "%",     "values": renew_pct_vals},
            {"label": "Energy Intensity",         "unit": "GJ/t",  "values": ym("EnergyIntensityGJpert", dp=2)},
        ],
        "notes": [
            "Energy intensity = Total energy consumed (GJ) / Total production (tonnes). Denominator: production volume in metric tonnes.",
            "Renewable Electricity % = sum(renewable GJ) / sum(renewable + non-renewable GJ) per year (GRI-weighted ratio, consistent with the dashboard KPI and SASB report calculations).",
        ],
    }
    outside_table = {
        "subtitle": "Upstream and downstream energy (GRI 302-2)", "row_header": "Source", "unit": "", "years": years,
        "rows": [
            {"label": "Upstream Energy",        "unit": "GJ",   "values": ys("UpstreamEnergyGJ")},
            {"label": "Downstream Energy",      "unit": "GJ",   "values": ys("DownstreamEnergyGJ")},
            {"label": "Upstream Intensity",     "unit": "GJ/t", "values": ym("UpstreamEnergyIntensityGJpert", dp=2)},
            {"label": "Downstream Intensity",   "unit": "GJ/t", "values": ym("DownstreamEnergyIntensityGJpert", dp=2)},
        ],
        "notes": [
            "Upstream: raw material extraction, transport, and refining (GHG Protocol Scope 3 Categories 1, 3, 4).",
            "Downstream: transport and distribution of finished products (Scope 3 Category 9).",
        ],
    }

    # ── GRI 302-4: Reduction of energy consumption vs baseline ────────────────
    baseline_y = years[0] if years else None
    reduction_rows = []
    if baseline_y and len(years) > 1:
        bl_consumed = float(df[df["ReportingYear"] == baseline_y]["TotalEnergyConsumedGJ"].sum())
        bl_df = df[df["ReportingYear"] == baseline_y]
        bl_intensity = weighted_ratio(bl_df, "TotalEnergyConsumedGJ")
        reductions, int_reductions = {}, {}
        for y in years:
            ydf = df[df["ReportingYear"] == y]
            reductions[y] = round(bl_consumed - float(ydf["TotalEnergyConsumedGJ"].sum()), 1)
            int_reductions[y] = round(bl_intensity - weighted_ratio(ydf, "TotalEnergyConsumedGJ"), 3)
        reduction_rows = [
            {"label": f"Total Energy Reduction vs {baseline_y} baseline (positive = saved)", "unit": "GJ", "values": reductions},
            {"label": f"Intensity Reduction vs {baseline_y} baseline (positive = improved)", "unit": "GJ/t", "values": int_reductions},
        ]

    reduction_table = {
        "subtitle": f"Energy reduction vs {baseline_y or 'baseline'} (GRI 302-4 indicator)",
        "row_header": "Metric", "unit": "GJ", "years": years,
        "rows": reduction_rows if reduction_rows else [
            {"label": "Insufficient data — at least two reporting years required", "values": {y: "N/A" for y in years}}
        ],
        "notes": [
            f"Baseline year: {baseline_y}. Positive values = energy saved vs baseline; negative = increased consumption.",
            "GRI 302-4 requires organisations to attribute specific reductions to individual conservation and efficiency initiatives.",
            "Initiative-specific attribution cannot be auto-computed from the dataset and must be provided manually by the reporting organisation.",
        ],
    }

    ma_302_5 = {
        "type": "narrative",
        "subtitle": "Disclosure 302-5: Reductions in energy requirements of products and services",
        "row_header": "Disclosure Element",
        "years": [],
        "rows": [
            {"label": "a. Reductions in product/service energy requirements",
             "guidance": "[Required if applicable — manual input] Report reductions in energy requirements of sold products or services compared to prior reporting periods, in joules or multiples. State the methodology and assumptions used to calculate reductions (e.g., lifecycle assessment, engineering estimates)."},
        ],
        "notes": [
            "If no product or service energy reduction programmes exist, state 'Not applicable' with a brief explanation.",
        ],
    }

    return [
        {"section_title": "GRI 302 Energy — Management Approach", "tables": [ma_302]},
        {"section_title": "GRI 302-1 Energy consumption within the organisation",
         "tables": [consumed_table, sold_table, intensity_table]},
        {"section_title": "GRI 302-2 Energy consumption outside the organisation",
         "tables": [outside_table]},
        {"section_title": "GRI 302-4 Reduction of energy consumption",
         "tables": [reduction_table]},
        {"section_title": "GRI 302-5 Reductions in energy requirements of products and services",
         "tables": [ma_302_5]},
    ]


def build_ghg_gri_sections(df, year=None, plant=None):
    df = safe_filter_plant(df, plant)
    if year:
        df = df[df["ReportingYear"] == year]
    years = sorted(int(y) for y in df["ReportingYear"].unique().tolist())
    air_years = [y for y in years if y >= 2022]
    base_year = years[0] if years else None

    def ys(col, yrs=None):
        yrs = yrs or years
        return {y: round(float(df[df["ReportingYear"] == y][col].sum()), 1) for y in yrs}
    def ym(col):
        numerator_col = INTENSITY_NUMERATOR.get(col)
        result = {}
        for y in years:
            ydf = df[df["ReportingYear"] == y]
            if not len(ydf):
                result[y] = 0
            elif numerator_col:
                result[y] = round(weighted_ratio(ydf, numerator_col), 4)
            else:
                result[y] = round(float(ydf[col].mean()), 4)
        return result

    s3 = {}
    for y in years:
        ydf = df[df["ReportingYear"] == y]
        if "Scope3DataAvailable" in df.columns and ydf["Scope3DataAvailable"].max() == 0:
            s3[y] = 0
        else:
            s3[y] = round(float(ydf["Scope3TotaltCO2e"].sum()), 1)

    # ── GRI 305 Management Approach ───────────────────────────────────────────
    ma_305 = {
        "type": "narrative",
        "subtitle": "GRI 305 Emissions — Management Approach",
        "row_header": "Disclosure Element",
        "years": [],
        "rows": [
            {"label": "Emissions reduction policy and targets",
             "guidance": "[Required — manual input] Describe the organization's GHG reduction targets (absolute and/or intensity-based), including any science-based targets (SBTs), net-zero commitments, or renewable energy procurement goals."},
            {"label": "GHG accounting methodology and base year",
             "guidance": "[Required — manual input] State which GHG Protocol standard is applied (e.g., Corporate Accounting and Reporting Standard, Scope 2 Guidance). Identify the base year, the reason for its selection, and the recalculation policy when significant structural changes occur."},
            {"label": "GWP values applied",
             "guidance": "[Required — manual input] Specify the IPCC Assessment Report GWP values used (e.g., AR4, AR5, AR6) and the 100-year time horizon. State whether biogenic CO₂ is included in or excluded from Scope 1."},
            {"label": "Scope 2 — market-based methodology",
             "guidance": "[Required if applicable] Describe whether the organization purchases renewable energy certificates (RECs/GOs) that affect market-based Scope 2 figures. Provide market-based Scope 2 values or explain why they equal the location-based figures."},
            {"label": "Emission reduction initiatives",
             "guidance": "[Required — manual input] Describe programmes in place to reduce direct (Scope 1) and indirect (Scope 2/3) GHG emissions, including energy efficiency upgrades, fuel switching, and carbon offset/removal activities."},
        ],
        "notes": [
            "GRI 305 management approach must explain how the organization manages its material emissions-related impacts.",
            "This section cannot be auto-generated from dataset values and must be completed by the reporting organization.",
            "Reference: GRI 305: Emissions 2016.",
        ],
    }

    scope_table = {
        "subtitle": "GHG emissions by scope (GRI 305-1/305-2/305-3)", "row_header": "Scope", "unit": "tCO₂e", "years": years,
        "rows": [
            {"label": "Scope 1 — Direct Emissions",        "values": ys("Scope1TotaltCO2e")},
            {"label": "Scope 2 — Location-based Indirect", "values": ys("Scope2LocationBasedtCO2e")},
            {"label": "Scope 3 — Other Indirect",          "values": s3},
        ],
        "notes": [
            f"GHG accounting standard: GHG Protocol Corporate Standard (confirm applicable version — manual verification required).",
            f"Base year: {base_year}. This represents the first year for which consistent GHG data is available in this system.",
            "GWP source: IPCC Fifth Assessment Report (AR5) 100-year values. [Confirm if AR6 or another version applies — manual verification required.]",
            "Scope 2 is reported on a location-based methodology per GHG Protocol Scope 2 Guidance.",
            "Scope 2 market-based: Requires supplier-specific emission factors or residual mix factors and is not currently tracked in the dataset. If equivalent to location-based, state that explicitly.",
            "Scope 3: A breakdown by GHG Protocol category (Categories 1–15) is not individually tracked. Where Scope 3 data was unavailable, values are shown as 0.",
            "GRI 305-5 (reduction of GHG emissions vs base year) and GRI 305-6 (ozone-depleting substance emissions) require additional dataset columns or manual input and are not currently reported.",
        ],
    }
    intensity_table = {
        "subtitle": "GHG emissions intensity (GRI 305-4)", "row_header": "Category", "unit": "tCO₂e/t", "years": years,
        "rows": [
            {"label": "Scope 1 Intensity", "values": ym("Scope1IntensitytCO2epert")},
            {"label": "Scope 2 Intensity", "values": ym("Scope2IntensitytCO2epert")},
            {"label": "Scope 3 Intensity", "values": ym("Scope3IntensitytCO2epert")},
        ],
        "notes": [
            "Intensity = tCO₂e / Production tonnes. Denominator: total production volume in metric tonnes.",
            "Scope 1 and Scope 2 intensity denominators use the same production volume for comparability.",
        ],
    }
    air_table = {
        "subtitle": "Significant air emissions (GRI 305-7)", "row_header": "Pollutant", "unit": "tonnes",
        "years": air_years,
        "rows": [
            {"label": "NOx",                              "values": ys("GrossNOxt", air_years)},
            {"label": "SOx",                              "values": ys("GrossSOxt", air_years)},
            {"label": "VOC (Volatile Organic Compounds)", "values": ys("GrossVOCt", air_years)},
            {"label": "PM (Particulate Matter)",          "values": ys("GrossPMt", air_years)},
        ],
        "notes": [
            "Air emission data available from 2022 onwards, when systematic stack monitoring and air quality measurement was established across all sites. Pre-2022 data is not available and is excluded from this disclosure.",
            "GRI 305-7 also requires disclosure of persistent organic pollutants (POPs), hazardous air pollutants (HAPs), and ozone-depleting substances (ODS). These pollutant categories are not currently tracked in the dataset.",
        ],
    }
    return [
        {"section_title": "GRI 305 Emissions — Management Approach", "tables": [ma_305]},
        {"section_title": "GRI 305-1/305-2/305-3 GHG Emissions", "tables": [scope_table, intensity_table]},
        {"section_title": "GRI 305-7 Significant air emissions", "tables": [air_table]},
    ]


def build_waste_gri_sections(df, year=None, plant=None):
    df = safe_filter_plant(df, plant)
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    years = sorted(int(y) for y in df["FiscalReportingPeriod"].unique().tolist())

    def ys(category, haz=None, indicator=None):
        mask = df["WasteCategory"] == category
        if haz:
            mask &= df["HazardousFlag"] == haz
        if indicator:
            mask &= df["IndicatorName"] == indicator
        sub = df[mask]
        return {y: round(float(sub[sub["FiscalReportingPeriod"] == y]["ValueNumber"].sum()), 1) for y in years}

    div_haz = ys("Diverted", "Hazardous")
    div_non = ys("Diverted", "Non-hazardous")
    dis_haz = ys("Disposed", "Hazardous")
    dis_non = ys("Disposed", "Non-hazardous")
    div_tot = ys("Diverted")
    dis_tot = ys("Disposed")
    gen_tot = {y: div_tot[y] + dis_tot[y] for y in years}
    gen_haz = {y: div_haz[y] + dis_haz[y] for y in years}
    gen_non = {y: div_non[y] + dis_non[y] for y in years}

    # ── GRI 306 Management Approach ───────────────────────────────────────────
    ma_306_1 = {
        "type": "narrative",
        "subtitle": "Disclosure 306-1: Waste generation and significant waste-related impacts",
        "row_header": "Disclosure Element",
        "years": [],
        "rows": [
            {"label": "a. Waste generation and significant impacts",
             "guidance": "[Required — manual input] Describe how the organization generates waste and the significant waste-related impacts in its own activities and value chain. Identify the significant waste types and the processes or activities that generate them."},
            {"label": "b. Material topic boundary",
             "guidance": "[Required — manual input] Describe the boundary of the material topic — identify which parts of the value chain generate the significant waste-related impacts (upstream, own operations, or downstream)."},
        ],
        "notes": [
            "GRI 306-1 is a management approach disclosure requiring qualitative organizational narrative.",
            "This section cannot be auto-generated from dataset values and must be completed by the reporting organization.",
            "Reference: GRI 306: Waste 2020, Disclosure 306-1.",
        ],
    }

    ma_306_2 = {
        "type": "narrative",
        "subtitle": "Disclosure 306-2: Management of significant waste-related impacts",
        "row_header": "Disclosure Element",
        "years": [],
        "rows": [
            {"label": "a. Actions to prevent waste generation",
             "guidance": "[Required — manual input] Describe actions taken to prevent waste generation, including waste prevention programmes, product design changes, and upstream collaboration with suppliers."},
            {"label": "b. Actions to manage waste that cannot be prevented",
             "guidance": "[Required — manual input] Describe how the organization manages waste that cannot be prevented, including the waste management hierarchy applied (reuse → recycle → other recovery → dispose)."},
            {"label": "c. Processes for tracking waste transferred to third parties",
             "guidance": "[Required — manual input] Describe the processes used to track and confirm the fate of waste transferred to contractors, including any auditing, chain-of-custody verification, or certifications required."},
        ],
        "notes": [
            "GRI 306-2 is a management approach disclosure requiring qualitative organizational narrative.",
            "This section cannot be auto-generated from dataset values and must be completed by the reporting organization.",
            "Reference: GRI 306: Waste 2020, Disclosure 306-2.",
        ],
    }

    generated_table = {
        "subtitle": "Total waste generated (GRI 306-3)", "row_header": "Category", "unit": "tonnes", "years": years,
        "rows": [
            {"label": "Total Waste Generated",  "values": gen_tot, "is_total": True},
            {"label": "  Hazardous",            "values": gen_haz},
            {"label": "  Non-hazardous",        "values": gen_non},
        ],
        "notes": [
            "Waste generation = Waste diverted from disposal + Waste directed to disposal (GRI 306-3 definition).",
            "GRI 306-3-b: If contextually necessary, report the composition of generated waste by material type. Waste composition data is not currently tracked in the dataset.",
        ],
    }
    diverted_table = {
        "subtitle": "Waste diverted from disposal by recovery operation and hazardous classification (GRI 306-4)",
        "row_header": "Category", "unit": "tonnes", "years": years,
        "rows": [
            {"label": "Reuse — Hazardous",                 "values": ys("Diverted", "Hazardous", "Reuse")},
            {"label": "Reuse — Non-hazardous",             "values": ys("Diverted", "Non-hazardous", "Reuse")},
            {"label": "Recycling — Hazardous",             "values": ys("Diverted", "Hazardous", "Recycling")},
            {"label": "Recycling — Non-hazardous",         "values": ys("Diverted", "Non-hazardous", "Recycling")},
            {"label": "Other Recovery — Hazardous",        "values": ys("Diverted", "Hazardous", "Other")},
            {"label": "Other Recovery — Non-hazardous",    "values": ys("Diverted", "Non-hazardous", "Other")},
            {"label": "Total Diverted — Hazardous",        "values": div_haz, "is_total": True},
            {"label": "Total Diverted — Non-hazardous",    "values": div_non, "is_total": True},
        ],
        "notes": [
            "Values split by Hazardous / Non-hazardous classification per GRI 306-4.",
            "GRI 306-4-b: Organizations must report whether diverted waste is treated onsite (facility operated by the reporting organization) or offsite (third-party facility). Onsite/offsite distinction is not currently available in the dataset.",
            "GRI 306-4-c: For offsite-diverted waste, organizations should confirm the final destination where possible. Contractor destination confirmation data is not currently tracked.",
        ],
    }
    disposed_table = {
        "subtitle": "Waste directed to disposal by disposal operation and hazardous classification (GRI 306-5)",
        "row_header": "Category", "unit": "tonnes", "years": years,
        "rows": [
            {"label": "Incineration — Hazardous",                       "values": ys("Disposed", "Hazardous", "Incineration")},
            {"label": "Incineration — Non-hazardous",                   "values": ys("Disposed", "Non-hazardous", "Incineration")},
            {"label": "Incineration (with recovery) — Hazardous",       "values": ys("Disposed", "Hazardous", "Incineration (with recovery)")},
            {"label": "Incineration (with recovery) — Non-hazardous",   "values": ys("Disposed", "Non-hazardous", "Incineration (with recovery)")},
            {"label": "Landfill — Hazardous",                           "values": ys("Disposed", "Hazardous", "Landfill")},
            {"label": "Landfill — Non-hazardous",                       "values": ys("Disposed", "Non-hazardous", "Landfill")},
            {"label": "Total Disposed — Hazardous",                     "values": dis_haz, "is_total": True},
            {"label": "Total Disposed — Non-hazardous",                 "values": dis_non, "is_total": True},
        ],
        "notes": [
            "Values split by Hazardous / Non-hazardous classification per GRI 306-5.",
            "GRI 306-5-b: Organizations must report whether disposed waste is treated onsite or offsite. Onsite/offsite distinction is not currently available in the dataset.",
            "GRI 306-5-c: Open burning (with or without energy recovery) must be reported separately if it occurs. If no open burning takes place, state 'None'. Open burning data is not currently tracked in the dataset.",
            "GRI 306-5-d: For offsite-disposed waste, organizations should confirm the final disposal destination where possible. Contractor destination confirmation data is not currently tracked.",
        ],
    }
    return [
        {"section_title": "GRI 306 Waste — Management Approach",       "tables": [ma_306_1, ma_306_2]},
        {"section_title": "GRI 306-3 Waste Generated",                 "tables": [generated_table]},
        {"section_title": "GRI 306-4 Waste Diverted From Disposal",    "tables": [diverted_table]},
        {"section_title": "GRI 306-5 Waste Directed to Disposal",      "tables": [disposed_table]},
    ]


def build_safety_gri_sections(df, year=None, plant=None):
    df = safe_filter_plant(df, plant)
    if year:
        df = df[df["FiscalReportingPeriod"] == year]
    years = sorted(int(y) for y in df["FiscalReportingPeriod"].unique().tolist())

    def ysum(wtype, col):
        return {y: round(float(df[(df["FiscalReportingPeriod"] == y) & (df["WorkerType"] == wtype)][col].sum()), 1)
                for y in years}

    def yrate(wtype, count_col):
        result = {}
        for y in years:
            sub = df[(df["FiscalReportingPeriod"] == y) & (df["WorkerType"] == wtype)]
            result[y] = rate_per_basis(sub[count_col].sum(), sub["HoursWorked"].sum())
        return result

    coverage_rows = []
    if "CoveredOHS" in df.columns and "Headcount" in df.columns:
        coverage_rows = [
            {"label": "Headcount — Employees",              "unit": "persons", "values": ysum("Employee",   "Headcount")},
            {"label": "Headcount — Contractors",            "unit": "persons", "values": ysum("Contractor", "Headcount")},
            {"label": "Covered by OHS System — Employees",  "unit": "persons", "values": ysum("Employee",   "CoveredOHS")},
            {"label": "Covered by OHS System — Contractors","unit": "persons", "values": ysum("Contractor", "CoveredOHS")},
        ]

    injury_table = {
        "subtitle": "Work-related injuries by worker type", "row_header": "Category", "unit": "", "years": years,
        "rows": [
            {"label": "Hours Worked — Employees",                     "unit": "hours",  "values": ysum("Employee",   "HoursWorked")},
            {"label": "Hours Worked — Contractors",                   "unit": "hours",  "values": ysum("Contractor", "HoursWorked")},
            {"label": "Recordable Injuries — Employees",              "unit": "count",  "values": ysum("Employee",   "RecordableInjuries")},
            {"label": "Recordable Injuries — Contractors",            "unit": "count",  "values": ysum("Contractor", "RecordableInjuries")},
            {"label": "Lost Time Injuries — Employees",               "unit": "count",  "values": ysum("Employee",   "LostTimeInjuries")},
            {"label": "Lost Time Injuries — Contractors",             "unit": "count",  "values": ysum("Contractor", "LostTimeInjuries")},
            {"label": "High Consequence Injuries — Employees",        "unit": "count",  "values": ysum("Employee",   "HighConsequenceInjuries")},
            {"label": "High Consequence Injuries — Contractors",      "unit": "count",  "values": ysum("Contractor", "HighConsequenceInjuries")},
            {"label": "Fatalities (Injury) — Employees",              "unit": "count",  "values": ysum("Employee",   "FatalitiesInjury")},
            {"label": "Fatalities (Injury) — Contractors",            "unit": "count",  "values": ysum("Contractor", "FatalitiesInjury")},
            {"label": "TRIR — Employees (per 200k hrs)",              "unit": "rate",   "values": yrate("Employee",   "RecordableInjuries")},
            {"label": "TRIR — Contractors (per 200k hrs)",            "unit": "rate",   "values": yrate("Contractor", "RecordableInjuries")},
            {"label": "LTIFR — Employees (per 200k hrs)",             "unit": "rate",   "values": yrate("Employee",   "LostTimeInjuries")},
            {"label": "LTIFR — Contractors (per 200k hrs)",           "unit": "rate",   "values": yrate("Contractor", "LostTimeInjuries")},
        ],
        "notes": [
            "TRIR = Total Recordable Injury Rate per 200,000 hours worked (GRI 403-9 formula).",
            "LTIFR = Lost Time Injury Frequency Rate per 200,000 hours worked (GRI 403-9 formula).",
            "Rate normalization basis: 200,000 hours = GRI 403-9 standard annual full-time equivalent.",
            "Worker inclusions: All employees and directly supervised contractors with recorded hours are included.",
            "Worker exclusions: Independent contractors who control their own work are excluded per GRI 403-9 compilation requirement.",
            "GRI 403-9-c requires identification of the main types of work-related hazard (e.g., working at height, chemical exposure, moving machinery) that caused injuries. Formal hazard-type classification requires additional input from the OHS management system and is not auto-derivable from incident count data alone.",
        ],
    }
    ill_table = {
        "subtitle": "Work-related ill health by worker type (GRI 403-10)", "row_header": "Category", "unit": "count", "years": years,
        "rows": [
            {"label": "Recordable Ill Health Cases — Employees",   "values": ysum("Employee",   "RecordableIllHealth")},
            {"label": "Recordable Ill Health Cases — Contractors", "values": ysum("Contractor", "RecordableIllHealth")},
            {"label": "Fatalities (Ill Health) — Employees",       "values": ysum("Employee",   "FatalitiesIllHealth")},
            {"label": "Fatalities (Ill Health) — Contractors",     "values": ysum("Contractor", "FatalitiesIllHealth")},
        ],
        "notes": [
            "GRI 403-10: Recordable ill health includes cases requiring medical treatment, restricted work, job transfer, or time away from work due to occupational disease.",
            "Worker inclusions/exclusions: Same as GRI 403-9 — all employees and directly supervised contractors are included; independent contractors who control their own work are excluded.",
            "GRI 403-10-c requires identification of the main types of work-related hazard causing ill health (e.g., chemical exposure, ergonomic factors, biological agents). Hazard-to-illness causal linkage requires formal occupational disease case analysis and is not currently tracked in the dataset.",
        ],
    }

    # ── GRI 403 Management Approach (403-1 through 403-7) ─────────────────────
    ma_403 = {
        "type": "narrative",
        "subtitle": "GRI 403 OHS — Management Approach (Disclosures 403-1 through 403-7)",
        "row_header": "Disclosure",
        "years": [],
        "rows": [
            {"label": "403-1: OHS management system",
             "guidance": "[Required — manual input] State whether the OHS management system is based on a legal requirement or a voluntary standard/programme (e.g., ISO 45001). Describe the scope of the system and identify any workers, activities, or workplaces that are excluded from it."},
            {"label": "403-2: Hazard identification, risk assessment, and incident investigation",
             "guidance": "[Required — manual input] Describe the processes for workers to identify and report work-related hazards, the risk assessment methodology applied (e.g., bow-tie analysis, HAZOP), and the incident investigation process including root cause analysis and corrective action tracking."},
            {"label": "403-3: Occupational health services",
             "guidance": "[Required — manual input] Describe how occupational health services facilitate workers' access to health promotion programmes, preventive health assessments, and appropriate medical care. Identify the functions these services perform."},
            {"label": "403-4: Worker participation, consultation, and communication",
             "guidance": "[Required — manual input] Describe the processes for worker participation and consultation in developing, implementing, and evaluating the OHS management system, including how workers are represented (e.g., OHS committees, worker safety representatives)."},
            {"label": "403-5: Worker training on occupational health and safety",
             "guidance": "[Required — manual input] Describe OHS training provided to workers, including specific training for hazardous activities, new worker induction training, and any refresher or recertification requirements."},
            {"label": "403-6: Promotion of worker health",
             "guidance": "[Required — manual input] Describe how the organization facilitates worker access to non-occupational health services (e.g., general health screenings, mental health support) and any voluntary health promotion programmes offered."},
            {"label": "403-7: Prevention and mitigation through business relationships",
             "guidance": "[Required — manual input] Describe measures taken to prevent or mitigate significant OHS impacts directly linked to operations through business relationships (e.g., contractor pre-qualification criteria, supplier OHS requirements, joint OHS audits)."},
        ],
        "notes": [
            "GRI 403-1 through 403-7 are management approach disclosures requiring qualitative organizational narrative.",
            "These disclosures cannot be auto-generated from the OHS incident dataset and must be completed by the reporting organization.",
            "Reference: GRI 403: Occupational Health and Safety 2018.",
        ],
    }

    sections = [
        {"section_title": "GRI 403 OHS — Management Approach", "tables": [ma_403]},
        {"section_title": "GRI 403-9 Work-related Injuries",    "tables": [injury_table]},
        {"section_title": "GRI 403-10 Work-related Ill Health", "tables": [ill_table]},
    ]
    if coverage_rows:
        coverage_table = {
            "subtitle": "Workers covered by an OHS management system (GRI 403-8)",
            "row_header": "Category", "unit": "persons", "years": years,
            "rows": coverage_rows,
            "notes": [
                "GRI 403-8: Number of all employees and workers who are not employees but whose work is controlled by the organization, who are covered by an OHS management system.",
                "If any workers are excluded from OHS system coverage, state the reason and number excluded.",
            ],
        }
        sections.insert(1, {"section_title": "GRI 403-8 OHS System Coverage", "tables": [coverage_table]})
    return sections


def build_report_sections(templates, year, plant):
    sections = []
    if "gri_303" in templates:
        df = load_water_data()
        sections.append({"name": "GRI 303 - Water", "kind": "gri_sections", "data": build_water_gri_sections(df, year=year, plant=plant)})
    if "gri_302" in templates:
        df = load_energy_data()
        sections.append({"name": "GRI 302 - Energy", "kind": "gri_sections", "data": build_energy_gri_sections(df, year=year, plant=plant)})
    if "gri_305" in templates:
        df = load_ghg_data()
        sections.append({"name": "GRI 305 - Emissions", "kind": "gri_sections", "data": build_ghg_gri_sections(df, year=year, plant=plant)})
    if "gri_306" in templates:
        df = load_waste_data()
        sections.append({"name": "GRI 306 - Waste", "kind": "gri_sections", "data": build_waste_gri_sections(df, year=year, plant=plant)})
    if "gri_403" in templates:
        df = load_safety_data()
        sections.append({"name": "GRI 403 - Safety", "kind": "gri_sections", "data": build_safety_gri_sections(df, year=year, plant=plant)})
    return sections

def build_sasb_report_sections(templates, year, plant):
    """SASB equivalent of build_report_sections() - reuses the same
    {section_title, tables} shape produced by sasb_report.py's section
    builders, which write_gri_sheet()/write_gri_csv() already consume
    generically (they were never GRI-specific in implementation)."""
    from sasb_report import (
        _ghg_air_section, _energy_mgmt_section, _water_mgmt_section,
        _hazardous_waste_section, _workforce_safety_section, _process_safety_section,
    )
    sections = []
    if "sasb_ghg_air" in templates:
        sections.append({"name": "RT-CH-110-120 - GHG and Air Quality", "kind": "gri_sections",
                          "data": _ghg_air_section(load_ghg_data(), year=year, plant=plant)})
    if "sasb_energy" in templates:
        sections.append({"name": "RT-CH-130 - Energy Management", "kind": "gri_sections",
                          "data": _energy_mgmt_section(load_energy_data(), year=year, plant=plant)})
    if "sasb_water" in templates:
        sections.append({"name": "RT-CH-140 - Water Management", "kind": "gri_sections",
                          "data": _water_mgmt_section(load_water_data(), year=year, plant=plant)})
    if "sasb_waste" in templates:
        sections.append({"name": "RT-CH-150 - Hazardous Waste Management", "kind": "gri_sections",
                          "data": _hazardous_waste_section(load_waste_data(), year=year, plant=plant)})
    if "sasb_safety" in templates:
        sections.append({"name": "RT-CH-320 - Workforce Health and Safety", "kind": "gri_sections",
                          "data": _workforce_safety_section(load_safety_data(), year=year, plant=plant)})
    if "sasb_process_safety" in templates:
        sections.append({"name": "RT-CH-540 - Process Safety", "kind": "gri_sections",
                          "data": _process_safety_section()})
    return sections


@app.post("/api/reports/generate")
def generate_report(req: ReportRequest):
    if not req.templates:
        raise HTTPException(status_code=400, detail="No valid templates selected")

    is_sasb = (req.framework or "GRI").upper() == "SASB"

    if req.format == "pdf":
        if is_sasb:
            from sasb_report import generate_sasb_pdf
            buf = generate_sasb_pdf(
                templates=req.templates,
                year=req.year,
                plant=req.plant,
                load_water_fn=load_water_data,
                load_waste_fn=load_waste_data,
                load_safety_fn=load_safety_data,
                load_energy_fn=load_energy_data,
                load_ghg_fn=load_ghg_data,
            )
            filename = f"esg_sasb_report_{req.year or 'all'}.pdf"
        else:
            from pdf_report import generate_gri_pdf
            buf = generate_gri_pdf(
                templates=req.templates,
                year=req.year,
                plant=req.plant,
                load_water_fn=load_water_data,
                load_waste_fn=load_waste_data,
                load_safety_fn=load_safety_data,
                load_energy_fn=load_energy_data,
                load_ghg_fn=load_ghg_data,
                build_water_sections_fn=build_water_gri_sections,
                build_energy_sections_fn=build_energy_gri_sections,
                build_ghg_sections_fn=build_ghg_gri_sections,
                build_waste_sections_fn=build_waste_gri_sections,
                build_safety_sections_fn=build_safety_gri_sections,
            )
            filename = f"esg_gri_report_{req.year or 'all'}.pdf"
        content = buf.getvalue()
        storage.save_report(content, filename, framework=("SASB" if is_sasb else "GRI"), format="pdf",
                             templates=req.templates, year=req.year, plant=req.plant)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    sections = build_sasb_report_sections(req.templates, req.year, req.plant) if is_sasb \
        else build_report_sections(req.templates, req.year, req.plant)
    if not sections:
        raise HTTPException(status_code=400, detail="No valid templates selected")

    if req.format == "excel":
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        for sec in sections:
            write_gri_sheet(wb, sec["data"], sheet_name=sec["name"])
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        filename = f"esg_{'sasb' if is_sasb else 'gri'}_report_{req.year or 'all'}.xlsx"
        storage.save_report(content, filename, framework=("SASB" if is_sasb else "GRI"), format="excel",
                             templates=req.templates, year=req.year, plant=req.plant)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        out = io.StringIO()
        for sec in sections:
            write_gri_csv(out, sec["data"])
        content = out.getvalue()
        filename = f"esg_{'sasb' if is_sasb else 'gri'}_report_{req.year or 'all'}.csv"
        storage.save_report(content.encode("utf-8"), filename, framework=("SASB" if is_sasb else "GRI"), format="csv",
                             templates=req.templates, year=req.year, plant=req.plant)
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

# ─── REPORT LIBRARY (persisted) ─────────────────────────────────────────────

_FORMAT_MEDIA = {
    "pdf": "application/pdf",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv",
}

@app.get("/api/reports/library")
def list_report_library():
    reports = storage.list_reports()
    for r in reports:
        r.pop("storage_path", None)
        r["templates"] = r["templates"].split(",") if r["templates"] else []
    return reports

@app.get("/api/reports/library/{report_id}/download")
def download_library_report(report_id: str):
    report = storage.get_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    with open(report["storage_path"], "rb") as f:
        content = f.read()
    return Response(
        content=content,
        media_type=_FORMAT_MEDIA.get(report["format"], "application/octet-stream"),
        headers={"Content-Disposition": f"attachment; filename={report['filename']}"},
    )

@app.delete("/api/reports/library/{report_id}")
def delete_library_report(report_id: str):
    if not storage.delete_report(report_id):
        raise HTTPException(status_code=404, detail="Report not found")
    return {"ok": True}

# ─── ALERT MANAGEMENT ────────────────────────────────────────────────────────
# Centralized alert feed with acknowledgement workflow and configurable thresholds.
# Alert IDs are deterministic: "{domain}:{metric}:{plant}:{year}".

_VALID_ACK_STATUSES = {"acknowledged", "resolved", "ignored"}

@app.get("/api/alerts")
def get_alerts(
    domain: Optional[str] = None,
    severity: Optional[str] = None,
    status: Optional[str] = None,
    year: Optional[int] = None,
    plant: Optional[str] = None,
    region: Optional[str] = None,
):
    """Aggregated alert feed with ack status overlay. Supports filtering by
    domain, severity, and status (open/acknowledged/resolved/ignored)."""
    acks = storage.list_alert_acks()
    domains_to_fetch = [domain] if (domain and domain in _OUTLIER_CONFIG) else list(_OUTLIER_CONFIG.keys())
    all_anomalies = []
    for d in domains_to_fetch:
        t = storage.get_domain_thresholds(d)
        result = _compute_outliers(d, year, plant, region, t["low"], t["medium"], t["high"], limit=None)
        for a in result["anomalies"]:
            alert_id = f"{d}:{a['metric']}:{a['plant']}:{a['year']}"
            ack = acks.get(alert_id)
            all_anomalies.append({
                **a,
                "domain": d,
                "alert_id": alert_id,
                "status": ack["status"] if ack else "open",
                "ack_note": ack["note"] if ack else "",
            })

    if severity and severity != "all":
        all_anomalies = [a for a in all_anomalies if a["severity"] == severity]
    if status and status != "all":
        all_anomalies = [a for a in all_anomalies if a["status"] == status]

    severity_rank = {"high": 0, "medium": 1, "low": 2}
    all_anomalies.sort(key=lambda x: (-x["year"], severity_rank.get(x["severity"], 3), -x["change_pct"]))
    return {"anomalies": all_anomalies, "total": len(all_anomalies)}


class AlertAckRequest(BaseModel):
    alert_id: str
    status: str
    note: Optional[str] = ""

class AlertIdRequest(BaseModel):
    alert_id: str


@app.post("/api/alerts/ack")
def ack_alert(req: AlertAckRequest):
    if req.status not in _VALID_ACK_STATUSES:
        raise HTTPException(status_code=400, detail=f"status must be one of {sorted(_VALID_ACK_STATUSES)}")
    storage.upsert_alert_ack(req.alert_id, req.status, req.note or "")
    return {"ok": True}


@app.post("/api/alerts/unack")
def unack_alert(req: AlertIdRequest):
    storage.delete_alert_ack(req.alert_id)
    return {"ok": True}


@app.get("/api/alerts/config")
def get_alert_config():
    return storage.get_alert_config()


class AlertConfigRequest(BaseModel):
    domain: str
    low: float
    medium: float
    high: float


@app.put("/api/alerts/config")
def update_alert_config(req: AlertConfigRequest):
    if not (0 < req.low < req.medium < req.high <= 200):
        raise HTTPException(status_code=400, detail="Thresholds must satisfy: 0 < low < medium < high <= 200")
    storage.save_domain_thresholds(req.domain, req.low, req.medium, req.high)
    return {"ok": True}


@app.delete("/api/alerts/config/{domain}")
def reset_alert_config(domain: str):
    storage.reset_domain_thresholds(domain)
    return {"ok": True}


# ─── CORRELATION ANALYSIS ────────────────────────────────────────────────────
# Cross-domain Pearson/Spearman correlation endpoints powering the KPI card
# correlation chips. Only GRI datasets are included (water, energy, ghg, waste,
# safety all share the Plant × Year index needed for a clean join).

CORR_METRICS = {
    "water_withdrawn":  {
        "loader": load_water_data, "year_col": "ReportingYear",
        "value_col": "TotalWaterWithdrawn", "scale": 1/1000, "label": "Water Withdrawn (ML)",
    },
    "water_consumed": {
        "loader": load_water_data, "year_col": "ReportingYear",
        "value_col": "WaterConsumed", "scale": 1/1000, "label": "Water Consumed (ML)",
    },
    "energy_consumed": {
        "loader": load_energy_data, "year_col": "ReportingYear",
        "value_col": "TotalEnergyConsumedGJ", "scale": 1, "label": "Energy Consumed (GJ)",
    },
    "scope1_ghg": {
        "loader": load_ghg_data, "year_col": "ReportingYear",
        "value_col": "Scope1TotaltCO2e", "scale": 1, "label": "Scope 1 GHG (tCO₂e)",
    },
    "waste_generated": {
        "loader": load_waste_data, "year_col": "FiscalReportingPeriod",
        "value_col": "__waste_generated__", "scale": 1, "label": "Waste Generated (t)",
    },
    "safety_incidents": {
        "loader": load_safety_data, "year_col": "FiscalReportingPeriod",
        "value_col": "RecordableInjuries", "scale": 1, "label": "Recordable Injuries",
    },
}


def _get_metric_series(metric_id: str, plant: Optional[str] = None):
    """Returns (DataFrame with year/PlantName/value columns, label) or None."""
    cfg = CORR_METRICS.get(metric_id)
    if cfg is None:
        return None
    try:
        df = safe_filter_plant(cfg["loader"](), plant)
        if cfg["value_col"] == "__waste_generated__":
            agg = (df.groupby([cfg["year_col"], "PlantName"])["ValueNumber"]
                     .sum().reset_index()
                     .rename(columns={cfg["year_col"]: "year", "ValueNumber": "value"}))
        else:
            agg = (df.groupby([cfg["year_col"], "PlantName"])[cfg["value_col"]]
                     .sum().reset_index()
                     .rename(columns={cfg["year_col"]: "year", cfg["value_col"]: "value"}))
        agg["value"] = agg["value"] * cfg["scale"]
        return agg, cfg["label"]
    except Exception:
        return None


@app.get("/api/kpi-correlations/{kpi_id}")
def get_kpi_correlations(kpi_id: str, plant: Optional[str] = None):
    """Returns top correlating metrics for a given KPI metric ID.
    Designed for KPI card inline correlation chips — one call per domain page load."""
    try:
        from scipy import stats as _stats
    except ImportError:
        return {"correlations": [], "error": "scipy not installed on backend"}

    base = _get_metric_series(kpi_id, plant if plant and plant != "all" else None)
    if base is None:
        return {"kpi_id": kpi_id, "correlations": []}

    base_df, base_label = base
    correlations = []
    for other_id in CORR_METRICS:
        if other_id == kpi_id:
            continue
        other = _get_metric_series(other_id, plant if plant and plant != "all" else None)
        if other is None:
            continue
        other_df, other_label = other
        merged = base_df.merge(other_df, on=["year", "PlantName"], suffixes=("_a", "_b"))
        if len(merged) < 4:
            continue
        try:
            r, p = _stats.pearsonr(merged["value_a"].values, merged["value_b"].values)
            if np.isnan(r):
                continue
            correlations.append({
                "metric_id": other_id,
                "label": other_label,
                "r": round(float(r), 3),
                "p": round(float(p), 6),
                "n": len(merged),
                "significant": bool(float(p) < 0.05),
                "strength": "strong" if abs(r) >= 0.7 else "moderate" if abs(r) >= 0.4 else "weak",
                "direction": "positive" if r > 0 else "negative",
            })
        except Exception:
            continue

    correlations.sort(key=lambda x: -abs(x["r"]))
    return {"kpi_id": kpi_id, "label": base_label, "correlations": correlations[:4]}


@app.get("/api/correlate")
def get_correlation(
    metric_x: str,
    metric_y: str,
    plant: Optional[str] = None,
    year_from: Optional[int] = None,
    year_to: Optional[int] = None,
):
    """Scatter-plot data + Pearson/Spearman stats for any two correlatable metrics."""
    try:
        from scipy import stats as _stats
    except ImportError:
        raise HTTPException(status_code=500, detail="scipy not installed. Run: pip install scipy")

    p_arg = plant if plant and plant != "all" else None
    res_x = _get_metric_series(metric_x, p_arg)
    res_y = _get_metric_series(metric_y, p_arg)
    if res_x is None or res_y is None:
        raise HTTPException(status_code=400, detail="Unknown metric_x or metric_y")

    df_x, label_x = res_x
    df_y, label_y = res_y
    merged = df_x.merge(df_y, on=["year", "PlantName"], suffixes=("_x", "_y"))
    if year_from:
        merged = merged[merged["year"] >= year_from]
    if year_to:
        merged = merged[merged["year"] <= year_to]

    if len(merged) < 3:
        return {
            "metric_x": label_x, "metric_y": label_y,
            "points": [], "pearson_r": None, "pearson_p": None, "spearman_r": None,
            "regression": None, "n": len(merged),
            "warning": "Not enough data points (minimum 3 required).",
        }

    x_vals = merged["value_x"].values
    y_vals = merged["value_y"].values
    r, p = _stats.pearsonr(x_vals, y_vals)
    sp_r, _ = _stats.spearmanr(x_vals, y_vals)
    slope, intercept, r_val, _, _ = _stats.linregress(x_vals, y_vals)

    abs_r = abs(float(r))
    strength = "strong" if abs_r >= 0.7 else "moderate" if abs_r >= 0.4 else "weak"
    direction = "positive" if r > 0 else "negative"
    sig_text = "statistically significant (p<0.05)" if float(p) < 0.05 else "not statistically significant"
    interpretation = (
        f"{strength.capitalize()} {direction} correlation (r={round(float(r),3)}, p={round(float(p),4)}). "
        f"{sig_text} across {len(merged)} plant-year observations."
    )

    points = [
        {"label": f"{row['PlantName']} / {int(row['year'])}", "x": round(float(row["value_x"]), 3), "y": round(float(row["value_y"]), 3)}
        for _, row in merged.iterrows()
    ]
    return {
        "metric_x": label_x, "metric_y": label_y,
        "points": points,
        "pearson_r": round(float(r), 3),
        "pearson_p": round(float(p), 6),
        "spearman_r": round(float(sp_r), 3),
        "regression": {"slope": round(float(slope), 6), "intercept": round(float(intercept), 3), "r_squared": round(float(r_val**2), 3)},
        "n": len(merged),
        "interpretation": interpretation,
    }


@app.get("/api/correlate/matrix")
def get_correlation_matrix(
    metrics: Optional[str] = None,
    plant: Optional[str] = None,
):
    """Full Pearson correlation matrix across all (or a subset of) correlatable metrics."""
    try:
        from scipy import stats as _stats
    except ImportError:
        raise HTTPException(status_code=500, detail="scipy not installed")

    metric_ids = [m.strip() for m in (metrics or ",".join(CORR_METRICS.keys())).split(",") if m.strip() in CORR_METRICS]
    p_arg = plant if plant and plant != "all" else None

    series_list = []
    for mid in metric_ids:
        result = _get_metric_series(mid, p_arg)
        if result:
            series_list.append((mid, result[0], result[1]))

    if len(series_list) < 2:
        return {"metrics": [s[2] for s in series_list], "metric_ids": [s[0] for s in series_list], "matrix": [], "p_values": [], "n": 0}

    base = series_list[0][1].rename(columns={"value": series_list[0][0]})
    for mid, df, _ in series_list[1:]:
        base = base.merge(df.rename(columns={"value": mid}), on=["year", "PlantName"], how="inner")

    n = len(base)
    matrix, p_matrix = [], []
    for i, (mi, _, _) in enumerate(series_list):
        row_r, row_p = [], []
        for j, (mj, _, _) in enumerate(series_list):
            if i == j:
                row_r.append(1.0); row_p.append(0.0)
            elif n >= 3:
                try:
                    r, p = _stats.pearsonr(base[mi].values, base[mj].values)
                    row_r.append(round(float(r), 3)); row_p.append(round(float(p), 6))
                except Exception:
                    row_r.append(None); row_p.append(None)
            else:
                row_r.append(None); row_p.append(None)
        matrix.append(row_r); p_matrix.append(row_p)

    return {
        "metrics": [s[2] for s in series_list],
        "metric_ids": [s[0] for s in series_list],
        "matrix": matrix,
        "p_values": p_matrix,
        "n": n,
    }


@app.get("/api/health")
def health():
    return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
